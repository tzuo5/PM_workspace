# -*- coding: utf-8 -*-
"""Deterministic Contract / CQP / TA comparison engine.

The engine codifies the business logic maintained in
``backend/config/contract_checker_prompt`` while preserving the response shape
used by the existing browser.  LLM output is narrative-only; all pass/blocker
states are decided here from extracted evidence.
"""

from __future__ import annotations

# CONTRACT_REVIEW_M4367_PATCH_V3: engine

import csv
import os
import re
import unicodedata
from collections import defaultdict
from dataclasses import dataclass
from difflib import SequenceMatcher
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple

from services.contract_llm_review import run_llm_contract_review
from services.contract_review_knowledge import (
    aliases_for_code,
    config_family,
    get_contract_review_knowledge,
    is_commercial_only_config,
    normalize_alias_text,
)
from services.pdf_evidence import (
    ParsedPDF,
    ParsedPage,
    locate_all_evidence,
    locate_evidence,
    match_key,
    normalize_for_match,
    ocr_pdf_pages,
    parse_pdf_with_evidence,
)

CATEGORY_CUSTOMER = "customer_information"
CATEGORY_PRODUCT = "product_information"
CATEGORY_OTHER = "other_information"
CATEGORY_LABELS = {
    CATEGORY_CUSTOMER: "客户信息",
    CATEGORY_PRODUCT: "产品信息",
    CATEGORY_OTHER: "其他信息",
}

SELLER_PREFIX_MAP = {
    "M": "ABB（上海）机器人投资有限公司",
    "K": "ABB机器人（珠海）有限公司",
}
CHECKED_MARKERS = ("☒", "☑", "■", "●", "✓", "√", "[x]", "[X]")
UNCHECKED_MARKERS = ("☐", "□", "○", "[ ]")
AMOUNT_ROUNDING_TOLERANCE = 1.0
EXPECTED_CN_ROBOT_VAT = 0.13

# Generic model recognition deliberately supports new IRB products without a
# Python release.  Small aliases only normalize known marketing-name variants.
MODEL_ALIASES: List[Tuple[str, re.Pattern[str]]] = [
    ("IRB 1200-7/0.7 Gen2", re.compile(r"IRB\s*1200\s*-\s*7\s*/\s*0[.]7\s*Gen\s*2", re.I)),
    ("IRB 1200-7/0.9 LPS", re.compile(r"IRB\s*1200\s*-\s*7\s*/\s*0[.]9\s*(?:LPS|Lite\s*[+＋])", re.I)),
    ("IRB 1100-4/0.58", re.compile(r"IRB\s*1100\s*-\s*4\s*/\s*0[.]58", re.I)),
]
GENERIC_MODEL_PATTERN = re.compile(
    # ASCII-only boundaries intentionally allow Chinese text such as “32台IRB”.
    r"(?<![A-Za-z0-9])IRB\s*\d{3,4}(?:\s*-\s*[\d.]+\s*/\s*[\d.]+)?"
    r"(?:\s*(?:Gen\s*\d+|LPS|Lite\s*[+＋]?))?(?![A-Za-z0-9])",
    re.I,
)


@dataclass
class DocumentSet:
    contract_physical: Optional[ParsedPDF]
    contract_body: Optional[ParsedPDF]
    cqp: Optional[ParsedPDF]
    ta: Optional[ParsedPDF]
    ta_embedded: bool = False


def parse_pdf(filepath: str) -> ParsedPDF:
    return parse_pdf_with_evidence(filepath)


def _slice_pdf(parsed: ParsedPDF, pages: Iterable[ParsedPage], suffix: str = "") -> ParsedPDF:
    selected = list(pages)
    return ParsedPDF(
        filepath=parsed.filepath + suffix,
        pages=selected,
        full_text="\n".join(page.text for page in selected),
    )


def _find_ta_start(parsed: ParsedPDF) -> Optional[int]:
    """Find an embedded TA cover, not a table-of-contents mention."""
    for page in parsed.pages:
        key = match_key(page.text)
        if (
            "technicalagreement" in key
            or "技术协议书" in key
            or "docno.3.02.f03" in key
            or "docno3.02.f03" in key
        ):
            return page.page_num
    return None


def _resolve_documents(
    parsed_by_path: Dict[str, ParsedPDF],
    file_roles: Optional[Dict[str, str]],
) -> DocumentSet:
    contract: Optional[ParsedPDF] = None
    cqp: Optional[ParsedPDF] = None
    ta: Optional[ParsedPDF] = None

    if file_roles:
        for role, path in file_roles.items():
            parsed = parsed_by_path.get(os.path.abspath(path))
            if role == "contract":
                contract = parsed
            elif role == "cqp":
                cqp = parsed
            elif role == "ta":
                ta = parsed

    for parsed in parsed_by_path.values():
        text_key = match_key(parsed.full_text)
        if contract is None and "销售合同" in parsed.full_text and re.search(r"[MK]\s*\d{4}\s*-\s*\d{4}", parsed.full_text):
            contract = parsed
        if cqp is None and re.search(r"CQ\d{7}", parsed.full_text) and ("报价" in parsed.full_text or "quotation" in text_key):
            cqp = parsed
        if ta is None and "技术协议" in parsed.full_text and "销售合同" not in parsed.full_text:
            ta = parsed

    contract_body = contract
    ta_embedded = False
    if contract:
        ta_start = _find_ta_start(contract)
        if ta_start:
            contract_body = _slice_pdf(contract, (page for page in contract.pages if page.page_num < ta_start), "#contract")
            if ta is None:
                ta_pages = [page for page in contract.pages if page.page_num >= ta_start]
                if ta_pages:
                    ta = _slice_pdf(contract, ta_pages, "#ta")
                    ta_embedded = True
    return DocumentSet(contract, contract_body, cqp, ta, ta_embedded)


# ---------------------------------------------------------------------------
# Text normalization and low-level extraction
# ---------------------------------------------------------------------------


def _page_text(parsed: Optional[ParsedPDF], page_num: int) -> str:
    if parsed is None:
        return ""
    page = next((item for item in parsed.pages if item.page_num == page_num), None)
    return page.text if page else ""


def _clean_inline(value: str) -> str:
    value = unicodedata.normalize("NFKC", value or "")
    value = re.sub(r"[\uf000-\uf8ff]", "", value)
    return re.sub(r"\s+", " ", value).strip(" _\t\r\n")


# CONTRACT_REVIEW_EVIDENCE_PATCH_V2
def _normalize_extraction_text(value: str) -> str:
    """Normalize template/OCR fillers without flattening line structure."""
    normalized = unicodedata.normalize("NFKC", value or "")
    normalized = re.sub(r"[＿_]+", " ", normalized)
    normalized = re.sub(r"(?<=\d)\s+(?=(?:%|周|个月|天|元)(?:\b|$))", "", normalized)
    normalized = re.sub(r"(?<=[:：])\s+", " ", normalized)
    return normalized


def _extract_payment_source(text: str) -> str:
    """Select the real Annex 2 body while preserving its original punctuation."""
    source = str(text or "").replace("\r\n", "\n").replace("\r", "\n")
    stop_pattern = (
        r"(?m)^\s*(?:附件\s*[三四五六](?:\s*[^\n]*)?|诚信条款|违约责任|合同解除|解除合同|取消条款|"
        r"质量保证|质保条款|保修条款|交货条款|贸易术语|签字盖章)\s*[:：]?"
    )
    candidates: List[Tuple[int, int, str]] = []
    for match in re.finditer(r"(?m)^\s*附件\s*二(?:\s*[^\n]*)?$", source, re.I):
        section = source[match.start(): match.start() + 10000]
        end = re.search(stop_pattern, section[1:], re.I)
        if end:
            section = section[: end.start() + 1]
        normalized_section = _normalize_extraction_text(section)
        compact_section = match_key(normalized_section)
        score = 0
        score += 8 if "付款条件" in normalized_section else 0
        score += 4 if "付款方式" in normalized_section else 0
        score += 5 if "合同总价" in normalized_section else 0
        score += min(12, 3 * len(re.findall(r"百分之|\d+(?:[.]\d+)?\s*%", normalized_section)))
        score += 3 if "电汇" in normalized_section else 0
        score += 3 if "承兑" in normalized_section else 0
        score += 3 if "发票" in normalized_section else 0
        score += 2 if any(term in compact_section for term in ("用户银行", "账户号码", "账号")) else 0
        candidates.append((score, match.start(), section.strip()))

    if candidates:
        score, _, section = max(candidates, key=lambda item: (item[0], item[1]))
        if score >= 8:
            return section[:8000]

    headings = list(re.finditer(r"(?m)^\s*(?:付款条件|付款方式)\s*(?:[（(][^\n）)]*[）)])?\s*[:：]?", source, re.I))
    if headings:
        start = headings[-1].start()
    else:
        inline = list(re.finditer(r"(?:付款条件|付款方式)\s*[:：]?", source, re.I))
        start = inline[-1].start() if inline else 0
    section = source[start:]
    end = re.search(stop_pattern, section[1:], re.I)
    if end:
        section = section[: end.start() + 1]
    return section[:8000].strip()


def _extract_labeled_money(text: str, labels: Sequence[str], *, not_preceded_by: str = "") -> float:
    """Extract one monetary value after an explicit label.

    This is intentionally independent from the broader `_match_first` helper so
    template underscores and optional currency tokens cannot shift capture groups.
    """
    normalized = _normalize_extraction_text(text)
    label_pattern = "|".join(re.escape(label) for label in labels if label)
    if not label_pattern:
        return 0.0
    prefix_guard = rf"(?<!{re.escape(not_preceded_by)})" if not_preceded_by else ""
    match = re.search(
        rf"{prefix_guard}(?:{label_pattern})(?:为)?\s*[:：]?\s*(?:CNY|RMB|人民币)?\s*[:：]?\s*([\d,]+(?:[.]\d{{1,2}})?)",
        normalized,
        re.I,
    )
    return _parse_money(match.group(1)) if match else 0.0


def _payment_terms_score(terms: Dict[str, Any]) -> Tuple[int, int, int]:
    state_rank = {"MISSING": 0, "EXTRACTION_FAILED": 1, "INCOMPLETE": 2, "EXTRACTED": 3}
    return (
        state_rank.get(str(terms.get("extraction_state", "")), 0),
        len(terms.get("percentages", []) or []),
        len(str(terms.get("raw", "") or "")),
    )


def _payment_ocr_candidate_pages(parsed: Optional[ParsedPDF], max_pages: int = 6) -> List[int]:
    if parsed is None:
        return []
    keyword_pages: List[int] = []
    low_text_pages: List[int] = []
    for page in parsed.pages:
        text = _normalize_extraction_text(page.text)
        if re.search(r"附件\s*二|付款方式|付款条件", text, re.I):
            keyword_pages.append(page.page_num)
        if page.page_num > 1 and len(match_key(text)) < 80:
            low_text_pages.append(page.page_num)
    ordered = keyword_pages + list(reversed(low_text_pages))
    result: List[int] = []
    for page_num in ordered:
        if page_num not in result:
            result.append(page_num)
        if len(result) >= max_pages:
            break
    return sorted(result)


def _extract_contract_delivery(parsed: Optional[ParsedPDF]) -> Dict[str, Any]:
    if parsed is None:
        return {"weeks": 0, "trigger": "", "raw": "", "page": 0}
    trigger_phrases = (
        "合同生效且收到预付款后",
        "合同生效并收到预付款后",
        "合同生效且预付款到账后",
        "自合同生效且收到预付款之日起",
        "自合同生效之日起",
        "合同生效",
        "预付款到账后",
    )
    for page in parsed.pages:
        text = _normalize_extraction_text(page.text)
        for anchor in re.finditer(r"(?:发运时间|发货时间|交货时间|交付周期|交期)\s*[:：]?", text, re.I):
            window = text[anchor.start(): anchor.start() + 520]
            weeks_match = re.search(r"(?<!\d)(\d{1,3})\s*周", window, re.I)
            if not weeks_match:
                continue
            trigger = next((phrase for phrase in trigger_phrases if phrase in window[: weeks_match.end() + 80]), "")
            return {
                "weeks": int(weeks_match.group(1)),
                "trigger": trigger,
                "raw": _clean_inline(window[: weeks_match.end()]),
                "page": page.page_num,
            }
    return {"weeks": 0, "trigger": "", "raw": "", "page": 0}


def _looks_like_configuration_code(code: str, line: str, description: str) -> bool:
    """Reject dates, contract numbers and physical ranges before comparison."""
    try:
        prefix_text, suffix_text = code.split("-", 1)
        prefix, suffix = int(prefix_text), int(suffix_text)
    except (TypeError, ValueError):
        return False

    normalized_line = _normalize_extraction_text(line)
    context = f"{normalized_line} {description}"
    if 1900 <= prefix <= 2099:
        return False
    if re.search(rf"[A-Za-z]\s*{re.escape(code)}", normalized_line):
        return False
    if re.search(r"(?:日期|date|合同编号|contract\s*no)\s*[:：]?\s*" + re.escape(code), context, re.I):
        return False
    physical_words = (
        r"电压|伏特|频率|温度|重量|负载|工作范围|行程|速度|精度|半径|长度|宽度|高度|"
        r"voltage|frequency|temperature|weight|payload|reach|range|speed"
    )
    physical_units = r"V|伏|Hz|kW|W|A|mA|mm|cm|kg|N|Nm|m/s|℃|°C"
    if re.search(rf"(?:{physical_words})[^\n]{{0,35}}{re.escape(code)}", context, re.I):
        return False
    if re.search(rf"{re.escape(code)}\s*(?:{physical_units})\b", context, re.I):
        return False
    if 100 <= prefix <= 1000 and 100 <= suffix <= 1000 and re.search(physical_units, context, re.I):
        return False
    return True


def _coherent_ship_to_tuple(
    contract: Dict[str, Any],
    cqp: Dict[str, Any],
    conclusion: str,
    delivery_location: str,
) -> Tuple[str, str, str]:
    """Return name/address from one coherent source; never mix unrelated fallbacks."""
    candidates: List[Tuple[str, str, str]] = [
        ("合同明确Ship-to", contract.get("ship_to_name", ""), contract.get("ship_to_address", "")),
        ("CQP明确Ship-to", cqp.get("ship_to_name", ""), cqp.get("ship_to_address", "")),
    ]
    if conclusion == "DDP":
        candidates.extend([
            ("合同最终用户", contract.get("end_customer_name", ""), contract.get("end_customer_address", "")),
            ("CQP最终用户", cqp.get("end_user", ""), cqp.get("end_user_address", "")),
        ])
        buyer_name, buyer_address = contract.get("buyer_name", ""), contract.get("buyer_address", "")
        if delivery_location and buyer_address and _address_match(delivery_location, buyer_address):
            candidates.append(("合同买方（交付地点与买方地址一致）", buyer_name, buyer_address))
    elif conclusion == "EXW":
        candidates.append(("合同买方", contract.get("buyer_name", ""), contract.get("buyer_address", "")))

    for source, name, address in candidates:
        clean_name, clean_address = _clean_inline(str(name or "")), _clean_inline(str(address or ""))
        if clean_name and clean_address:
            return clean_name, clean_address, source
    return "", "", ""


def _clean_entity(value: str) -> str:
    return re.sub(r"\s+", "", _clean_inline(value)).strip("_：:")


def _compact(value: str) -> str:
    return re.sub(r"[^a-z0-9\u3400-\u9fff]+", "", normalize_for_match(value))


def _contains_equivalent(left: str, right: str) -> bool:
    lkey, rkey = _compact(left), _compact(right)
    return bool(lkey and rkey and (lkey == rkey or lkey in rkey or rkey in lkey))


def _match_first(text: str, patterns: Sequence[str], flags: int = re.I | re.S) -> str:
    for pattern in patterns:
        match = re.search(pattern, text or "", flags)
        if match:
            return _clean_inline(match.group(1))
    return ""


def _first_line_value(text: str, label: str) -> str:
    stop_labels = (
        "ABB 公司名称", "ABB单位名称", "ABB 单位名称", "地址 / 街道", "地址/街道",
        "联络人", "电话号码", "联系人电子邮箱", "电子邮件", "邮政编码", "城市",
        "国家或地区", "项目名称", "报价编号", "报价单编号", "日期", "报价修订版本",
        "初始编号", "询价日期", "报价日期", "负责人", "页码",
    )
    for raw_line in (text or "").splitlines():
        line = _clean_inline(raw_line)
        pos = line.find(label)
        if pos < 0:
            continue
        value = line[pos + len(label):].lstrip(" :：")
        cut = len(value)
        for stop in stop_labels:
            idx = value.find(stop)
            if idx > 0:
                cut = min(cut, idx)
        if value[:cut].strip():
            return value[:cut].strip()
    return ""


def _parse_money(value: str) -> float:
    try:
        return float(str(value or "").replace(",", "").strip())
    except (TypeError, ValueError):
        return 0.0


def _canonical_model(value: str) -> str:
    normalized = unicodedata.normalize("NFKC", value or "")
    for canonical, pattern in MODEL_ALIASES:
        if pattern.search(normalized):
            return canonical
    match = GENERIC_MODEL_PATTERN.search(normalized)
    if not match:
        return _clean_inline(value)
    raw = _clean_inline(match.group(0))
    parsed = re.match(
        r"IRB\s*(\d{3,4})(?:\s*-\s*([\d.]+)\s*/\s*([\d.]+))?(?:\s*(.*))?$",
        raw,
        re.I,
    )
    if not parsed:
        return raw
    base = f"IRB {parsed.group(1)}"
    if parsed.group(2) and parsed.group(3):
        base += f"-{parsed.group(2)}/{parsed.group(3)}"
    suffix = _clean_inline(parsed.group(4) or "")
    if re.search(r"Lite\s*[+＋]?", suffix, re.I):
        suffix = "LPS"
    elif re.search(r"Gen\s*(\d+)", suffix, re.I):
        suffix = "Gen" + re.search(r"Gen\s*(\d+)", suffix, re.I).group(1)
    elif suffix.upper() == "LPS":
        suffix = "LPS"
    return f"{base} {suffix}".strip()


def _models_in_text(text: str) -> List[str]:
    found: List[str] = []
    for match in GENERIC_MODEL_PATTERN.finditer(unicodedata.normalize("NFKC", text or "")):
        model = _canonical_model(match.group(0))
        if model and model not in found:
            found.append(model)
    return found


def _extract_model_quantities(parsed: Optional[ParsedPDF]) -> Dict[str, int]:
    """Extract quantities only from explicit, model-bound evidence.

    Nearby specification numbers are intentionally ignored.  Ambiguous
    top-confidence values produce no quantity rather than a guessed value.
    """
    if parsed is None:
        return {}
    candidates: Dict[str, List[Tuple[int, int, int, str]]] = defaultdict(list)
    spec_words = re.compile(
        r"负载|工作范围|重复定位|电压|频率|重量|速度|轴数|防护等级|payload|reach|range|"
        r"repeatability|voltage|frequency|weight|speed",
        re.I,
    )
    spec_unit = re.compile(r"^\s*(?:kg|g|mm|cm|m\b|N\b|Nm|kW|W\b|V\b|Hz|℃|°C|轴|级)", re.I)

    for page in parsed.pages:
        page_text = _normalize_extraction_text(page.text)
        lines = [line.strip() for line in page_text.splitlines() if line.strip()]
        page_is_table = bool(re.search(r"供货范围|设备清单|机器人清单|产品清单|数量|qty|quantity", page_text, re.I))
        for line in lines:
            for model_match in GENERIC_MODEL_PATTERN.finditer(line):
                model = _canonical_model(model_match.group(0))
                before, after = line[: model_match.start()], line[model_match.end() :]
                occurrence: List[Tuple[int, int]] = []

                before_match = re.search(r"(\d{1,4})\s*(?:台|套|pcs?|units?)\s*$", before, re.I)
                if before_match:
                    occurrence.append((100, int(before_match.group(1))))

                labelled = re.search(
                    r"(?:数量|qty|quantity)\s*[:：]?\s*(\d{1,4})(?:\s*(?:台|套|pcs?|units?))?",
                    after[:160],
                    re.I,
                )
                if labelled:
                    occurrence.append((100, int(labelled.group(1))))

                unit_after = re.search(r"(?:^|[|｜:：,，;；])\s*(\d{1,4})\s*(?:台|套|pcs?|units?)\b", after[:160], re.I)
                if unit_after:
                    occurrence.append((95, int(unit_after.group(1))))

                if page_is_table and not spec_words.search(line):
                    table_qty = re.match(r"\s*(?:[|｜:：,，;；-]\s*)?(\d{1,4})(?![\d./-])", after)
                    if table_qty and not spec_unit.match(after[table_qty.end() :]):
                        occurrence.append((70, int(table_qty.group(1))))

                valid = [(score, qty) for score, qty in occurrence if 0 < qty < 1000]
                if valid:
                    score, qty = max(valid, key=lambda item: item[0])
                    candidates[model].append((score, qty, page.page_num, _clean_inline(line)))

    result: Dict[str, int] = {}
    for model, values in candidates.items():
        top_score = max(item[0] for item in values)
        top_values = [item[1] for item in values if item[0] == top_score]
        counts = {value: top_values.count(value) for value in set(top_values)}
        ranked = sorted(counts.items(), key=lambda item: (-item[1], item[0]))
        if len(ranked) == 1 or ranked[0][1] > ranked[1][1]:
            result[model] = ranked[0][0]
    return result


def _extract_cqp_products(parsed: Optional[ParsedPDF]) -> List[Dict[str, Any]]:
    if parsed is None:
        return []
    rows: List[Dict[str, Any]] = []
    for page in parsed.pages:
        lines = [unicodedata.normalize("NFKC", line).strip() for line in page.text.splitlines()]
        for index, line in enumerate(lines):
            for model_match in GENERIC_MODEL_PATTERN.finditer(line):
                model = _canonical_model(model_match.group(0))
                tail = line[model_match.end():]
                numbers = re.findall(r"(?<![A-Za-z0-9])([\d,]+[.]\d+|\d+)(?![A-Za-z0-9])", tail)
                if len(numbers) < 3:
                    continue
                qty = int(numbers[0].replace(",", ""))
                if not 0 < qty < 1000:
                    continue
                item_code = ""
                for nearby in lines[index:index + 4]:
                    code_match = re.search(r"(3HAC[\d-]+)", nearby, re.I)
                    if code_match:
                        item_code = code_match.group(1).upper()
                        break
                rows.append({
                    "model": model,
                    "item_code": item_code,
                    "qty": qty,
                    "unit_price": _parse_money(numbers[1]),
                    "line_total": _parse_money(numbers[2]),
                    "page": page.page_num,
                })
    if not rows:
        return [
            {"model": model, "qty": qty, "item_code": "", "unit_price": 0.0, "line_total": 0.0, "page": 0}
            for model, qty in _extract_model_quantities(parsed).items()
        ]
    merged: Dict[str, Dict[str, Any]] = {}
    for row in rows:
        existing = merged.setdefault(row["model"], dict(row))
        if existing is not row and existing != row:
            # setdefault already inserted the first row; only aggregate later rows.
            if existing.get("page") != row.get("page") or existing.get("line_total") != row.get("line_total"):
                existing["qty"] += row["qty"]
                existing["line_total"] = round(existing.get("line_total", 0.0) + row.get("line_total", 0.0), 2)
                if not existing.get("item_code"):
                    existing["item_code"] = row.get("item_code", "")
    return list(merged.values())


def _extract_delivery_schedule(parsed: Optional[ParsedPDF]) -> List[Dict[str, Any]]:
    if parsed is None:
        return []
    schedule: Dict[str, Dict[str, Any]] = {}
    for page in parsed.pages:
        text = _normalize_extraction_text(page.text)
        page_has_delivery = bool(re.search(r"发货时间|交货时间|交付周期|交期", text, re.I))
        for model_match in GENERIC_MODEL_PATTERN.finditer(text):
            model = _canonical_model(model_match.group(0))
            before = text[max(0, model_match.start() - 160) : model_match.start()]
            after = text[model_match.end() : model_match.end() + 650]
            context = after
            qty_match = re.search(r"(\d{1,4})\s*(?:台|套)\s*$", before)
            if not qty_match:
                qty_match = re.search(r"(?:数量\s*[:：]?\s*)?(\d{1,4})\s*(?:台|套)\b", after[:120])
            weeks_match = re.search(
                r"(?:发货时间|交货时间|交付周期|交期)\s*[:：]?[^\n]{0,180}?(\d{1,3})\s*周",
                context,
                re.I,
            )
            if not weeks_match and page_has_delivery:
                weeks_match = re.search(r"(?<!\d)(\d{1,3})\s*周", context[:260], re.I)
            if not weeks_match:
                continue
            entry = {
                "model": model,
                "qty": int(qty_match.group(1)) if qty_match else 0,
                "weeks": int(weeks_match.group(1)),
                "condition": _clean_inline(context[: weeks_match.start()]),
                "page": page.page_num,
            }
            schedule.setdefault(model, entry)
    return list(schedule.values())


def _extract_payment_terms(text: str) -> Dict[str, Any]:
    section_raw = _extract_payment_source(text)
    section = _normalize_extraction_text(section_raw)

    chinese_digits = {"零": 0, "一": 1, "二": 2, "三": 3, "四": 4, "五": 5, "六": 6, "七": 7, "八": 8, "九": 9}

    def chinese_percent(token: str) -> Optional[float]:
        if token == "一百":
            return 100.0
        if "十" in token:
            left, _, right = token.partition("十")
            tens = chinese_digits.get(left, 1) if left else 1
            ones = chinese_digits.get(right, 0) if right else 0
            return float(tens * 10 + ones)
        if token in chinese_digits:
            return float(chinese_digits[token])
        return None

    def percentages(clause: str) -> List[float]:
        arabic = [float(value) for value in re.findall(r"(\d+(?:[.]\d+)?)\s*%", clause)]
        if arabic:
            return arabic
        output: List[float] = []
        for token in re.findall(r"百分之([零一二三四五六七八九十百]+)", clause):
            value = chinese_percent(token)
            if value is not None:
                output.append(value)
        return output

    clauses = [
        _clean_inline(part)
        for part in re.split(r"(?=(?:^|\n)\s*(?:\d+|[一二三四五六七八九十]+)[）).、])", section)
        if _clean_inline(part)
    ]
    if len(clauses) <= 1:
        clauses = [_clean_inline(line) for line in section.splitlines() if _clean_inline(line)]

    payment_hints = ("合同总价", "合同价款", "货款", "预付款", "付款", "支付", "电汇", "承兑", "发货", "交付", "验收", "开票")
    exclusions = ("违约金", "罚金", "罚款", "增值税率", "税率", "利息", "取消费", "赔偿")
    values: List[float] = []
    installments: List[Dict[str, Any]] = []
    for clause in clauses:
        found = percentages(clause)
        if not found or not any(hint in clause for hint in payment_hints):
            continue
        if any(word in clause for word in exclusions) and not any(word in clause for word in ("预付款", "付款", "货款", "合同总价")):
            continue
        values.extend(found)
        for percent in found:
            installments.append({
                "percent": int(percent) if percent.is_integer() else percent,
                "text": clause,
                "trigger": _match_first(clause, [r"(?:在|于)?(.{0,100}?(?:后|前|时|之日))"]),
                "method": "银行承兑汇票" if "承兑" in clause else ("电汇" if "电汇" in clause else ""),
            })

    total = sum(values)
    if not section_raw:
        extraction_state = "MISSING"
    elif not values:
        extraction_state = "EXTRACTION_FAILED"
    elif abs(total - 100) >= 0.01:
        extraction_state = "INCOMPLETE"
    else:
        extraction_state = "EXTRACTED"
    return {
        "installments": installments,
        "percentages": values,
        "raw": section_raw,
        "bank": _match_first(section, [r"用户银行[:：]\s*([^\n]+)", r"(?:结算银行|开户银行)[:：]?\s*([^\n]+)"]),
        "account_name": _match_first(section, [r"(?:账户名称|户名|名称)[:：]\s*([^\n]+)"]),
        "account_number": re.sub(r"\s+", "", _match_first(section, [r"(?:账户号码|帐号|账号)[:：]\s*([\d\s-]+)"])),
        "complete": bool(section_raw and values and abs(total - 100) < 0.01),
        "extraction_state": extraction_state,
    }


def _extract_configurations(parsed: Optional[ParsedPDF], source_type: str) -> List[Dict[str, Any]]:
    """Extract model-bound option rows while rejecting numeric noise."""
    if parsed is None:
        return []
    code_pattern = re.compile(r"(?<![A-Za-z0-9.])(\d{3,4}-\d{1,4})(?![\d.])")
    configs: List[Dict[str, Any]] = []
    for page in parsed.pages:
        lines = [
            _normalize_extraction_text(line).strip()
            for line in page.text.splitlines()
            if _normalize_extraction_text(line).strip()
        ]
        page_models = _models_in_text("\n".join(lines))
        current_model = page_models[0] if len(page_models) == 1 else ""
        for index, line in enumerate(lines):
            model_matches = list(GENERIC_MODEL_PATTERN.finditer(line))
            if model_matches:
                current_model = _canonical_model(model_matches[0].group(0))
            for match in code_pattern.finditer(line):
                if any(model.start() <= match.start() < model.end() for model in model_matches):
                    continue
                code_value = match.group(1)
                description = line[match.end() :].strip(" :：-—|｜")
                if not description and index + 1 < len(lines):
                    next_line = lines[index + 1]
                    if not code_pattern.search(next_line) and not GENERIC_MODEL_PATTERN.search(next_line):
                        description = next_line
                if not _looks_like_configuration_code(code_value, line, description):
                    continue
                if code_value.startswith("3300-"):
                    detected = _canonical_model(description)
                    if detected.startswith("IRB "):
                        current_model = detected
                if not current_model:
                    continue
                configs.append({
                    "model": current_model,
                    "code": code_value,
                    "description": _clean_inline(description),
                    "page": page.page_num,
                    "source": source_type,
                })

    unique: List[Dict[str, Any]] = []
    seen = set()
    for config in configs:
        key = (config.get("model"), config.get("code"), normalize_alias_text(config.get("description", "")))
        if key not in seen:
            unique.append(config)
            seen.add(key)
    return unique


def _model_section_text(parsed: Optional[ParsedPDF], model: str) -> str:
    """Return bounded model sections instead of whole multi-model pages."""
    if parsed is None:
        return ""
    target = _canonical_model(model)
    sections: List[str] = []
    for page in parsed.pages:
        lines = [_normalize_extraction_text(line) for line in page.text.splitlines()]
        markers: List[Tuple[int, str]] = []
        for index, line in enumerate(lines):
            models = _models_in_text(line)
            if models:
                markers.append((index, models[0]))
        for marker_index, (start_index, found_model) in enumerate(markers):
            if found_model != target:
                continue
            next_index = markers[marker_index + 1][0] if marker_index + 1 < len(markers) else len(lines)
            start = max(0, start_index - 2)
            end = min(next_index, start_index + 80)
            sections.append("\n".join(lines[start:end]))
    return "\n".join(sections)[:16000]


def _extract_named_clause(text: str, clause_number: str, max_chars: int = 4000) -> str:
    normalized = unicodedata.normalize("NFKC", text or "").replace("\r\n", "\n").replace("\r", "\n")
    start = re.search(rf"(?m)^\s*{re.escape(clause_number)}\s*[.、]?\s*", normalized)
    if not start:
        start = re.search(rf"(?<!\d){re.escape(clause_number)}\s*[.、]?\s*", normalized)
    if not start:
        return ""
    major, _, minor = clause_number.partition(".")
    tail = normalized[start.start():start.start() + max_chars]
    if minor.isdigit():
        next_minor = str(int(minor) + 1)
        end = re.search(rf"(?m)^\s*{re.escape(major)}[.]\s*{next_minor}\b", tail[1:])
        if end:
            tail = tail[:end.start() + 1]
    return tail.strip()


def _warranty_class_from_text(text: str, periods: Optional[Sequence[Dict[str, Any]]] = None) -> str:
    normalized = match_key(text or "")
    if any(term in normalized for term in ("extendedwarranty", "延长质保", "延保", "延长保修")):
        return "Extended Warranty"
    periods = list(periods or [])
    if any(int(item.get("first_months", 0)) > 18 or int(item.get("second_months", 0)) > 12 for item in periods):
        return "Extended Warranty"
    if any(term in normalized for term in ("standardwarranty", "标准质保", "标准保修", "litewarranty")):
        return "Standard Warranty"
    if periods:
        return "Standard Warranty"
    return "Unknown"


def _extract_warranty_clause(text: str) -> Dict[str, Any]:
    raw = _extract_named_clause(text, "5.2")
    if not raw:
        marker = re.search(r"(?:质量保证|质保条款|保修期)", text or "", re.I)
        raw = (text[marker.start():marker.start() + 3000] if marker else "").strip()
    normalized = unicodedata.normalize("NFKC", raw)
    for word, number in {"十二": 12, "十五": 15, "十八": 18, "二十四": 24, "三十六": 36}.items():
        normalized = normalized.replace(word, str(number))
    periods: List[Dict[str, Any]] = []
    pair_pattern = re.compile(
        r"(\d{1,3})\s*[（(]?\s*\d*\s*[）)]?\s*个?月.{0,180}?(\d{1,3})\s*[（(]?\s*\d*\s*[）)]?\s*个?月",
        re.S,
    )
    for match in pair_pattern.finditer(normalized):
        first, second = int(match.group(1)), int(match.group(2))
        periods.append({
            "period": f"{first}/{second}",
            "first_months": first,
            "second_months": second,
            "context": _clean_inline(normalized[max(0, match.start() - 100):match.end() + 80]),
        })
    for match in re.finditer(r"(?<!\d)(\d{1,3})\s*/\s*(\d{1,3})(?!\d)", normalized):
        period = f"{int(match.group(1))}/{int(match.group(2))}"
        if not any(item["period"] == period for item in periods):
            periods.append({
                "period": period,
                "first_months": int(match.group(1)),
                "second_months": int(match.group(2)),
                "context": _clean_inline(normalized[max(0, match.start() - 100):match.end() + 80]),
            })
    return {
        "raw": raw,
        "periods": periods,
        "primary_period": periods[0]["period"] if periods else "",
        "classification": _warranty_class_from_text(raw, periods),
    }


def _warranty_for_model(warranty: Dict[str, Any], model: str) -> Dict[str, Any]:
    periods = warranty.get("periods", [])
    model_key = _compact(model)
    for item in periods:
        if model_key and model_key in _compact(item.get("context", "")):
            return item
    if "LPS" in model.upper():
        for item in periods:
            context = match_key(item.get("context", ""))
            if "lps" in context or "lite" in context:
                return item
    return periods[0] if periods else {"period": "", "first_months": 0, "second_months": 0, "context": ""}


def _warranty_config_details(configs: Sequence[Dict[str, Any]]) -> Dict[str, Dict[str, str]]:
    result: Dict[str, Dict[str, str]] = {}
    for config in configs:
        code = str(config.get("code", ""))
        description = str(config.get("description", ""))
        if code.startswith("438-") or re.search(r"warranty|质保|保修", description, re.I):
            result[str(config.get("model", ""))] = {
                "code": code,
                "description": description,
                "classification": _warranty_class_from_text(description),
            }
    return result


def _extract_warranty_codes_by_model(configs: Sequence[Dict[str, Any]]) -> Dict[str, str]:
    return {model: data["code"] for model, data in _warranty_config_details(configs).items()}


def _detect_contract_incoterm(page_text: str) -> Dict[str, Any]:
    """Read checkbox state from the transport clause.

    Party wording has varied across prompt examples, so the legal trade basis is
    derived from the unambiguous price term: 到货价 => DDP, 出厂价 => EXW.
    Both option texts may coexist in OCR; only an actual checked marker selects
    an option.
    """
    lines: List[Dict[str, str]] = []
    selected: List[str] = []
    for raw_line in unicodedata.normalize("NFKC", page_text or "").splitlines():
        line = _clean_inline(raw_line)
        if "到货价" not in line and "出厂价" not in line:
            continue
        term = "DDP" if "到货价" in line else "EXW"
        checked = any(marker in line for marker in CHECKED_MARKERS)
        unchecked = any(marker in line for marker in UNCHECKED_MARKERS)
        state = "checked" if checked and not unchecked else ("unchecked" if unchecked and not checked else "unknown")
        lines.append({"term": term, "state": state, "text": line})
        if state == "checked":
            selected.append(term)
    selected_terms = sorted(set(selected))
    return {
        "selected": selected_terms[0] if len(selected_terms) == 1 else "",
        "selected_terms": selected_terms,
        "conflict": len(selected_terms) > 1,
        "lines": lines,
        "candidates": sorted(set(item["term"] for item in lines)),
    }


def _normalize_incoterm(value: str) -> str:
    normalized = match_key(value or "")
    if "ddp" in normalized or "到货" in normalized:
        return "DDP"
    if "exw" in normalized or "出厂" in normalized:
        return "EXW"
    return ""


def _extract_incoterm_named_place(value: str, expected_term: str = "") -> str:
    """Extract the named place from text such as ``DDP Suzhou ...``."""
    text = _clean_inline(value)
    term = expected_term or _normalize_incoterm(text)
    if not term:
        return ""
    match = re.search(rf"\b{term}\b\s*[,，-]?\s*(.+)$", text, re.I)
    if not match:
        return ""
    place = re.split(r"(?:以\s*INCOTERMS|INCOTERMS|为准|特殊指定|\(|（|;|；)", match.group(1), maxsplit=1, flags=re.I)[0]
    return _clean_inline(place).strip(" ,，。")


def _infer_incoterm(contract: Dict[str, Any], cqp: Dict[str, Any]) -> Dict[str, Any]:
    detection = contract.get("incoterm_detection", {})
    selected = detection.get("selected", "")
    cqp_text = cqp.get("delivery_term", "")
    cqp_term = _normalize_incoterm(cqp_text)
    cqp_place = _extract_incoterm_named_place(cqp_text, cqp_term)
    delivery_location = _clean_inline(contract.get("delivery_location", ""))
    explicit_ship_address = _clean_inline(contract.get("ship_to_address", ""))
    destination_evidence = any([
        cqp_place,
        delivery_location,
        explicit_ship_address,
        contract.get("end_customer_address", ""),
        cqp.get("ship_to_address", ""),
        cqp.get("end_user_address", ""),
    ])
    conflict = bool(detection.get("conflict"))

    status, severity, conclusion = "UNDETERMINED", "blocker", ""
    reason = "证据不足，无法确认DDP或EXW。"
    if conflict:
        reason = "合同中DDP与EXW均被识别为已勾选，合同证据互相矛盾。"
    elif selected:
        conclusion = selected
        if cqp_term and cqp_term != selected:
            status, severity = "MISMATCH", "blocker"
            reason = f"合同明确选择{selected}，但CQP为{cqp_term}。"
        elif selected == "DDP" and not destination_evidence:
            status, severity = "UNDETERMINED", "blocker"
            reason = "合同选择DDP，但合同/CQP均未提取到可信目的地。"
        elif selected == "EXW" and delivery_location:
            status, severity = "WARNING", "warning"
            reason = "合同明确选择EXW；交付地点字段虽有内容，但不得据此推翻勾选结果，需确认该字段用途。"
        elif cqp_term:
            status, severity = "PASS", "info"
            reason = f"合同勾选与CQP均为{selected}。"
        else:
            status, severity = "WARNING", "warning"
            reason = f"合同明确选择{selected}，但CQP贸易术语未提取。"
    elif cqp_term == "EXW" and not delivery_location:
        conclusion, status, severity = "EXW", "WARNING", "warning"
        reason = "合同勾选未识别，但交付地点为空且CQP为EXW，可合理推定为EXW。"
    elif cqp_term == "DDP" and destination_evidence:
        conclusion, status, severity = "DDP", "WARNING", "warning"
        reason = "合同勾选未识别，但CQP为DDP且存在可信目的地，可合理推定为DDP。"
    elif cqp_term == "EXW" and delivery_location:
        conclusion, status, severity = "", "MISMATCH", "blocker"
        reason = "合同勾选未识别；CQP为EXW但合同填写交付地点，fallback证据冲突。"
    elif cqp_term == "DDP":
        conclusion, status, severity = "", "UNDETERMINED", "blocker"
        reason = "CQP为DDP，但named place和目的地证据均缺失。"

    if conclusion == "EXW":
        incoterm_2 = cqp_place or _clean_inline(contract.get("seller_origin", "")) or "Shanghai"
    elif conclusion == "DDP":
        incoterm_2 = cqp_place or delivery_location or explicit_ship_address or _clean_inline(contract.get("end_customer_address", ""))
    else:
        incoterm_2 = ""

    ship_to_name, ship_to_address, ship_to_source = _coherent_ship_to_tuple(
        contract, cqp, conclusion, delivery_location
    )
    if conclusion in {"DDP", "EXW"} and status != "MISMATCH" and not (ship_to_name and ship_to_address):
        status, severity = "UNDETERMINED", "blocker"
        reason = reason.rstrip("。") + "；未找到同一来源的完整Ship-to名称和地址。"

    ship_to_rule = (
        "Ship-to名称与地址必须来自同一证据源；DDP优先合同/CQP明确Ship-to，其次完整最终用户；"
        "EXW优先明确Ship-to，其次完整买方信息。禁止跨来源拼接名称和地址。"
    )
    return {
        "conclusion": conclusion or "UNDETERMINED",
        "status": status,
        "severity": severity,
        "reason": reason,
        "contract_selected": selected,
        "cqp_term": cqp_term,
        "cqp_named_place": cqp_place,
        "delivery_location": delivery_location,
        "ship_to_name": ship_to_name,
        "ship_to_address": ship_to_address,
        "ship_to_source": ship_to_source,
        "incoterm_2": incoterm_2,
        "ship_to_rule": ship_to_rule,
    }


def _payment_terms_consistency(contract_terms: Dict[str, Any], cqp_terms: Dict[str, Any]) -> Tuple[str, str, str]:
    if not contract_terms.get("raw"):
        return "UNDETERMINED", "blocker", "合同附件二付款条件未定位，属于来源/抽取缺失，不能判定为条款冲突。"
    if not contract_terms.get("percentages") or not contract_terms.get("complete"):
        return "UNDETERMINED", "blocker", "合同附件二已定位，但付款比例未完整抽取或合计不为100%，需人工确认原文。"
    if not cqp_terms.get("raw") or not cqp_terms.get("percentages"):
        return "WARNING", "warning", "合同付款条件已提取；CQP付款条件未完整提取。BT09仍应原文照抄合同附件二。"
    if sorted(contract_terms.get("percentages", [])) != sorted(cqp_terms.get("percentages", [])):
        return "MISMATCH", "blocker", "合同与CQP均已提取付款比例，且比例明确不一致。"

    def signatures(terms: Dict[str, Any]) -> List[Tuple[Any, Tuple[str, ...], str]]:
        keywords = ("签订合同", "合同生效", "预付款", "发货前", "发货后", "交付", "验收", "开票", "工作日")
        output = []
        for entry in terms.get("installments", []):
            text = str(entry.get("text", ""))
            output.append((entry.get("percent"), tuple(k for k in keywords if k in text), str(entry.get("method", ""))))
        return sorted(output, key=lambda item: (float(item[0] or 0), item[1], item[2]))

    left, right = signatures(contract_terms), signatures(cqp_terms)
    if left and right and left != right:
        return "MISMATCH", "blocker", "合同与CQP付款触发条件或付款方式明确不一致。"
    return "PASS", "info", "合同与CQP付款核心条件一致；BT09付款条件使用合同附件二原文。"


def _address_match(left: str, right: str) -> bool:
    left_key, right_key = _compact(left), _compact(right)
    if not left_key or not right_key:
        return False
    if left_key in right_key or right_key in left_key:
        return True
    left_digits, right_digits = set(re.findall(r"\d+", left_key)), set(re.findall(r"\d+", right_key))
    digit_ok = not left_digits or not right_digits or bool(left_digits & right_digits)
    return digit_ok and SequenceMatcher(None, left_key, right_key).ratio() >= 0.58


def _amount_status(left: float, right: float) -> Tuple[str, str, float]:
    if not left or not right:
        return "UNDETERMINED", "blocker", round(abs((left or 0) - (right or 0)), 2)
    diff = round(abs(left - right), 2)
    if diff == 0:
        return "PASS", "info", diff
    if diff < AMOUNT_ROUNDING_TOLERANCE:
        return "WARNING", "warning", diff
    return "MISMATCH", "blocker", diff


def _alias_match_score(left_values: Sequence[str], right_values: Sequence[str]) -> float:
    """Return a conservative semantic score for bilingual configuration text."""
    left = [normalize_alias_text(value) for value in left_values if normalize_alias_text(value)]
    right = [normalize_alias_text(value) for value in right_values if normalize_alias_text(value)]
    best = 0.0
    for lkey in left:
        for rkey in right:
            if lkey == rkey:
                return 1.0
            # Similar wording with different numeric parameters is a real
            # technical conflict (30m vs 10m, Base 67 vs Base 54, V250 vs V100).
            left_numbers = set(re.findall(r"\d+(?:[.]\d+)?", lkey))
            right_numbers = set(re.findall(r"\d+(?:[.]\d+)?", rkey))
            if left_numbers and right_numbers and left_numbers != right_numbers:
                continue
            if len(lkey) >= 4 and (lkey in rkey or rkey in lkey):
                best = max(best, 0.92)
            else:
                best = max(best, SequenceMatcher(None, lkey, rkey).ratio())
    return best


def _config_match(
    cqp_config: Dict[str, Any],
    ta_configs: Sequence[Dict[str, Any]],
    ta_section: str,
) -> Dict[str, Any]:
    """Match one CQP option against TA by code, maintained aliases, or family.

    A different code from the same option family is not treated as a translation
    match.  It is returned as an explicit conflict so the UI can show both sides
    (for example 3016-3 / 30m versus 3016-2 / 10m).
    """
    model = str(cqp_config.get("model", ""))
    code = str(cqp_config.get("code", ""))
    description = str(cqp_config.get("description", ""))
    same_model = [item for item in ta_configs if item.get("model") == model]

    exact = next((item for item in same_model if item.get("code") == code), None)
    if exact:
        return {"matched": True, "method": "相同代码", "ta": exact, "translation_only": False, "conflict": None}

    cqp_aliases = aliases_for_code(code, description)
    for candidate in same_model:
        candidate_aliases = aliases_for_code(str(candidate.get("code", "")), str(candidate.get("description", "")))
        score = _alias_match_score(cqp_aliases, candidate_aliases)
        if score >= 0.86:
            return {
                "matched": True,
                "method": "中英文/别名映射",
                "ta": candidate,
                "translation_only": True,
                "conflict": None,
                "score": round(score, 3),
            }

    # Some TAs omit the code but include a prose description on the model page.
    section_score = _alias_match_score(cqp_aliases, [ta_section])
    if section_score >= 0.90:
        return {
            "matched": True,
            "method": "TA正文等价描述",
            "ta": None,
            "translation_only": True,
            "conflict": None,
            "score": round(section_score, 3),
        }

    family = config_family(code)
    conflict = None
    if family:
        family_candidates = [item for item in same_model if config_family(str(item.get("code", ""))) == family]
        if family_candidates:
            conflict = family_candidates[0]
    return {"matched": False, "method": "未匹配", "ta": None, "translation_only": False, "conflict": conflict}


# ---------------------------------------------------------------------------
# Structured document extraction
# ---------------------------------------------------------------------------


def extract_contract(parsed: Optional[ParsedPDF]) -> Dict[str, Any]:
    if parsed is None:
        return {}
    source_text = parsed.full_text or "\n".join(page.text for page in parsed.pages)
    text = _normalize_extraction_text(source_text)
    first_pages = "\n".join(_normalize_extraction_text(page.text) for page in parsed.pages[:3])
    transport_text = "\n".join(_normalize_extraction_text(page.text) for page in parsed.pages[:5])
    signature_text = "\n".join(_normalize_extraction_text(page.text) for page in parsed.pages[-3:])

    contract_match = re.search(r"\b([MK])\s*(\d{4})\s*-\s*(\d{4})\b", text, re.I)
    contract_number = "" if not contract_match else f"{contract_match.group(1).upper()}{contract_match.group(2)}-{contract_match.group(3)}"
    buyer = _match_first(first_pages, [r"买方[:：]\s*(.*?)\s*地址[:：]", r"甲方[（(]买方[）)][:：]\s*([^\n]+)"], re.S)
    seller = _match_first(first_pages, [r"卖方[:：]\s*(.*?)\s*地址[:：]", r"乙方[（(]卖方[）)][:：]\s*([^\n]+)"], re.S)
    buyer_address = _match_first(first_pages, [r"买方[:：].*?地址[:：]\s*(.*?)\s*卖方[:：]", r"买方地址[:：]\s*([^\n]+)"], re.S)
    seller_address = _match_first(first_pages, [r"卖方[:：].*?地址[:：]\s*(.*?)(?:目录|合同编号|$)", r"卖方地址[:：]\s*([^\n]+)"], re.S)
    quantities = _extract_model_quantities(parsed)
    delivery_schedule = _extract_delivery_schedule(parsed)
    contract_delivery = _extract_contract_delivery(parsed)

    untaxed_amount = _extract_labeled_money(
        text,
        ("不含增值税总额", "不含增值税金额", "未税金额", "未税总额"),
    )
    tax_included_amount = _extract_labeled_money(
        text,
        ("合同价格的含增值税总额", "含增值税总额", "含增值税金额", "含税总额", "含税金额"),
        not_preceded_by="不",
    )
    vat_text = _match_first(text, [r"(?:增值税(?:税率)?|税率)\s*[:：]?\s*(\d{1,2}(?:[.]\d+)?)\s*%"])

    delivery_trigger = str(contract_delivery.get("trigger", ""))
    if not delivery_trigger:
        for phrase in ("合同生效且收到预付款后", "合同生效并收到预付款后", "合同生效且预付款到账后", "预付款到账后"):
            if phrase in text:
                delivery_trigger = phrase
                break

    placeholder_text = re.sub(r"\s+", "", signature_text + "\n" + text)
    known_placeholders = {"@@@Chop_ABB", "@@@Chop_Customer", "@@@Sign_ABBPerson", "@@@Sign_CustomerPerson"}

    payment_terms = _extract_payment_terms(text)
    payment_terms.setdefault("ocr_used", False)
    payment_terms.setdefault("ocr_pages", [])
    if payment_terms.get("extraction_state") != "EXTRACTED":
        candidate_pages = _payment_ocr_candidate_pages(parsed)
        ocr_by_page = ocr_pdf_pages(parsed.filepath, candidate_pages)
        if ocr_by_page:
            ocr_text = "\n".join(f"[OCR PAGE {page}]\n{ocr_by_page[page]}" for page in sorted(ocr_by_page))
            ocr_terms = _extract_payment_terms(text + "\n" + ocr_text)
            if _payment_terms_score(ocr_terms) > _payment_terms_score(payment_terms):
                payment_terms = ocr_terms
                payment_terms["ocr_used"] = True
                payment_terms["ocr_pages"] = sorted(ocr_by_page)

    delivery_weeks = int(contract_delivery.get("weeks", 0) or 0)
    return {
        "contract_number": contract_number,
        "cqp_reference": _match_first(text, [r"(?:单价信息|报价单|CQP)[:：]?\s*(CQ\d{7})", r"\b(CQ\d{7})\b"]),
        "buyer_name": _clean_entity(buyer),
        "buyer_address": _clean_entity(buyer_address),
        "seller_name": _clean_entity(seller),
        "seller_address": _clean_entity(seller_address),
        "project_name": _match_first(text, [r"买方基于(.{2,80}?)项目需求", r"项目名称[:：]\s*([^\n]+)"]),
        "end_customer_name": _clean_entity(_match_first(text, [r"(?:最终用户|终端客户)(?:名称)?[:：]\s*([^\n]+)"])),
        "end_customer_address": _clean_entity(_match_first(text, [r"(?:最终用户|终端客户)地址[:：]\s*([^\n]+)"])),
        "installation_location": _clean_entity(_match_first(text, [r"(?:设备)?安装地点[:：]\s*([^\n]+)"])),
        "delivery_location": _clean_entity(_match_first(text, [r"(?:交付|交货)地点[:：]\s*([^\n]+)"])),
        "ship_to_name": _clean_entity(_match_first(text, [r"(?:Ship[- ]?to|收货方)(?:名称)?[:：]\s*([^\n]+)"])),
        "ship_to_address": _clean_entity(_match_first(text, [r"(?:Ship[- ]?to|收货方)地址[:：]\s*([^\n]+)"])),
        "seller_origin": _clean_inline(_match_first(transport_text, [r"从([^\n，。]{2,60}?)发出", r"在([^\n，。]{2,60}?)工厂内包装完毕"])),
        "sales_person": _clean_inline(_match_first(text, [r"(?:销售人员|销售负责人|Sales)[:：]\s*([^\n]+)"])),
        "pm": _clean_inline(_match_first(text, [r"(?:项目经理|PM)[:：]\s*([^\n]+)"])),
        "products": [{"model": model, "qty": qty} for model, qty in quantities.items()],
        "total_qty": sum(quantities.values()),
        "delivery_schedule": delivery_schedule,
        "delivery_weeks": delivery_weeks,
        "delivery_raw": contract_delivery.get("raw", ""),
        "delivery_page": int(contract_delivery.get("page", 0) or 0),
        "delivery_trigger": delivery_trigger,
        "split_delivery": bool(re.search(r"允许分批(?:装运|发货)", text)),
        "incoterm_detection": _detect_contract_incoterm(transport_text),
        "untaxed_amount": untaxed_amount,
        "vat_rate": float(vat_text) / 100 if vat_text else 0.0,
        "tax_included_amount": tax_included_amount,
        "payment_terms": payment_terms,
        "warranty": _extract_warranty_clause(text),
        "signature_placeholders": sorted(token for token in known_placeholders if token in placeholder_text),
        "blank_signature_dates": bool(re.search(r"日期[:：]\s+日期[:：]", signature_text) or re.search(r"日期[:：]\s*(?:\n|$)", signature_text)),
        "attachments": {
            "ta": bool(re.search(r"附件一[^\n]{0,80}技术协议", text)),
            "payment": bool(re.search(r"附件二(?:[^\n]{0,80}|\s{0,20})(?:付款方式|付款条件)", text, re.S)),
            "integrity": bool(re.search(r"附件三[^\n]{0,80}诚信条款", text)),
        },
        "file_priority": "本合同优先于附件" if re.search(r"本合同.*?优先于.*?附件|文件优先性", text, re.S) else "",
        "extraction_state": {
            "products": "EXTRACTED" if quantities else "EXTRACTION_FAILED",
            "delivery_schedule": "EXTRACTED" if (delivery_weeks or delivery_schedule) else "EXTRACTION_FAILED",
            "delivery_period": "EXTRACTED" if delivery_weeks else "EXTRACTION_FAILED",
            "untaxed_amount": "EXTRACTED" if untaxed_amount else "EXTRACTION_FAILED",
            "vat_rate": "EXTRACTED" if vat_text else "EXTRACTION_FAILED",
            "tax_included_amount": "EXTRACTED" if tax_included_amount else "EXTRACTION_FAILED",
            "payment_terms": payment_terms.get("extraction_state", "EXTRACTION_FAILED"),
        },
    }


def extract_cqp(parsed: Optional[ParsedPDF]) -> Dict[str, Any]:
    if parsed is None:
        return {}
    text = _normalize_extraction_text(parsed.full_text)
    products = _extract_cqp_products(parsed)
    configs = _extract_configurations(parsed, "cqp")
    number = re.search(r"(?:报价单编号|报价编号)[:：]?\s*(CQ\d{7})", text) or re.search(r"\b(CQ\d{7})\b", text)
    customer = _first_line_value(text, "客户") or _match_first(text, [r"客户(?:名称)?[:：]\s*([^\n]+)"])
    customer_address = _first_line_value(text, "联络地址") or _match_first(text, [r"客户地址[:：]\s*([^\n]+)"])
    seller = _match_first(text, [r"ABB\s*公司名称[:：]?\s*([^\n]+)", r"ABB\s*单位名称[:：]?\s*([^\n]+)"])
    money_values = [_parse_money(value) for value in re.findall(r"CNY\s*[:：]?\s*([\d,]+(?:[.]\d{1,2})?)", text, re.I)]
    untaxed = _parse_money(_match_first(text, [r"(?:未税|不含税)(?:总额|金额)?[:：]?\s*(?:CNY|RMB)?\s*[:：]?\s*([\d,]+(?:[.]\d{1,2})?)"]))
    tax_amount = _parse_money(_match_first(text, [r"(?:增值税额|税额)[:：]?\s*(?:CNY|RMB)?\s*[:：]?\s*([\d,]+(?:[.]\d{1,2})?)"]))
    gross = _parse_money(_match_first(text, [r"(?:含税总额|含税金额)[:：]?\s*(?:CNY|RMB)?\s*[:：]?\s*([\d,]+(?:[.]\d{1,2})?)"]))
    if not untaxed and money_values:
        untaxed = money_values[0]
    if not tax_amount and len(money_values) >= 2:
        tax_amount = money_values[1]
    if not gross and len(money_values) >= 3:
        gross = money_values[2]
    delivery_time = _match_first(text, [r"交货时间[:：]?\s*([^\n]+)", r"交期[:：]?\s*([^\n]+)"])
    delivery_term = _match_first(text, [r"交货条款[:：]?\s*([^\n]+)", r"贸易术语[:：]?\s*([^\n]+)"])
    weeks = re.search(r"(\d+)\s*周", delivery_time)
    vat = re.search(r"增值税(?:率)?\s*[:：]?\s*(\d{1,2}(?:[.]\d+)?)\s*%", text)
    payment_terms = _extract_payment_terms(text)
    return {
        "cqp_number": number.group(1) if number else "",
        "version": _match_first(text, [r"报价修订版本[:：]?\s*([A-Za-z0-9._-]+)", r"版本号[:：]?\s*([A-Za-z0-9._-]+)"]),
        "project_name": _clean_inline(_first_line_value(text, "项目名称") or _match_first(text, [r"项目名称[:：]?\s*([^\n]+)"])),
        "customer_name": _clean_inline(customer),
        "customer_address": _clean_inline(customer_address),
        "end_user": _clean_entity(_match_first(text, [r"(?:最终用户|终端客户)(?:名称)?[:：]\s*([^\n]+)"])),
        "end_user_address": _clean_entity(_match_first(text, [r"(?:最终用户|终端客户)地址[:：]\s*([^\n]+)"])),
        "ship_to_name": _clean_entity(_match_first(text, [r"(?:Ship[- ]?to|收货方)(?:名称)?[:：]\s*([^\n]+)"])),
        "ship_to_address": _clean_entity(_match_first(text, [r"(?:Ship[- ]?to|收货方)地址[:：]\s*([^\n]+)"])),
        "seller_name": _clean_entity(seller),
        "sales_person": _clean_inline(_match_first(text, [r"(?:负责人|销售人员|Sales)[:：]\s*([^\n]+)"])),
        "products": products,
        "total_qty": sum(int(product.get("qty", 0)) for product in products),
        "untaxed_total": untaxed,
        "tax_amount": tax_amount,
        "tax_included_total": gross,
        "vat_rate": float(vat.group(1)) / 100 if vat else 0.0,
        "payment_terms": payment_terms,
        "delivery_time": _clean_inline(delivery_time),
        "delivery_weeks": int(weeks.group(1)) if weeks else 0,
        "delivery_trigger": "预付款到账" if "预付款" in delivery_time else ("合同生效" if "合同生效" in delivery_time else ""),
        "delivery_term": _clean_inline(delivery_term),
        "warranty_terms": _clean_inline(_match_first(text, [r"质量保证[:：]?\s*([^\n]+)", r"质保[:：]?\s*([^\n]+)"])),
        "configurations": configs,
        "warranty_codes_by_model": _extract_warranty_codes_by_model(configs),
        "warranty_details_by_model": _warranty_config_details(configs),
        "extraction_state": {
            "products": "EXTRACTED" if products else "EXTRACTION_FAILED",
            "configurations": "EXTRACTED" if configs else "EXTRACTION_FAILED",
            "payment_terms": payment_terms.get("extraction_state", "EXTRACTION_FAILED"),
        },
    }


def extract_ta(parsed: Optional[ParsedPDF]) -> Dict[str, Any]:
    if parsed is None:
        return {}
    text = _normalize_extraction_text(parsed.full_text)
    configs = _extract_configurations(parsed, "ta")
    quantities = _extract_model_quantities(parsed)
    contract_match = re.search(r"合同编号[:：]?\s*([MK]\s*\d{4}\s*-\s*\d{4})", text, re.I)
    compact_text = re.sub(r"\s+", "", text)
    placeholders = {"@@@Chop_ABB", "@@@Chop_Customer", "@@@Sign_ABBPerson", "@@@Sign_CustomerPerson"}
    return {
        "contract_number": re.sub(r"\s+", "", contract_match.group(1)).upper() if contract_match else "",
        "buyer_name": _clean_entity(_match_first(text, [r"甲方[（(]买方[）)][:：]\s*([^\n]+)", r"买方[:：]\s*([^\n]+)"])),
        "buyer_address": _clean_entity(_match_first(text, [r"买方地址[:：]\s*([^\n]+)"])),
        "seller_name": _clean_entity(_match_first(text, [r"卖方[（(]乙方[）)][:：]\s*([^\n]+)", r"乙方[（(]卖方[）)][:：]\s*([^\n]+)"])),
        "products": [{"model": model, "qty": qty} for model, qty in quantities.items()],
        "total_qty": sum(quantities.values()),
        "configurations": configs,
        "warranty_codes_by_model": _extract_warranty_codes_by_model(configs),
        "warranty_details_by_model": _warranty_config_details(configs),
        "lps_name_in_supply": "LPS" if re.search(r"IRB\s*\d{3,4}[^\n]{0,40}\bLPS\b", text, re.I) else "",
        "lps_name_in_parameters": "Lite+" if re.search(r"IRB\s*\d{3,4}[^\n]{0,40}Lite\s*[+＋]", text, re.I) else "",
        "signature_placeholders": sorted(token for token in placeholders if token in compact_text),
        "blank_signature_dates": bool(re.search(r"日期[:：]\s+日期[:：]", text) or re.search(r"日期[:：]\s*(?:\n|$)", text)),
        "responsibilities": {
            "buyer_integration": bool(re.search(r"买方.*?负责.*?系统集成", text, re.S)),
            "buyer_installation": bool(re.search(r"买方.*?负责.*?(?:卸货|起吊|就位|现场安装)", text, re.S)),
            "seller_not_integration": bool(re.search(r"卖方.*?不承担.*?系统集成", text, re.S)),
        },
        "extraction_state": {
            "products": "EXTRACTED" if quantities else "EXTRACTION_FAILED",
            "configurations": "EXTRACTED" if configs else "EXTRACTION_FAILED",
        },
    }

# ---------------------------------------------------------------------------
# Evidence and review item construction
# ---------------------------------------------------------------------------


def _evidence(
    parsed: Optional[ParsedPDF],
    document_type: str,
    label: str,
    terms: Sequence[str],
    *,
    page_hint: Optional[int] = None,
    quote: str = "",
) -> Dict[str, Any]:
    return locate_evidence(parsed, document_type, label, terms, page_hint=page_hint, quote=quote)


def _evidence_many(parsed: Optional[ParsedPDF], document_type: str, label: str, terms: Sequence[str]) -> List[Dict[str, Any]]:
    return locate_all_evidence(parsed, document_type, label, terms)


def _valid_evidence(entries: Iterable[Dict[str, Any]]) -> List[Dict[str, Any]]:
    output: List[Dict[str, Any]] = []
    seen = set()
    for entry in entries:
        if not entry:
            continue
        key = (entry.get("document_type"), entry.get("page"), entry.get("label"), entry.get("quote"))
        if key not in seen:
            output.append(entry)
            seen.add(key)
    return output


def _item(
    item_id: str,
    category: str,
    title: str,
    status: str,
    severity: str,
    summary: str,
    values: Optional[Dict[str, Any]] = None,
    evidence: Optional[List[Dict[str, Any]]] = None,
    sub_items: Optional[List[Dict[str, Any]]] = None,
    decision_state: str = "",
) -> Dict[str, Any]:
    default_state = {
        "PASS": "MATCH",
        "MISMATCH": "MISMATCH",
        "UNDETERMINED": "EXTRACTION_FAILED",
        "INFO": "NOT_APPLICABLE",
        "WARNING": "REVIEW_REQUIRED",
    }.get(status, "REVIEW_REQUIRED")
    result: Dict[str, Any] = {
        "id": item_id,
        "category": category,
        "category_label": CATEGORY_LABELS[category],
        "title": title,
        "status": status,
        "severity": severity,
        "decision_state": decision_state or default_state,
        "summary": summary,
        "values": values or {},
        "evidence": _valid_evidence(evidence or []),
    }
    if sub_items:
        result["sub_items"] = sub_items
    return result


def _product_map(products: Sequence[Dict[str, Any]]) -> Dict[str, Dict[str, Any]]:
    return {str(product.get("model", "")): product for product in products if product.get("model")}


def build_review_items(
    documents: DocumentSet,
    contract: Dict[str, Any],
    cqp: Dict[str, Any],
    ta: Dict[str, Any],
) -> List[Dict[str, Any]]:
    items: List[Dict[str, Any]] = []
    contract_pdf, cqp_pdf, ta_pdf = documents.contract_physical, documents.cqp, documents.ta
    ta_has_content = bool(documents.ta and (ta.get("products") or ta.get("configurations") or _find_ta_start(documents.ta)))

    source_ok = bool(documents.contract_physical and documents.cqp and ta_has_content)
    items.append(_item(
        "source_completeness", CATEGORY_CUSTOMER, "文件来源完整性",
        "PASS" if source_ok else "MISMATCH", "info" if source_ok else "blocker",
        "合同、CQP和TA均可读取；TA位于合同附件中。" if source_ok and documents.ta_embedded else (
            "合同、CQP和独立TA均可读取。" if source_ok else "合同、CQP或TA来源缺失；这是来源缺失，不是字段不一致。"
        ),
        {"合同": "已找到" if documents.contract_physical else "缺失", "CQP": "已找到" if documents.cqp else "缺失", "TA": "合同内嵌" if documents.ta_embedded else ("独立文件" if ta_has_content else "缺失")},
        [_evidence(contract_pdf, "contract", "合同", [contract.get("contract_number", ""), "销售合同"]), _evidence(cqp_pdf, "cqp", "CQP", [cqp.get("cqp_number", ""), "报价"]), _evidence(ta_pdf, "ta", "TA", ["Technical Agreement", "技术协议书", "技术协议"])],
        decision_state="MATCH" if source_ok else "SOURCE_MISSING",
    ))

    contract_no = contract.get("contract_number", "")
    cqp_ref, cqp_no, ta_no = contract.get("cqp_reference", ""), cqp.get("cqp_number", ""), ta.get("contract_number", "")
    explicit_link_mismatch = bool(contract_no and ta_no and contract_no != ta_no)
    link_missing = not contract_no or (ta_has_content and not ta_no)
    link_status = "MISMATCH" if explicit_link_mismatch else ("UNDETERMINED" if link_missing else "PASS")
    items.append(_item(
        "document_linkage", CATEGORY_CUSTOMER, "合同编号关联", link_status,
        "blocker" if explicit_link_mismatch else ("warning" if link_missing else "info"),
        "合同与TA合同编号明确冲突。" if explicit_link_mismatch else (
            "合同或TA合同编号未提取，需人工确认。" if link_missing else
            "合同与TA合同编号一致；CQP编号和版本仅展示，不参与合同编号判定。"
        ),
        {
            "合同编号": contract_no or "未提取",
            "TA合同编号": ta_no or "未提取",
            "CQP编号（仅展示）": cqp_no or "未提取",
            "CQP版本（仅展示）": cqp.get("version") or "未提取",
            "合同内CQP引用（仅展示）": cqp_ref or "未提取",
        },
        [
            _evidence(contract_pdf, "contract", "合同编号", [contract_no]),
            _evidence(ta_pdf, "ta", "TA合同编号", [ta_no]),
            _evidence(cqp_pdf, "cqp", "CQP编号（仅展示）", [cqp_no]),
        ],
        decision_state="MATCH" if link_status == "PASS" else ("MISMATCH" if link_status == "MISMATCH" else "EXTRACTION_FAILED"),
    ))

    seller = contract.get("seller_name", "")
    expected_seller = SELLER_PREFIX_MAP.get(str(contract_no)[:1], "")
    other_sellers = [value for value in (cqp.get("seller_name", ""), ta.get("seller_name", "")) if value]
    seller_conflict = bool(seller and any(not _contains_equivalent(seller, value) for value in other_sellers))
    prefix_conflict = bool(expected_seller and seller and not _contains_equivalent(expected_seller, seller))
    seller_status = "MISMATCH" if seller_conflict or prefix_conflict else ("UNDETERMINED" if not seller else "PASS")
    items.append(_item(
        "seller_entity", CATEGORY_CUSTOMER, "卖方法定实体", seller_status,
        "blocker" if seller_status == "MISMATCH" else ("warning" if seller_status == "UNDETERMINED" else "info"),
        "卖方主体与其他文件或合同号前缀冲突。" if seller_status == "MISMATCH" else ("卖方主体无法识别，需检查盖章页。" if seller_status == "UNDETERMINED" else "卖方主体与合同号前缀规则一致。"),
        {"合同": seller or "未提取", "CQP": cqp.get("seller_name") or "未提取", "TA": ta.get("seller_name") or "未提取", "合同号预期实体": expected_seller or "无映射"},
        [_evidence(contract_pdf, "contract", "合同卖方", [seller, expected_seller]), _evidence(cqp_pdf, "cqp", "CQP卖方", [cqp.get("seller_name", "")]), _evidence(ta_pdf, "ta", "TA卖方", [ta.get("seller_name", "")])],
    ))

    buyer, customer, ta_buyer = contract.get("buyer_name", ""), cqp.get("customer_name", ""), ta.get("buyer_name", "")
    if not buyer:
        buyer_status, buyer_severity, buyer_summary = "UNDETERMINED", "blocker", "合同买方未提取，无法确认签约主体。"
    elif buyer and ta_buyer and not _contains_equivalent(buyer, ta_buyer):
        buyer_status, buyer_severity, buyer_summary = "MISMATCH", "blocker", "合同与TA买方名称明确不一致。"
    elif customer and not _contains_equivalent(buyer, customer):
        buyer_status, buyer_severity, buyer_summary = "WARNING", "warning", "CQP可能使用客户简称；需由客户主数据确认。"
    elif not customer:
        buyer_status, buyer_severity, buyer_summary = "WARNING", "warning", "CQP客户名称未提取，合同买方作为法律依据。"
    else:
        buyer_status, buyer_severity, buyer_summary = "PASS", "info", "买方主体一致。"
    items.append(_item(
        "buyer_customer_identity", CATEGORY_CUSTOMER, "买方与CQP客户", buyer_status, buyer_severity, buyer_summary,
        {"合同买方": buyer or "未提取", "CQP客户": customer or "未提取", "TA买方": ta_buyer or "未提取"},
        [_evidence(contract_pdf, "contract", "合同买方", [buyer]), _evidence(cqp_pdf, "cqp", "CQP客户", [customer]), _evidence(ta_pdf, "ta", "TA买方", [ta_buyer])],
    ))

    address_ok = _address_match(contract.get("buyer_address", ""), cqp.get("customer_address", ""))
    items.append(_item(
        "customer_address", CATEGORY_CUSTOMER, "客户地址", "PASS" if address_ok else "WARNING", "info" if address_ok else "warning",
        "合同与CQP地址核心信息一致。" if address_ok else "客户地址缺失或仅能部分匹配，需要人工确认。",
        {"合同": contract.get("buyer_address") or "未提取", "CQP": cqp.get("customer_address") or "未提取"},
        [_evidence(contract_pdf, "contract", "合同客户地址", [contract.get("buyer_address", "")]), _evidence(cqp_pdf, "cqp", "CQP客户地址", [cqp.get("customer_address", "")])],
    ))

    end_user, cqp_end_user = contract.get("end_customer_name", ""), cqp.get("end_user", "")
    if end_user and cqp_end_user and not _contains_equivalent(end_user, cqp_end_user):
        end_status, end_severity, end_summary = "MISMATCH", "blocker", "合同与CQP最终用户名称不一致。"
    elif not end_user and not cqp_end_user:
        end_status, end_severity, end_summary = "UNDETERMINED", "warning", "最终用户信息未提取。"
    elif end_user and not cqp_end_user:
        end_status, end_severity, end_summary = "WARNING", "warning", "合同写明最终用户，CQP未列示；以合同为主。"
    else:
        end_status, end_severity, end_summary = "PASS", "info", "最终用户名称一致。"
    items.append(_item(
        "project_end_user", CATEGORY_CUSTOMER, "项目、最终用户与安装地点", end_status, end_severity, end_summary,
        {"合同项目": contract.get("project_name") or "未提取", "CQP项目": cqp.get("project_name") or "未提取", "合同最终用户": end_user or "未提取", "CQP最终用户": cqp_end_user or "未提取", "安装地点": contract.get("installation_location") or "未提取"},
        [_evidence(contract_pdf, "contract", "最终用户", [end_user]), _evidence(cqp_pdf, "cqp", "CQP最终用户", [cqp_end_user])],
    ))

    c_products = _product_map(contract.get("products", []))
    q_products = _product_map(cqp.get("products", []))
    t_products = _product_map(ta.get("products", []))
    contract_package_models = set(c_products) | set(t_products)
    expected_models = set(q_products) if q_products else set(contract_package_models)
    contract_internal_model_conflict = bool(c_products and t_products and set(c_products) != set(t_products))

    if not expected_models and not contract_package_models:
        model_status, model_severity = "UNDETERMINED", "blocker" if source_ok else "warning"
        model_summary = "CQP和合同包均未提取到可信机器人型号。"
    elif contract_internal_model_conflict:
        model_status, model_severity = "MISMATCH", "blocker"
        model_summary = "合同正文与TA中的机器人型号明确不一致。"
    elif not q_products:
        model_status, model_severity = "UNDETERMINED", "blocker" if source_ok else "warning"
        model_summary = "CQP未提取到机器人型号，无法与合同包比较。"
    elif not contract_package_models:
        model_status, model_severity = "UNDETERMINED", "blocker" if source_ok else "warning"
        model_summary = "合同正文和TA均未提取到机器人型号。"
    elif set(q_products) != contract_package_models:
        model_status, model_severity = "MISMATCH", "blocker"
        model_summary = "CQP与合同包（合同正文+TA）的机器人型号明确不一致。"
    else:
        model_status, model_severity = "PASS", "info"
        model_summary = "CQP与合同包（合同正文+TA）的机器人型号一致。"
    items.append(_item(
        "product_models", CATEGORY_PRODUCT, "机器人型号", model_status, model_severity, model_summary,
        {
            "合同正文": "、".join(c_products) or "未提取",
            "TA": "、".join(t_products) or "未提取",
            "合同包": "、".join(sorted(contract_package_models)) or "未提取",
            "CQP": "、".join(q_products) or "未提取",
        },
        _evidence_many(contract_pdf, "contract", "合同正文型号", list(c_products))
        + _evidence_many(ta_pdf, "ta", "TA型号", list(t_products))
        + _evidence_many(cqp_pdf, "cqp", "CQP型号", list(q_products)),
        decision_state="MATCH" if model_status == "PASS" else ("MISMATCH" if model_status == "MISMATCH" else "EXTRACTION_FAILED"),
    ))

    quantity_subitems: List[Dict[str, Any]] = []
    quantity_evidence: List[Dict[str, Any]] = []
    quantity_states: List[str] = []
    for model in sorted(expected_models | contract_package_models):
        contract_qty = int(c_products.get(model, {}).get("qty", 0))
        ta_qty = int(t_products.get(model, {}).get("qty", 0)) if ta_has_content else 0
        cqp_qty = int(q_products.get(model, {}).get("qty", 0))
        package_qty = ta_qty or contract_qty
        internal_conflict = bool(contract_qty and ta_qty and contract_qty != ta_qty)
        if internal_conflict:
            sub_status, sub_state = "MISMATCH", "MISMATCH"
            sub_summary = f"合同正文 {contract_qty} / TA {ta_qty}；合同包内部数量冲突。"
        elif not package_qty or not cqp_qty:
            sub_status, sub_state = "UNDETERMINED", "EXTRACTION_FAILED"
            sub_summary = f"合同包 {package_qty or '未提取'} / CQP {cqp_qty or '未提取'}；证据不完整。"
        elif package_qty != cqp_qty:
            sub_status, sub_state = "MISMATCH", "MISMATCH"
            sub_summary = f"合同包 {package_qty} / CQP {cqp_qty}；数量明确不一致。"
        else:
            sub_status, sub_state = "PASS", "MATCH"
            sub_summary = f"合同包与CQP均为 {package_qty}。"
        quantity_states.append(sub_status)
        evidence = _valid_evidence(
            _evidence_many(contract_pdf, "contract", model + "合同正文数量", [model, str(contract_qty)])
            + _evidence_many(ta_pdf, "ta", model + "TA数量", [model, str(ta_qty)])
            + _evidence_many(cqp_pdf, "cqp", model + "CQP数量", [model, str(cqp_qty)])
        )
        quantity_subitems.append({
            "id": "quantity_" + re.sub(r"[^a-z0-9]+", "_", model.lower()).strip("_"),
            "title": model,
            "status": sub_status,
            "decision_state": sub_state,
            "summary": sub_summary,
            "values": {
                "合同正文": contract_qty or "未提取",
                "TA": ta_qty or "未提取",
                "合同包": package_qty or "未提取",
                "CQP": cqp_qty or "未提取",
            },
            "evidence": evidence,
        })
        quantity_evidence.extend(evidence)
    if "MISMATCH" in quantity_states:
        quantity_status, quantity_severity, quantity_summary = "MISMATCH", "blocker", "至少一个型号在合同包与CQP之间存在明确数量冲突。"
    elif not quantity_states or "UNDETERMINED" in quantity_states:
        quantity_status = "UNDETERMINED"
        quantity_severity = "warning" if model_status != "PASS" else "blocker"
        quantity_summary = "合同包或CQP的型号数量证据不完整。"
    else:
        quantity_status, quantity_severity, quantity_summary = "PASS", "info", "合同包与CQP的各型号数量一致。"

    contract_total = int(contract.get("total_qty", 0) or 0)
    ta_total = int(ta.get("total_qty", 0) or 0)
    cqp_total = int(cqp.get("total_qty", 0) or 0)
    package_total = ta_total or contract_total
    items.append(_item(
        "product_quantities", CATEGORY_PRODUCT, "各型号数量", quantity_status, quantity_severity, quantity_summary,
        {"合同正文总数": contract_total or "未提取", "TA总数": ta_total or "未提取", "合同包总数": package_total or "未提取", "CQP总数": cqp_total or "未提取"},
        quantity_evidence, quantity_subitems,
        decision_state="MATCH" if quantity_status == "PASS" else ("MISMATCH" if quantity_status == "MISMATCH" else "EXTRACTION_FAILED"),
    ))

    if quantity_status != "PASS":
        total_status, total_severity = "UNDETERMINED", "warning"
        total_summary = "总数量依赖各型号数量；上游证据未通过，本项不重复生成阻塞。"
    elif not package_total or not cqp_total:
        total_status, total_severity = "UNDETERMINED", "blocker"
        total_summary = "各型号数量已通过，但合同包或CQP总数量未生成。"
    elif package_total != cqp_total:
        total_status, total_severity = "MISMATCH", "blocker"
        total_summary = "合同包与CQP的机器人总数量明确不一致。"
    else:
        total_status, total_severity = "PASS", "info"
        total_summary = f"合同包与CQP的机器人总数量均为{package_total}。"
    items.append(_item(
        "total_quantity", CATEGORY_PRODUCT, "机器人总数量", total_status, total_severity, total_summary,
        {"合同正文": contract_total or "未提取", "TA": ta_total or "未提取", "合同包": package_total or "未提取", "CQP": cqp_total or "未提取"},
        quantity_evidence,
    ))

    contract_warranty = contract.get("warranty", {})
    cqp_warranty = cqp.get("warranty_details_by_model", {})
    ta_warranty = ta.get("warranty_details_by_model", {})
    warranty_subitems: List[Dict[str, Any]] = []
    warranty_states: List[str] = []
    warranty_notes = False
    for model in sorted(expected_models):
        period_info = _warranty_for_model(contract_warranty, model)
        period = period_info.get("period", "")
        contract_class = _warranty_class_from_text(period_info.get("context", "") or contract_warranty.get("raw", ""), [period_info] if period else [])
        c_detail, t_detail = cqp_warranty.get(model, {}), ta_warranty.get(model, {})
        c_class, t_class = c_detail.get("classification", "Unknown"), t_detail.get("classification", "Unknown")
        codes_differ = bool(c_detail.get("code") and t_detail.get("code") and c_detail.get("code") != t_detail.get("code"))
        if not period or c_class == "Unknown" or t_class == "Unknown":
            sub_status, sub_state = "UNDETERMINED", "EXTRACTION_FAILED"
            sub_summary = "质保来源或分类未完整提取，不能判定为冲突。"
        elif len({contract_class, c_class, t_class}) != 1:
            sub_status, sub_state = "MISMATCH", "MISMATCH"
            sub_summary = "合同、CQP与TA质保分类均已提取且明确冲突。"
        elif codes_differ:
            sub_status, sub_state = "WARNING", "REVIEW_REQUIRED"
            sub_summary = "质保分类一致；CQP/TA代码不同但均映射到同一分类。"
            warranty_notes = True
        else:
            sub_status, sub_state = "PASS", "MATCH"
            sub_summary = "合同5.2与CQP/TA质保分类一致。"
        warranty_states.append(sub_status)
        warranty_subitems.append({
            "id": "warranty_" + re.sub(r"[^a-z0-9]+", "_", model.lower()).strip("_"),
            "title": model,
            "status": sub_status,
            "decision_state": sub_state,
            "summary": sub_summary,
            "values": {"合同质保": period or "未提取", "合同分类": contract_class, "CQP代码": c_detail.get("code") or "未提取", "CQP分类": c_class, "TA代码": t_detail.get("code") or "未提取", "TA分类": t_class},
            "evidence": _valid_evidence([_evidence(contract_pdf, "contract", model + "合同质保", [period, "5.2", "质保"]), _evidence(cqp_pdf, "cqp", model + " CQP质保", [c_detail.get("code", ""), c_detail.get("description", "")]), _evidence(ta_pdf, "ta", model + " TA质保", [t_detail.get("code", ""), t_detail.get("description", "")])]),
        })
    if "MISMATCH" in warranty_states:
        warranty_status, warranty_severity, warranty_summary = "MISMATCH", "blocker", "合同、CQP或TA质保分类存在明确冲突。"
    elif not contract_warranty.get("raw") or not warranty_states or "UNDETERMINED" in warranty_states:
        warranty_status, warranty_severity, warranty_summary = "UNDETERMINED", "blocker", "质保证据未完整提取；这是证据不足，不是已证实冲突。"
    elif warranty_notes:
        warranty_status, warranty_severity, warranty_summary = "WARNING", "warning", "质保分类一致，但配置代码存在可解释差异。"
    else:
        warranty_status, warranty_severity, warranty_summary = "PASS", "info", "合同5.2质保分类与CQP/TA一致。"
    items.append(_item(
        "warranty_by_model", CATEGORY_PRODUCT, "按型号校对质保", warranty_status, warranty_severity, warranty_summary,
        {"合同5.2原文": contract_warranty.get("raw") or "未提取", "合同质保分类": contract_warranty.get("classification") or "Unknown"},
        [entry for sub in warranty_subitems for entry in sub["evidence"]], warranty_subitems,
    ))

    cqp_configs, ta_configs = cqp.get("configurations", []), ta.get("configurations", [])
    config_subitems: List[Dict[str, Any]] = []
    config_evidence: List[Dict[str, Any]] = []
    explicit_conflicts: List[str] = []
    unresolved: List[str] = []
    translation_matches: List[str] = []
    commercial_notes: List[str] = []
    matched_keys = set()
    for config in cqp_configs:
        model, code_value = str(config.get("model", "")), str(config.get("code", ""))
        if not model or not code_value or code_value.startswith("438-"):
            continue
        if is_commercial_only_config(code_value, str(config.get("description", ""))):
            commercial_notes.append(f"{model}:{code_value}")
            continue
        match = _config_match(config, ta_configs, _model_section_text(ta_pdf, model))
        candidate = match.get("ta") or match.get("conflict") or {}
        if match.get("matched"):
            sub_status, sub_state = "PASS", "MATCH"
            if candidate.get("model") and candidate.get("code"):
                matched_keys.add((candidate.get("model"), candidate.get("code")))
            if match.get("translation_only"):
                translation_matches.append(f"{model}:{code_value}")
        elif match.get("conflict"):
            sub_status, sub_state = "MISMATCH", "MISMATCH"
            explicit_conflicts.append(f"{model}:{code_value} ↔ TA {candidate.get('code')}")
        else:
            sub_status, sub_state = "UNDETERMINED", "EXTRACTION_FAILED"
            unresolved.append(f"{model}:{code_value}")

        evidence = _valid_evidence([
            _evidence(cqp_pdf, "cqp", f"{model} {code_value}", [code_value], page_hint=config.get("page")),
            _evidence(ta_pdf, "ta", f"{model} 对应配置", [candidate.get("code", ""), candidate.get("description", ""), *aliases_for_code(code_value, str(config.get("description", "")))]),
        ])
        config_evidence.extend(evidence)
        # Unresolved options are already listed once in the parent card's 待补证 field.
        # Do not render a second child card that repeats the identical issue.
        if sub_status != "UNDETERMINED":
            config_subitems.append({
                "id": "config_" + re.sub(r"[^a-z0-9]+", "_", (model + "_" + code_value).lower()).strip("_"),
                "title": f"{model} · {code_value}",
                "status": sub_status,
                "decision_state": sub_state,
                "summary": (
                    "CQP与TA代码一致。" if match.get("method") == "相同代码" else
                    "中英文/代码别名可映射为同一配置。" if match.get("matched") else
                    f"同类配置明确冲突：CQP {code_value} / TA {candidate.get('code')}。"
                ),
                "values": {"CQP代码": code_value, "CQP描述": config.get("description") or "", "TA代码": candidate.get("code") or "未匹配", "TA描述": candidate.get("description") or "未匹配", "匹配方式": match.get("method")},
                "evidence": evidence,
            })

    cqp_keys = {(item.get("model"), item.get("code")) for item in cqp_configs}
    ta_only = sorted({
        (str(item.get("model", "")), str(item.get("code", "")))
        for item in ta_configs
        if (item.get("model"), item.get("code")) not in cqp_keys
        and (item.get("model"), item.get("code")) not in matched_keys
        and not str(item.get("code", "")).startswith("438-")
        and not is_commercial_only_config(str(item.get("code", "")), str(item.get("description", "")))
    })
    config_unavailable = not cqp_configs or not ta_configs
    if config_unavailable:
        config_status = "UNDETERMINED"
        config_severity = "warning" if model_status != "PASS" else "blocker"
        config_summary = "无法从CQP或TA提取足够配置项；按抽取失败处理，不判定配置冲突。"
    elif explicit_conflicts:
        config_status, config_severity = "MISMATCH", "blocker"
        config_summary = "存在同一配置族的明确参数/代码冲突。"
    elif unresolved:
        config_status, config_severity = "UNDETERMINED", "blocker"
        config_summary = "部分CQP配置未找到可信TA对应项；需要补充证据，不能直接判定冲突。"
    elif ta_only:
        config_status, config_severity = "WARNING", "warning"
        config_summary = "TA存在CQP未列示的附加技术项，需确认范围。"
    else:
        config_status, config_severity = "PASS", "info"
        config_summary = "CQP与TA技术配置一致。"
    items.append(_item(
        "configuration_consistency", CATEGORY_PRODUCT, "CQP与TA技术配置", config_status, config_severity, config_summary,
        {"明确冲突": "、".join(explicit_conflicts) or "无", "待补证": "、".join(unresolved) or "无", "翻译/别名一致": "、".join(translation_matches) or "无", "商务项备注": "、".join(commercial_notes) or "无", "TA附加项": "、".join(f"{m}:{c}" for m, c in ta_only) or "无", "知识库来源": "、".join(get_contract_review_knowledge().source_files) or "内置fallback"},
        config_evidence, config_subitems,
    ))

    naming_warning = bool(ta.get("lps_name_in_supply") and ta.get("lps_name_in_parameters") and ta.get("lps_name_in_supply") != ta.get("lps_name_in_parameters"))
    items.append(_item(
        "lps_lite_naming", CATEGORY_PRODUCT, "LPS / Lite+型号命名", "WARNING" if naming_warning else "PASS", "warning" if naming_warning else "info",
        "TA不同章节使用LPS与Lite+；已按等价别名处理。" if naming_warning else "未发现LPS/Lite+命名冲突。",
        {"供货范围": ta.get("lps_name_in_supply") or "未提取", "技术参数": ta.get("lps_name_in_parameters") or "未提取"},
        [_evidence(ta_pdf, "ta", "LPS/Lite+命名", ["LPS", "Lite+"])],
    ))

    for item_id, title, left_key, right_key in (("untaxed_amount", "未税金额", "untaxed_amount", "untaxed_total"), ("tax_included_amount", "含税金额", "tax_included_amount", "tax_included_total")):
        left, right = float(contract.get(left_key, 0) or 0), float(cqp.get(right_key, 0) or 0)
        status, severity, diff = _amount_status(left, right)
        summary = "合同与CQP金额一致。" if status == "PASS" else (f"合同与CQP相差人民币{diff:.2f}元，小于1元，按舍入差异处理。" if status == "WARNING" else ("合同或CQP金额未提取，按抽取失败处理，不能判定金额冲突。" if status == "UNDETERMINED" else f"合同与CQP金额相差人民币{diff:.2f}元。"))
        items.append(_item(item_id, CATEGORY_OTHER, title, status, severity, summary, {"合同": left or "未提取", "CQP": right or "未提取", "差异": diff}, [_evidence(contract_pdf, "contract", "合同" + title, [str(left)]), _evidence(cqp_pdf, "cqp", "CQP" + title, [str(right)])]))

    contract_vat, cqp_vat = float(contract.get("vat_rate", 0) or 0), float(cqp.get("vat_rate", 0) or 0)
    if not contract_vat or not cqp_vat:
        vat_status, vat_severity, vat_summary = "UNDETERMINED", "blocker", "合同或CQP增值税率未提取；按抽取失败处理。"
    elif abs(contract_vat - cqp_vat) > 0.0001 or abs(contract_vat - EXPECTED_CN_ROBOT_VAT) > 0.0001:
        vat_status, vat_severity, vat_summary = "MISMATCH", "blocker", "增值税率均已提取，但不一致或不是13%。"
    else:
        vat_status, vat_severity, vat_summary = "PASS", "info", "合同与CQP增值税率均为13%。"
    items.append(_item("vat_rate", CATEGORY_OTHER, "增值税率", vat_status, vat_severity, vat_summary, {"合同": contract_vat or "未提取", "CQP": cqp_vat or "未提取"}, [_evidence(contract_pdf, "contract", "合同VAT", ["13%"]), _evidence(cqp_pdf, "cqp", "CQP VAT", ["13%"]) ]))

    payment_status, payment_severity, payment_summary = _payment_terms_consistency(contract.get("payment_terms", {}), cqp.get("payment_terms", {}))
    items.append(_item(
        "payment_terms", CATEGORY_OTHER, "付款条件与开票", payment_status, payment_severity, payment_summary,
        {"合同附件二原文": contract.get("payment_terms", {}).get("raw") or "未提取", "CQP付款条件": cqp.get("payment_terms", {}).get("raw") or "未提取", "合同抽取状态": contract.get("payment_terms", {}).get("extraction_state") or "未知", "BT09填写规则": "原文照抄合同附件二，不得简写比例或省略备注"},
        [_evidence(contract_pdf, "contract", "合同付款条件", [contract.get("payment_terms", {}).get("raw", ""), "付款条件"]), _evidence(cqp_pdf, "cqp", "CQP付款条件", [cqp.get("payment_terms", {}).get("raw", ""), "付款条件"])],
    ))

    schedule = contract.get("delivery_schedule", [])
    direct_weeks = int(contract.get("delivery_weeks", 0) or 0)
    q_weeks = int(cqp.get("delivery_weeks", 0) or 0)
    schedule_weeks = sorted({int(entry.get("weeks", 0)) for entry in schedule if entry.get("weeks")})
    contract_weeks = [direct_weeks] if direct_weeks else schedule_weeks
    contract_trigger = str(contract.get("delivery_trigger", "") or "")
    cqp_trigger = str(cqp.get("delivery_trigger", "") or "")
    if not contract_weeks:
        delivery_status, delivery_severity, delivery_summary = "UNDETERMINED", "blocker", "合同交期无法提取，按抽取失败处理。"
    elif not q_weeks:
        delivery_status, delivery_severity, delivery_summary = "WARNING", "warning", "CQP交期未提取；BT09按合同交期填写。"
    elif any(week != q_weeks for week in contract_weeks) or (contract_trigger and cqp_trigger and contract_trigger != cqp_trigger):
        delivery_status, delivery_severity, delivery_summary = "WARNING", "warning", "合同与CQP交期或起算条件不同；非阻塞，BT09以合同为准。"
    else:
        delivery_status, delivery_severity, delivery_summary = "PASS", "info", "合同与CQP交期一致；BT09以合同为准。"
    contract_display = f"{direct_weeks}周" if direct_weeks else "；".join(f"{entry.get('model')} {entry.get('weeks')}周" for entry in schedule)
    items.append(_item(
        "delivery_period", CATEGORY_OTHER, "交付周期与起算条件", delivery_status, delivery_severity, delivery_summary,
        {"合同": contract_display or "未提取", "合同起算": contract_trigger or "未提取", "CQP": cqp.get("delivery_time") or "未提取", "CQP起算": cqp_trigger or "未提取", "BT09来源": "合同"},
        _evidence_many(contract_pdf, "contract", "合同交付周期", [str(week) + "周" for week in contract_weeks] + [contract_trigger])
        + _evidence_many(cqp_pdf, "cqp", "CQP交付周期", [cqp.get("delivery_time", "")]),
    ))

    incoterm = _infer_incoterm(contract, cqp)
    items.append(_item(
        "incoterm_delivery_place", CATEGORY_OTHER, "贸易术语、交付地点与Ship-to", incoterm["status"], incoterm["severity"], incoterm["reason"],
        {"结论": incoterm["conclusion"], "合同勾选": incoterm["contract_selected"] or "未识别", "合同交付地点": incoterm["delivery_location"] or "未提取", "CQP": cqp.get("delivery_term") or "未提取", "CQP named place": incoterm.get("cqp_named_place") or "未提取", "BT09 Incoterm 2": incoterm["incoterm_2"] or "待确认", "BT09 Ship-to名称": incoterm["ship_to_name"] or "待确认", "BT09 Ship-to地址": incoterm["ship_to_address"] or "待确认", "Ship-to来源": incoterm.get("ship_to_source") or "未确定", "规则": incoterm["ship_to_rule"]},
        [_evidence(contract_pdf, "contract", "合同贸易术语选项", [line.get("text", "") for line in contract.get("incoterm_detection", {}).get("lines", [])]), _evidence(contract_pdf, "contract", "合同交付地点", [contract.get("delivery_location", "")]), _evidence(cqp_pdf, "cqp", "CQP贸易术语", [cqp.get("delivery_term", "")])],
    ))

    signature_placeholders = sorted(set(contract.get("signature_placeholders", []) + ta.get("signature_placeholders", [])))
    signature_issue = bool(signature_placeholders or contract.get("blank_signature_dates") or ta.get("blank_signature_dates"))
    items.append(_item("signature_completeness", CATEGORY_OTHER, "签字、盖章与日期完整性", "WARNING" if signature_issue else "PASS", "warning" if signature_issue else "info", "文件仍含签字/盖章占位符或空日期；最终签署版必须补全。" if signature_issue else "签署信息完整。", {"占位符": "、".join(signature_placeholders) or "无", "合同日期空白": bool(contract.get("blank_signature_dates")), "TA日期空白": bool(ta.get("blank_signature_dates"))}, []))

    attachments = contract.get("attachments", {})
    source_missing: List[str] = []
    extraction_failed: List[str] = []
    if not documents.ta:
        source_missing.append("技术协议")
    elif not ta_has_content:
        extraction_failed.append("技术协议内容")
    if not attachments.get("payment"):
        source_missing.append("附件二付款条件标题")
    elif not contract.get("payment_terms", {}).get("raw"):
        extraction_failed.append("附件二付款条件内容")
    integrity_missing = not attachments.get("integrity")
    if source_missing:
        attachment_status, attachment_severity = "MISMATCH", "blocker"
        attachment_summary = "缺少关键附件来源：" + "、".join(source_missing)
        attachment_state = "SOURCE_MISSING"
    elif extraction_failed:
        attachment_status, attachment_severity = "UNDETERMINED", "warning"
        attachment_summary = "附件存在但内容未可靠抽取：" + "、".join(extraction_failed) + "；阻塞由对应字段项承担，本项不重复阻塞。"
        attachment_state = "EXTRACTION_FAILED"
    elif integrity_missing:
        attachment_status, attachment_severity = "WARNING", "warning"
        attachment_summary = "诚信条款附件未识别。"
        attachment_state = "REVIEW_REQUIRED"
    else:
        attachment_status, attachment_severity = "PASS", "info"
        attachment_summary = "关键附件内容完整。"
        attachment_state = "MATCH"
    items.append(_item("attachment_completeness", CATEGORY_OTHER, "附件完整性与文件优先级", attachment_status, attachment_severity, attachment_summary, {"技术协议内容": ta_has_content, "付款条件内容": bool(contract.get("payment_terms", {}).get("raw")), "诚信条款": bool(attachments.get("integrity")), "优先级": contract.get("file_priority") or "未提取"}, [], decision_state=attachment_state))

    responsibilities = ta.get("responsibilities", {})
    responsibility_ok = bool(responsibilities) and all(responsibilities.values())
    items.append(_item("scope_responsibility", CATEGORY_OTHER, "供货范围与责任边界", "PASS" if responsibility_ok else "UNDETERMINED", "info" if responsibility_ok else "warning", "TA已提取双方责任边界。" if responsibility_ok else "责任边界未完整提取，需人工确认TA条款。", responsibilities, []))
    return items

# ---------------------------------------------------------------------------
# Customer master, BT09 draft, optional LLM, and orchestration
# ---------------------------------------------------------------------------


def _legacy_check(item: Dict[str, Any]) -> Dict[str, Any]:
    return {
        "check_name": item.get("title", ""),
        "status": item.get("status", ""),
        "detail": item.get("summary", ""),
        "is_blocker": item.get("severity") == "blocker" and item.get("status") != "PASS",
    }


def _header_key(value: Any) -> str:
    return _compact(str(value or ""))


def _load_customer_master_rows(path: str) -> List[Dict[str, Any]]:
    extension = os.path.splitext(path)[1].lower()
    rows: List[Dict[str, Any]] = []
    if extension in {".xlsx", ".xlsm"}:
        from openpyxl import load_workbook
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            for sheet in workbook.worksheets:
                iterator = sheet.iter_rows(values_only=True)
                headers = next(iterator, None)
                if not headers:
                    continue
                header_names = [str(value or "").strip() for value in headers]
                for values in iterator:
                    row = {header_names[index]: values[index] for index in range(min(len(header_names), len(values))) if header_names[index]}
                    if any(value not in (None, "") for value in row.values()):
                        row["__sheet__"] = sheet.title
                        rows.append(row)
        finally:
            workbook.close()
        return rows
    if extension == ".csv":
        with open(path, "r", encoding="utf-8-sig", newline="") as handle:
            for row in csv.DictReader(handle):
                if any(value not in (None, "") for value in row.values()):
                    rows.append(dict(row))
        return rows
    raise ValueError("客户主数据仅支持 .xlsx、.xlsm 或 .csv。")


def _row_values_by_headers(row: Dict[str, Any], aliases: Sequence[str]) -> List[str]:
    alias_keys = [_header_key(alias) for alias in aliases]
    values: List[str] = []
    for header, value in row.items():
        if str(header).startswith("__") or value in (None, ""):
            continue
        key = _header_key(header)
        if any(alias == key or alias in key for alias in alias_keys):
            values.append(_clean_inline(str(value)))
    return values


def _match_customer_master(path: Optional[str], contract: Dict[str, Any], incoterm: Dict[str, Any]) -> Dict[str, Any]:
    if not path:
        return {"status": "not_configured", "available": False, "matched": False, "note": "未提供客户主数据表。"}
    if not os.path.exists(path):
        return {"status": "error", "available": False, "matched": False, "note": f"客户主数据表不存在：{path}"}
    try:
        rows = _load_customer_master_rows(path)
    except Exception as exc:
        return {"status": "error", "available": False, "matched": False, "note": f"客户主数据读取失败：{exc}"}

    target_names = [contract.get("buyer_name", ""), contract.get("end_customer_name", ""), incoterm.get("ship_to_name", "")]
    target_addresses = [contract.get("buyer_address", ""), contract.get("end_customer_address", ""), incoterm.get("ship_to_address", "")]
    name_aliases = ("客户名称", "买方名称", "最终用户名称", "收货方名称", "ship-to name", "ship to name", "customer name", "end customer")
    address_aliases = ("客户地址", "买方地址", "最终用户地址", "收货地址", "ship-to address", "ship to address", "customer address")
    ship_to_aliases = ("ship-to id", "ship to id", "shiptoid", "收货方id", "收货客户id")
    end_customer_aliases = ("end customer id", "endcustomerid", "最终用户id", "终端客户id")
    gis_aliases = ("gis号", "gis number", "gis no", "gis")

    best: Optional[Tuple[int, Dict[str, Any]]] = None
    for row in rows:
        row_names = _row_values_by_headers(row, name_aliases)
        row_addresses = _row_values_by_headers(row, address_aliases)
        name_match = any(target and any(_contains_equivalent(target, candidate) for candidate in row_names) for target in target_names)
        if not name_match:
            # Business rule: address alone must never produce a customer ID match.
            continue
        score = 6
        if any(target and any(_address_match(target, candidate) for candidate in row_addresses) for target in target_addresses):
            score += 2
        if best is None or score > best[0]:
            best = (score, row)
    if not best:
        return {"status": "not_found", "available": True, "matched": False, "note": "客户主数据中未找到名称可信匹配，ID/GIS保持为空。", "rows_checked": len(rows)}

    row = best[1]
    def first_value(aliases: Sequence[str]) -> str:
        values = _row_values_by_headers(row, aliases)
        return values[0] if values else ""
    return {
        "status": "matched", "available": True, "matched": True, "score": best[0], "sheet": row.get("__sheet__", ""),
        "ship_to_id": first_value(ship_to_aliases), "end_customer_id": first_value(end_customer_aliases), "gis_number": first_value(gis_aliases),
        "note": "客户主数据已按名称优先、地址辅助匹配；空字段仍需人工补充。",
    }


def _customer_master_review_item(result: Dict[str, Any]) -> Dict[str, Any]:
    if result.get("status") == "matched":
        status, severity, summary = "PASS", "info", "客户主数据已匹配，已提取可用ID/GIS。"
    elif result.get("status") == "not_configured":
        status, severity, summary = "INFO", "info", "未提供客户主数据表；ID保持为空，不影响合同审核结论。"
    else:
        status, severity, summary = "WARNING", "warning", result.get("note") or "客户主数据未匹配。"
    return _item("customer_master_ids", CATEGORY_CUSTOMER, "客户ID与GIS匹配", status, severity, summary, {"Ship-to ID": result.get("ship_to_id") or "未匹配", "End Customer ID": result.get("end_customer_id") or "未匹配", "GIS": result.get("gis_number") or "未匹配", "匹配说明": result.get("note") or ""}, [])


def _build_bt09_draft(contract: Dict[str, Any], cqp: Dict[str, Any], ta: Dict[str, Any], items: Sequence[Dict[str, Any]], customer_master: Optional[Dict[str, Any]] = None) -> Dict[str, Any]:
    del ta
    customer_master = customer_master or {}
    by_id = {item.get("id"): item for item in items}
    incoterm_item = by_id.get("incoterm_delivery_place", {})
    values = incoterm_item.get("values", {})
    blockers = [item.get("id") for item in items if item.get("severity") == "blocker" and item.get("status") != "PASS"]
    incoterm = values.get("结论", "UNDETERMINED")
    return {
        "ready": not blockers,
        "blocked_by": blockers,
        "template_type": f"BT09_{incoterm}" if incoterm in {"DDP", "EXW"} else "",
        "contract_number": contract.get("contract_number", ""),
        "cqp_number": cqp.get("cqp_number", ""),
        "buyer_name": contract.get("buyer_name", ""),
        "buyer_address": contract.get("buyer_address", ""),
        "end_customer_name": contract.get("end_customer_name", ""),
        "end_customer_address": contract.get("end_customer_address", ""),
        "products": contract.get("products", []),
        "total_qty": contract.get("total_qty", 0),
        "incoterm": incoterm,
        "incoterm_2": values.get("BT09 Incoterm 2", ""),
        "ship_to_name": values.get("BT09 Ship-to名称", ""),
        "ship_to_address": values.get("BT09 Ship-to地址", ""),
        "delivery_terms": by_id.get("delivery_period", {}).get("values", {}).get("合同", ""),
        "delivery_trigger": contract.get("delivery_trigger", ""),
        "payment_terms_verbatim": contract.get("payment_terms", {}).get("raw", ""),
        "warranty_summary": by_id.get("warranty_by_model", {}).get("summary", ""),
        "vat_rate": contract.get("vat_rate", 0),
        "untaxed_amount": contract.get("untaxed_amount", 0),
        "tax_included_amount": contract.get("tax_included_amount", 0),
        "sales_person": contract.get("sales_person") or cqp.get("sales_person", ""),
        "pm": contract.get("pm", ""),
        "ship_to_id": customer_master.get("ship_to_id", ""),
        "end_customer_id": customer_master.get("end_customer_id", ""),
        "gis_number": customer_master.get("gis_number", ""),
        "gm": "",
        "nm": "",
        "customer_master_note": customer_master.get("note", "未提供客户主数据表时保持为空，不得编造。"),
    }


def _format_bt09_draft(fields: Dict[str, Any]) -> str:
    product_text = "、".join(f"{item.get('qty', '')}台 {item.get('model', '')}" for item in fields.get("products", []) if item.get("model"))
    lines = [
        f"模板：{fields.get('template_type') or '待确认'}",
        f"合同号：{fields.get('contract_number') or '待确认'}",
        f"CQP号：{fields.get('cqp_number') or '待确认'}",
        f"买方：{fields.get('buyer_name') or '待确认'}",
        f"机器人：{product_text or '待确认'}",
        f"PM：{fields.get('pm') or '待确认'}",
        f"Sales：{fields.get('sales_person') or '待确认'}",
        f"Incoterm：{fields.get('incoterm') or '待确认'}",
        f"Incoterm 2：{fields.get('incoterm_2') or '待确认'}",
        f"Ship-to：{fields.get('ship_to_name') or '待确认'}",
        f"Ship-to地址：{fields.get('ship_to_address') or '待确认'}",
        f"Ship-to ID：{fields.get('ship_to_id') or '待补充'}",
        f"End Customer ID：{fields.get('end_customer_id') or '待补充'}",
        f"GIS：{fields.get('gis_number') or '待补充'}",
        f"GM：{fields.get('gm') or '待补充'}",
        f"NM：{fields.get('nm') or '待补充'}",
        f"交期：{fields.get('delivery_terms') or '待确认'}",
        f"交期起算：{fields.get('delivery_trigger') or '待确认'}",
        "付款条件（合同附件二原文）：",
        fields.get("payment_terms_verbatim") or "待确认",
        f"质保：{fields.get('warranty_summary') or '待确认'}",
    ]
    if fields.get("blocked_by"):
        lines.extend(("", "当前不可发起BT09，阻塞项：" + "、".join(str(item) for item in fields["blocked_by"])))
    elif fields.get("customer_master_note"):
        lines.extend(("", "客户主数据：" + str(fields["customer_master_note"])))
    return "\n".join(lines)


def _run_optional_llm(contract: Dict[str, Any], cqp: Dict[str, Any], ta: Dict[str, Any], items: List[Dict[str, Any]]) -> Dict[str, Any]:
    by_id = {item.get("id"): item for item in items}
    incoterm_item = by_id.get("incoterm_delivery_place", {})
    try:
        result = run_llm_contract_review(
            contract_data=contract,
            cqp_data=cqp,
            ta_data=ta,
            incoterm_result={
                "conclusion": incoterm_item.get("values", {}).get("结论", "UNDETERMINED"),
                "contract_evidence": incoterm_item.get("values", {}).get("合同勾选", ""),
                "cqp_evidence": incoterm_item.get("values", {}).get("CQP", ""),
                "consistent": incoterm_item.get("status") == "PASS",
                "status": incoterm_item.get("status", ""),
                "reason": incoterm_item.get("summary", ""),
            },
            consistency_results=[_legacy_check(item) for item in items],
            warranty_result={"consistent": by_id.get("warranty_by_model", {}).get("status") in {"PASS", "WARNING"}, "detail": by_id.get("warranty_by_model", {}).get("summary", ""), "contract_warranty": contract.get("warranty", {}), "cqp_warranty_codes": cqp.get("warranty_codes_by_model", {}), "ta_warranty_codes": ta.get("warranty_codes_by_model", {})},
            config_result={"overall_consistent": by_id.get("configuration_consistency", {}).get("status") in {"PASS", "WARNING"}, "models_compared": by_id.get("configuration_consistency", {}).get("sub_items", [])},
            financial_result={"vat_check": by_id.get("vat_rate", {}), "untaxed_check": by_id.get("untaxed_amount", {}), "tax_included_check": by_id.get("tax_included_amount", {})},
        )
        result["knowledge_sources"] = list(get_contract_review_knowledge().source_files)
        return result
    except Exception as exc:
        return {"error": str(exc), "overall_assessment": "Unknown", "summary": "AI审核不可用，规则检查结果不受影响。"}


def _legacy_incoterm_payload(item: Dict[str, Any]) -> Dict[str, Any]:
    values = item.get("values", {})
    return {"conclusion": values.get("结论", "UNDETERMINED"), "contract_evidence": "；".join(value for value in (f"勾选={values.get('合同勾选')}" if values.get("合同勾选") else "", f"交付地点={values.get('合同交付地点')}" if values.get("合同交付地点") else "") if value), "cqp_evidence": values.get("CQP", ""), "consistent": item.get("status") == "PASS", "status": item.get("status", ""), "detail": item.get("summary", "")}


def _legacy_warranty_payload(item: Dict[str, Any], cqp: Dict[str, Any]) -> Dict[str, Any]:
    return {"consistent": item.get("status") in {"PASS", "WARNING"}, "detail": item.get("summary", ""), "cqp_warranty_codes": sorted(set(cqp.get("warranty_codes_by_model", {}).values())), "status": item.get("status", "")}


def _legacy_financial_payload(items: Sequence[Dict[str, Any]], contract: Dict[str, Any], cqp: Dict[str, Any]) -> Dict[str, Any]:
    by_id = {item.get("id"): item for item in items}
    def amount_payload(item: Dict[str, Any]) -> Dict[str, Any]:
        raw_diff = item.get("values", {}).get("差异", 0)
        try:
            diff = float(raw_diff)
        except (TypeError, ValueError):
            diff = 0.0
        return {"status": item.get("status", ""), "diff": diff, "is_rounding": item.get("status") == "WARNING" and diff < AMOUNT_ROUNDING_TOLERANCE, "detail": item.get("summary", "")}
    return {"vat_check": {"status": by_id.get("vat_rate", {}).get("status", ""), "contract_vat": float(contract.get("vat_rate", 0) or 0), "cqp_vat": float(cqp.get("vat_rate", 0) or 0), "detail": by_id.get("vat_rate", {}).get("summary", "")}, "untaxed_check": amount_payload(by_id.get("untaxed_amount", {})), "tax_included_check": amount_payload(by_id.get("tax_included_amount", {}))}


def run_review(
    pdf_paths: List[str],
    customer_db_path: str = None,
    template_path: str = None,
    file_roles: Optional[Dict[str, str]] = None,
) -> Dict[str, Any]:
    # template_path is retained for API compatibility.  The deterministic draft
    # exposes all fields; DOCX rendering remains a presentation-layer concern.
    del template_path
    customer_db_path = customer_db_path or os.environ.get("PM_CUSTOMER_MASTER_PATH", "")
    parsed_by_path = {os.path.abspath(path): parse_pdf(path) for path in pdf_paths}
    documents = _resolve_documents(parsed_by_path, file_roles)
    if documents.contract_physical is None or documents.cqp is None:
        raise ValueError("Contract 和 CQP 都必须上传且能够读取。")
    contract, cqp, ta = extract_contract(documents.contract_body), extract_cqp(documents.cqp), extract_ta(documents.ta)
    review_items = build_review_items(documents, contract, cqp, ta)
    incoterm_data = _infer_incoterm(contract, cqp)
    customer_master = _match_customer_master(customer_db_path, contract, incoterm_data)
    review_items.append(_customer_master_review_item(customer_master))
    blockers = [{"type": item["id"], "detail": item["summary"]} for item in review_items if item.get("severity") == "blocker" and item.get("status") != "PASS"]
    non_blockers = [{"type": item["id"], "detail": item["summary"]} for item in review_items if item.get("severity") == "warning" and item.get("status") != "PASS"]
    conclusion = "Blocked" if blockers else ("Pass with notes" if non_blockers else "Pass")
    by_id = {item.get("id"): item for item in review_items}
    bt09_fields = _build_bt09_draft(contract, cqp, ta, review_items, customer_master)
    return {
        "conclusion": conclusion,
        "review_categories": [{"id": CATEGORY_CUSTOMER, "title": CATEGORY_LABELS[CATEGORY_CUSTOMER], "order": 1}, {"id": CATEGORY_PRODUCT, "title": CATEGORY_LABELS[CATEGORY_PRODUCT], "order": 2}, {"id": CATEGORY_OTHER, "title": CATEGORY_LABELS[CATEGORY_OTHER], "order": 3}],
        "document_sources": {"contract": {"physical_role": "contract", "embedded": False}, "cqp": {"physical_role": "cqp", "embedded": False}, "ta": {"physical_role": "contract" if documents.ta_embedded else "ta", "embedded": documents.ta_embedded}},
        "source_recognition": {"contract": {"status": "found", "page_count": len(documents.contract_physical.pages)}, "cqp": {"status": "found", "page_count": len(documents.cqp.pages)}, "ta": {"status": "embedded" if documents.ta_embedded else ("found" if documents.ta else "not_found"), "page_count": len(documents.ta.pages) if documents.ta else 0}},
        "extracted_data": {"contract": contract, "cqp": cqp, "ta": ta},
        "customer_master": customer_master,
        "key_checks": [_legacy_check(item) for item in review_items],
        "blockers": blockers,
        "non_blockers": non_blockers,
        "review_items": review_items,
        "llm_review": _run_optional_llm(contract, cqp, ta, review_items),
        "incoterm": _legacy_incoterm_payload(by_id.get("incoterm_delivery_place", {})),
        "warranty": _legacy_warranty_payload(by_id.get("warranty_by_model", {}), cqp),
        "configuration": by_id.get("configuration_consistency", {}),
        "financial": _legacy_financial_payload(review_items, contract, cqp),
        "bt09_fields": bt09_fields,
        "bt09_draft": _format_bt09_draft(bt09_fields),
        "rule_knowledge": {"source_files": list(get_contract_review_knowledge().source_files), "config_alias_count": len(get_contract_review_knowledge().aliases_by_code)},
    }


__all__ = [
    "DocumentSet", "parse_pdf", "extract_contract", "extract_cqp", "extract_ta", "build_review_items", "run_review",
    "_amount_status", "_canonical_model", "_extract_configurations", "_extract_cqp_products", "_extract_model_quantities", "_extract_payment_terms", "_extract_warranty_clause",
    "_find_ta_start", "_infer_incoterm", "_match_customer_master",
]
