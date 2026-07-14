# -*- coding: utf-8 -*-
"""Outlook crawler + contract progress classifier for PM Workplace.

This module is intentionally import-safe on non-Windows machines. pywin32 is
imported only when a sync job runs.
"""

from __future__ import annotations

import os
import re
import hashlib
import shutil
import traceback
from collections import defaultdict
from datetime import datetime, timedelta, time
from typing import Any, Callable, Dict, Iterable, List, Optional, Sequence, Tuple

import openpyxl

from . import llm_review, project_db

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DATA_DIR = os.path.join(BASE_DIR, "data")
ATTACHMENT_DIR = os.path.join(DATA_DIR, "attachments")
CONFIG_DIR = os.path.join(BASE_DIR, "config")
DEFAULT_PATTERN_PATH = os.path.join(CONFIG_DIR, "Email Pattern11.xlsx")
MAIL_CLASS = 43

CANDIDATE_KEYWORDS = [
    "开启", "流程", "合同流程", "CHECK", "check", "Approved", "Rejected", "Your ABB Order",
    "Request For Change", "预付", "预付款", "BT09", "BTC", "book in SAP", "下单",
    "iprocess", "QueueID", "AuthorizationHistory", "CQP", "4367",
]

REVIEW_STAGE_ID = "review-required"

STAGE_ORDER = [
    "sales-contract", "pm-bt09", "pa-so-bt09", "iprocess", "book-order", "factory-bt", "factory-oa", REVIEW_STAGE_ID,
]

PROGRESS_TO_STAGE = {
    # Canonical front-end stage labels
    "开启流程": "sales-contract",
    "销售开启合同": "sales-contract",
    "PM开启BT09": "pm-bt09",
    "PA回复SO/BT09": "pa-so-bt09",
    "iProcess审批": "iprocess",
    "iprocess审批": "iprocess",
    "Book订单申请": "book-order",
    "工厂BT回复": "factory-bt",
    "工厂反馈OA": "factory-oa",

    # Legacy / Email Pattern progress names
    "销售合同已开启": "sales-contract",
    "BT09待创建": "pm-bt09",
    "BT09邮件已发送": "pa-so-bt09",
    "iprocess已上传": "iprocess",
    "预付款已到": "book-order",
    "合同待Aimee Book": "book-order",
    "合同已批准（待下单）": "factory-bt",
    "已下单给工厂": "factory-bt",
    "RFC已完成": "factory-oa",
    "OA已反馈": "factory-oa",
    "Your ABB Order": "factory-oa",
    "合同已取消": "sales-contract",
    "待人工确认": REVIEW_STAGE_ID,
    "待人工审核": REVIEW_STAGE_ID,
}


def to_str(value: Any) -> str:
    return "" if value is None else str(value).strip()


def safe_lower(value: Any) -> str:
    return to_str(value).lower()


def to_float(value: Any, default: float = 0.0) -> float:
    try:
        if value is None or value == "":
            return default
        return float(value)
    except Exception:
        return default


def normalize_role(role_text: Any) -> str:
    role_text = to_str(role_text)
    if not role_text:
        return ""
    role_map = {
        "sales": "Sales",
        "pm": "PM",
        "pa": "PA",
        "sa": "SA",
        "iprocess": "Iprocess",
        "global": "Global",
    }
    return role_map.get(role_text.lower(), role_text)


def parse_role_cell(role_text: Any) -> List[str]:
    text = to_str(role_text)
    if not text:
        return []
    roles: List[str] = []
    for part in re.split(r"[,&/]| and |\+|、", text, flags=re.I):
        role = normalize_role(part.strip())
        if role and role not in roles:
            roles.append(role)
    return roles


def normalize_topic(subject: Any) -> str:
    subject = to_str(subject)
    subject = re.sub(r"^\s*((RE|FW|FWD)\s*:\s*)+", "", subject, flags=re.I)
    subject = re.sub(r"\s+", " ", subject).strip().lower()
    return subject


def unique_list(seq: Iterable[Any]) -> List[Any]:
    seen = set()
    result = []
    for item in seq:
        if item not in seen:
            result.append(item)
            seen.add(item)
    return result


CONTRACT_TYPE_TO_ALIAS_TYPE = {
    "normal": "PROJECT_NUMBER",
    "project": "PROJECT_NUMBER",
    "project_number": "PROJECT_NUMBER",
    "ocr": "PROJECT_NUMBER",
    "cq": "CQ_NUMBER",
    "sales_order": "SALES_ORDER",
    "so": "SALES_ORDER",
    "bt": "BT_NUMBER",
    "subject": "FLOW_SUBJECT",
}

ALIAS_TYPE_RANK = {
    "PROJECT_NUMBER": 10,
    "CQ_NUMBER": 20,
    "SALES_ORDER": 30,
    "BT_NUMBER": 40,
    "OCR_NUMBER": 50,
    "FLOW_SUBJECT": 90,
    "UNKNOWN": 99,
}


def alias_type_for_contract(value: Any, contract_type: Any = "") -> str:
    ctype = to_str(contract_type).lower()
    if ctype in CONTRACT_TYPE_TO_ALIAS_TYPE:
        return CONTRACT_TYPE_TO_ALIAS_TYPE[ctype]
    return project_db.infer_alias_type(value)


def normalize_contract_alias(value: Any, contract_type: Any = "") -> Tuple[str, str]:
    value_text = project_db.normalize_alias_value(value)
    alias_type = alias_type_for_contract(value_text, contract_type)
    return alias_type, value_text


def contract_alias_rank(item: Tuple[str, str]) -> Tuple[int, str]:
    value, contract_type = item
    alias_type, alias_value = normalize_contract_alias(value, contract_type)
    return (ALIAS_TYPE_RANK.get(alias_type, 99), alias_value)


def sort_contract_aliases(items: Iterable[Tuple[str, str]]) -> List[Tuple[str, str]]:
    return unique_list(sorted(list(items), key=contract_alias_rank))


def choose_primary_contract(aliases: Iterable[Tuple[str, str]]) -> str:
    sorted_aliases = sort_contract_aliases(aliases)
    return sorted_aliases[0][0] if sorted_aliases else ""


def aliases_for_db(aliases: Iterable[Tuple[str, str]]) -> List[Dict[str, str]]:
    result: List[Dict[str, str]] = []
    seen = set()
    for value, contract_type in aliases:
        alias_type, alias_value = normalize_contract_alias(value, contract_type)
        if not alias_value:
            continue
        key = (alias_type, alias_value)
        if key in seen:
            continue
        seen.add(key)
        result.append({"alias_type": alias_type, "alias_value": alias_value})
    return result


def event_type_from_progress(progress: Any) -> str:
    progress_text = to_str(progress)
    mapping = {
        "开启流程": "CONTRACT_FLOW_OPENED",
        "销售开启合同": "CONTRACT_FLOW_OPENED",
        "销售合同已开启": "CONTRACT_FLOW_OPENED",
        "PM开启BT09": "BT09_REQUESTED",
        "BT09待创建": "BT09_REQUESTED",
        "PA回复SO/BT09": "SO_BT09_RECEIVED",
        "BT09邮件已发送": "SO_BT09_RECEIVED",
        "iProcess审批": "IPROCESS_APPROVAL",
        "iprocess审批": "IPROCESS_APPROVAL",
        "iprocess已上传": "IPROCESS_APPROVAL",
        "Book订单申请": "CHECK_BOOK_REQUESTED",
        "预付款已到": "CHECK_BOOK_REQUESTED",
        "合同待Aimee Book": "CHECK_BOOK_REQUESTED",
        "工厂BT回复": "FACTORY_BT_RECEIVED",
        "合同已批准（待下单）": "FACTORY_BT_RECEIVED",
        "已下单给工厂": "FACTORY_BT_RECEIVED",
        "工厂反馈OA": "FACTORY_OA_RECEIVED",
        "RFC已完成": "FACTORY_OA_RECEIVED",
        "OA已反馈": "FACTORY_OA_RECEIVED",
        "Your ABB Order": "FACTORY_OA_RECEIVED",
        "待人工审核": "REVIEW_REQUIRED",
        "待人工确认": "REVIEW_REQUIRED",
    }
    return mapping.get(progress_text, "EMAIL_EVIDENCE")


def parse_date(value: str, end_of_day: bool = False) -> datetime:
    text = to_str(value)
    if not text:
        base = datetime.now()
    else:
        try:
            base = datetime.strptime(text[:10], "%Y-%m-%d")
        except ValueError:
            base = datetime.fromisoformat(text)
    if end_of_day:
        return datetime.combine(base.date(), time.max.replace(microsecond=0))
    return datetime.combine(base.date(), time.min)


def format_dt(value: Any) -> str:
    value = coerce_datetime(value)
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    return to_str(value)


def coerce_datetime(value: Any) -> Optional[datetime]:
    """Convert Outlook / pywin32 date values to naive Python datetime.

    Outlook COM sometimes returns pywintypes datetime values with timezone info.
    Comparing those directly with HTML date inputs can silently fail or raise.
    """
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.replace(tzinfo=None)
    try:
        # pywintypes Time usually behaves like datetime, but keep a defensive path.
        if all(hasattr(value, attr) for attr in ("year", "month", "day", "hour", "minute", "second")):
            return datetime(value.year, value.month, value.day, value.hour, value.minute, value.second)
    except Exception:
        pass
    text = to_str(value)
    if not text:
        return None
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%dT%H:%M:%S", "%m/%d/%Y %I:%M %p", "%m/%d/%Y %H:%M"):
        try:
            return datetime.strptime(text[:19], fmt)
        except Exception:
            continue
    try:
        return datetime.fromisoformat(text.replace("Z", "+00:00")).replace(tzinfo=None)
    except Exception:
        return None


def get_item_time(item: Any) -> Optional[datetime]:
    for attr in ("ReceivedTime", "SentOn", "CreationTime", "LastModificationTime"):
        try:
            value = coerce_datetime(getattr(item, attr, None))
            if value:
                return value
        except Exception:
            continue
    return None


def date_only(value: Any) -> str:
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    text = to_str(value)
    return text[:10] if len(text) >= 10 else text


def stage_index(stage: Any) -> int:
    stage_text = to_str(stage)
    return STAGE_ORDER.index(stage_text) if stage_text in STAGE_ORDER else -1


def max_datetime(*values: Any) -> Optional[datetime]:
    parsed = [coerce_datetime(value) for value in values]
    parsed = [value for value in parsed if value is not None]
    return max(parsed) if parsed else None


def is_newer_datetime(candidate: Any, reference: Any) -> bool:
    candidate_dt = coerce_datetime(candidate)
    reference_dt = coerce_datetime(reference)
    if not candidate_dt:
        return False
    if not reference_dt:
        return True
    return candidate_dt > reference_dt


def choose_latest_value(existing_value: Any, new_value: Any, use_new: bool) -> Any:
    return new_value if use_new else existing_value


def sanitize_path_part(value: Any, fallback: str) -> str:
    text = to_str(value)
    text = re.sub(r'[<>:"/\\|?*]', "_", text)
    text = re.sub(r"\s+", " ", text).strip(" .")
    return text[:120] or fallback


def build_unique_file_path(directory: str, filename: str) -> str:
    stem, ext = os.path.splitext(filename)
    candidate = os.path.join(directory, filename)
    counter = 1
    while os.path.exists(candidate):
        candidate = os.path.join(directory, f"{stem}_{counter}{ext}")
        counter += 1
    return candidate


class SyncCancelled(Exception):
    """Raised when the user cancels an Outlook sync job."""


def cleanup_paths(paths: Iterable[str]) -> None:
    for path in unique_list([to_str(p) for p in paths if p]):
        try:
            abs_path = os.path.abspath(path)
            root = os.path.abspath(ATTACHMENT_DIR)
            if abs_path.startswith(root + os.sep) and os.path.isdir(abs_path):
                shutil.rmtree(abs_path, ignore_errors=True)
        except Exception:
            continue


def ensure_not_cancelled(cancel_check: Optional[Callable[[], bool]], cleanup: Optional[Callable[[], None]] = None) -> None:
    if cancel_check and cancel_check():
        if cleanup:
            cleanup()
        raise SyncCancelled("用户已取消 Outlook 整合，本次无数据写入。")


# ---------------------------------------------------------------------------
# Pattern workbook
# ---------------------------------------------------------------------------


def keyword_rule_hit(rule: Any, text: Any) -> bool:
    rule = to_str(rule)
    text = to_str(text)
    if not rule or not text:
        return False
    rule_l = rule.lower()
    text_l = text.lower()
    if rule_l == "4367":
        return "4367" in text_l or re.search(r"[MK]4367-\d{4}", text, re.I) is not None
    if rule_l == "re":
        return bool(re.search(r"^\s*re\s*:", text, re.I))
    if rule_l == "btxxxxxx":
        return bool(re.search(r"\bBT\d{4,}\b", text, re.I))
    if rule_l == "50xxxxxxx":
        return bool(re.search(r"\b50\d{4,}\b", text))
    if rule_l == "check - x4367-xxxx- xxxx有限公司 - xx.xx":
        return bool(re.search(r"check\s*-\s*[xmk]?4367-\d{4}\s*-\s*.*?有限公司\s*-\s*[\d,.]+", text, re.I))
    return rule_l in text_l


def normalize_attachment_rule(rule: Any) -> str:
    return to_str(rule).replace(" ", "").replace("≥", ">=")


def match_attachment_rule(rule: Any, attachment_count: int) -> bool:
    rule = normalize_attachment_rule(rule)
    if not rule:
        return False
    if rule in ("0", "=0"):
        return attachment_count == 0
    if rule in ("1", "=1"):
        return attachment_count == 1
    if rule == ">=1":
        return attachment_count >= 1
    match = re.match(r"([<>]=?|=)?(\d+)", rule)
    if not match:
        return False
    op = match.group(1) or "="
    num = int(match.group(2))
    return {
        "=": attachment_count == num,
        ">": attachment_count > num,
        ">=": attachment_count >= num,
        "<": attachment_count < num,
        "<=": attachment_count <= num,
    }.get(op, False)


def match_image_rule(rule: Any, has_image: bool) -> bool:
    rule = safe_lower(rule)
    if rule == "y":
        return has_image is True
    if rule == "n":
        return has_image is False
    return False


def load_config_from_pattern_excel(path: str = DEFAULT_PATTERN_PATH) -> Tuple[Dict[str, Dict[str, Any]], Dict[str, str]]:
    if not os.path.exists(path):
        raise FileNotFoundError(f"未找到 Email Pattern 文件: {path}")
    wb = openpyxl.load_workbook(path, data_only=True)

    sender_role_map: Dict[str, str] = {}
    for role_name in ["PM", "PA", "Sales", "SA", "Iprocess", "Global"]:
        if role_name not in wb.sheetnames:
            continue
        ws = wb[role_name]
        for row in ws.iter_rows(values_only=True):
            for cell in row:
                email = to_str(cell)
                if "@" in email:
                    sender_role_map[email.lower()] = role_name

    if "Email Pattern" not in wb.sheetnames:
        raise ValueError("未找到 'Email Pattern' sheet")

    ws = wb["Email Pattern"]
    patterns: Dict[str, Dict[str, Any]] = {}
    current_progress = ""
    for row in ws.iter_rows(min_row=2, values_only=True):
        threshold = row[0] if len(row) > 0 else None
        progress = to_str(row[1]) if len(row) > 1 else ""
        typical_sender = to_str(row[2]) if len(row) > 2 else ""
        sender_weight = to_float(row[3], 0) if len(row) > 3 else 0
        subject_rule = to_str(row[4]) if len(row) > 4 else ""
        subject_weight = to_float(row[5], 0) if len(row) > 5 else 0
        body_rule = to_str(row[6]) if len(row) > 6 else ""
        body_weight = to_float(row[7], 0) if len(row) > 7 else 0
        attachment_rule = to_str(row[8]) if len(row) > 8 else ""
        attachment_weight = to_float(row[9], 0) if len(row) > 9 else 0
        image_rule = to_str(row[10]) if len(row) > 10 else ""
        image_weight = to_float(row[11], 0) if len(row) > 11 else 0

        if progress:
            current_progress = progress
            patterns[current_progress] = {
                "threshold": to_float(threshold, 0),
                "sender_roles": set(parse_role_cell(typical_sender)),
                "sender_weight": sender_weight,
                "subject_rules": [],
                "body_rules": [],
                "attachment_rule": "",
                "attachment_weight": 0,
                "image_rule": "",
                "image_weight": 0,
            }
        if not current_progress:
            continue
        cfg = patterns[current_progress]
        if subject_rule:
            cfg["subject_rules"].append((subject_rule, subject_weight))
        if body_rule:
            cfg["body_rules"].append((body_rule, body_weight))
        if attachment_rule != "":
            cfg["attachment_rule"] = attachment_rule
            cfg["attachment_weight"] = attachment_weight
        if image_rule != "":
            cfg["image_rule"] = image_rule
            cfg["image_weight"] = image_weight
    return patterns, sender_role_map


# ---------------------------------------------------------------------------
# Outlook helpers
# ---------------------------------------------------------------------------


def _clear_gen_py_cache() -> None:
    """Remove corrupted pywin32 gen_py cache to fix CLSIDToPackageMap errors."""
    try:
        import win32com
        gen_py = getattr(win32com, "__gen_path__", None)
        if not gen_py:
            # pywin32 >= 306 moved gen_py to a different location; try common paths.
            import sysconfig
            temp = sysconfig.get_config_var("TEMP") or os.path.join(os.environ.get("LOCALAPPDATA", ""), "Temp")
            gen_py = os.path.join(temp, "gen_py")
        if os.path.isdir(gen_py):
            # Remove only Outlook-related cache folders (CLSID 00062FFF).
            for entry in os.listdir(gen_py):
                entry_path = os.path.join(gen_py, entry)
                if os.path.isdir(entry_path) and "00062FFF" in entry.upper():
                    shutil.rmtree(entry_path, ignore_errors=True)
    except Exception:
        pass


def connect_outlook() -> Any:
    try:
        import win32com.client  # type: ignore
    except ImportError as exc:
        raise RuntimeError("当前环境未安装 pywin32；Outlook 同步必须在 Windows + 传统 Outlook 环境运行。") from exc

    for attempt in range(2):
        try:
            return win32com.client.GetActiveObject("Outlook.Application")
        except Exception:
            try:
                return win32com.client.Dispatch("Outlook.Application")
            except AttributeError as e:
                if attempt == 0 and "CLSIDToPackageMap" in str(e):
                    _clear_gen_py_cache()
                    continue
                raise


def get_root_folder(namespace: Any, mailbox: str) -> Any:
    mailbox = to_str(mailbox)
    if not mailbox:
        raise ValueError("邮箱账号不能为空。")
    try:
        return namespace.Folders.Item(mailbox)
    except Exception:
        # Fallback: allow fuzzy match on DisplayName for Outlook profiles that expose a display label.
        for i in range(1, namespace.Folders.Count + 1):
            folder = namespace.Folders.Item(i)
            if mailbox.lower() in safe_lower(getattr(folder, "Name", "")):
                return folder
        raise ValueError(f"未找到 Outlook 邮箱根目录: {mailbox}")


def get_folder_by_root_path(namespace: Any, mailbox: str, folder_path: str) -> Any:
    folder = get_root_folder(namespace, mailbox)
    for part in [p.strip() for p in to_str(folder_path).replace("\\", "/").split("/") if p.strip()]:
        folder = folder.Folders.Item(part)
    return folder


def get_sender_smtp(item: Any) -> str:
    try:
        sender = to_str(item.SenderEmailAddress)
        if sender and "@" in sender:
            return sender.lower()
    except Exception:
        pass
    try:
        smtp = item.PropertyAccessor.GetProperty("http://schemas.microsoft.com/mapi/proptag/0x39FE001E")
        if smtp and "@" in smtp:
            return smtp.lower()
    except Exception:
        pass
    try:
        if item.Sender:
            ex_user = item.Sender.GetExchangeUser()
            if ex_user and ex_user.PrimarySmtpAddress:
                return ex_user.PrimarySmtpAddress.lower()
    except Exception:
        pass
    return safe_lower(getattr(item, "SenderEmailAddress", ""))


def iter_folders(folder: Any, include_subfolders: bool) -> Iterable[Tuple[Any, str]]:
    yield folder, to_str(getattr(folder, "FolderPath", "") or getattr(folder, "Name", ""))
    if not include_subfolders:
        return
    try:
        for sub in folder.Folders:
            yield from iter_folders(sub, include_subfolders=True)
    except Exception:
        return


def get_folder_item_count(folder: Any) -> int:
    try:
        return int(folder.Items.Count)
    except Exception:
        return -1


def get_items_in_date_range(folder: Any, start_dt: datetime, end_dt: datetime) -> Iterable[Any]:
    """Yield Outlook MailItem objects in the requested date window.

    This intentionally does not depend on Outlook Items.Restrict. Restrict is
    fragile across Windows locale/date formats and was the likely reason a
    folder with valid messages could return zero rows. We sort descending and
    stop once messages are older than the start date.
    """
    try:
        items = folder.Items
    except Exception:
        return
    try:
        items.Sort("[ReceivedTime]", True)
    except Exception:
        try:
            items.Sort("[SentOn]", True)
        except Exception:
            pass

    try:
        count = int(items.Count)
        iterator = (items.Item(i) for i in range(1, count + 1))
    except Exception:
        iterator = iter(items)

    old_seen_after_hits = 0
    for item in iterator:
        try:
            mail_time = get_item_time(item)
            if not mail_time:
                continue
            if mail_time > end_dt:
                continue
            if mail_time < start_dt:
                old_seen_after_hits += 1
                # With descending sorting, enough older items means the rest are also older.
                if old_seen_after_hits >= 25:
                    break
                continue
            yield item
        except Exception:
            continue


# ---------------------------------------------------------------------------
# Mail parsing/classification
# ---------------------------------------------------------------------------


def extract_current_body(body: Any) -> str:
    body = to_str(body)
    separators = [
        "-----Original Message-----", "---Original Message---", "---------- Forwarded message ---------",
        "From:", "发件人:", "Sent:", "发送时间:", "To:", "收件人:", "Cc:", "抄送:", "Subject:", "主题:",
    ]
    current_lines = []
    for line in body.splitlines():
        if any(sep.lower() in line.strip().lower() for sep in separators):
            break
        current_lines.append(line)
    result = "\n".join(current_lines).strip()
    return result if result else body


def extract_contracts_with_type(subject: Any, current_body: Any) -> List[Tuple[str, str]]:
    """Extract all stable identifiers visible in the current mail.

    Earlier versions returned only one identifier, preferring M/K4367 over CQ.
    That split one real project into separate rows when the opening mail only had
    ``CQ1106414`` and the later check mail introduced ``M4367-3569`` / SO.

    The new behavior returns every strong identifier, ordered by primary-key
    quality.  Downstream project-object assignment uses the full alias set to
    merge CQ-only, M/K, SO and BT evidence into one project object.
    """
    subject = to_str(subject)
    current_body = to_str(current_body)
    text_sources = [subject, current_body]
    found: List[Tuple[str, str]] = []

    def add(value: str, ctype: str) -> None:
        value = project_db.normalize_alias_value(value)
        if value:
            found.append((value, ctype))

    # Full ABB project / contract number. Preserve OCR marker as ctype, but keep
    # the stable key as M/K4367-xxxx so later non-OCR mails merge correctly.
    for source in text_sources:
        for match in re.finditer(r"(?<![A-Za-z0-9])([MK]4367-\d{4})\s*(OCR\d*)?(?![A-Za-z0-9])", source, re.I):
            add(match.group(1), "ocr" if match.group(2) else "normal")

    # CQ number from sales workflow subjects, forwarded threads and check mails.
    for source in text_sources:
        for match in re.finditer(r"(?<![A-Za-z0-9])(CQ\d{5,})(?![A-Za-z0-9])", source, re.I):
            add(match.group(1), "cq")

    # Sales order and BT identifiers are secondary aliases.  They should not win
    # over M/CQ as the display contract, but they allow later SO/BT-only replies
    # in the same thread to attach to the project.
    for source in text_sources:
        for match in re.finditer(r"(?<!\d)(50\d{5,})(?!\d)", source):
            add(match.group(1), "sales_order")
        for match in re.finditer(r"\b((?:BT[A-Z]?|RTY)\d{4,})\b", source, re.I):
            add(match.group(1), "bt")

    # Short OCR flow key, e.g. 请开启迅亚3248合同OCR流程. Use a neutral key only
    # when no full M/K project number is visible.
    if not any(alias_type_for_contract(value, ctype) == "PROJECT_NUMBER" for value, ctype in found):
        combined = subject + "\n" + current_body
        for match in re.finditer(r"(?<![A-Za-z0-9-])(\d{4})(?!\d).{0,12}(?:合同)?OCR\d*.{0,8}流程", combined, re.I):
            add(f"OCR-{match.group(1)}", "ocr")

    if found:
        return sort_contract_aliases(found)

    # Strong workflow-opening subjects with no explicit visible identifier
    # (common for procurement invitation subjects). Keep them in manual review
    # under a stable subject-derived key instead of dropping them silently.
    subject_key_source = compact_text(strip_reply_forward_prefix(subject))
    if subject_key_source and is_sales_workflow_base_subject(subject):
        digest = hashlib.sha1(subject_key_source.encode("utf-8")).hexdigest()[:8].upper()
        add(f"FLOW-{digest}", "subject")
        return sort_contract_aliases(found)

    return []

def extract_client_name(subject: Any, body: Any) -> str:
    subject_text = to_str(subject)
    text = subject_text + "\n" + to_str(body)
    # Common title pattern: CHECK - M4367-1234 - 客户有限公司 - amount
    match = re.search(r"[A-Z]*4367-\d{4}\s*-\s*([^\n\r-]{2,80}?(?:有限公司|集团|公司|Limited|Ltd\.?|Inc\.?|LLC))", text, re.I)
    if match:
        return sanitize_path_part(match.group(1), "")
    match = re.search(r"([\u4e00-\u9fa5A-Za-z0-9（）()&.·\-\s]{2,80}?(?:有限公司|集团|公司|Limited|Ltd\.?|Inc\.?|LLC))", text, re.I)
    if match:
        return re.sub(r"\s+", " ", match.group(1)).strip(" -_，,。")

    # Subject shorthand patterns from Outlook screenshots.
    clean_subject = strip_reply_forward_prefix(subject_text)
    compact_subject = compact_text(clean_subject)
    match = re.search(r"(?:请|烦请|麻烦|申请)?(?:开启|开通|启动|启用|发起)([\u4e00-\u9fa5A-Za-z0-9]{2,30}?)(?:CQ\d{5,}|[MK]4367-\d{4}|\d{4}).{0,12}(?:合同|OCR)", compact_subject, re.I)
    if match:
        return sanitize_path_part(match.group(1), "")
    match = re.search(r"邀请.*?(?:开启|发起)([\u4e00-\u9fa5A-Za-z0-9]{2,30}?)(?:采购|采购订单)", compact_subject)
    if match:
        return sanitize_path_part(match.group(1), "")
    match = re.search(r"^(.{2,30}?)(?:CQ\d{5,}|[MK]4367-\d{4}).{0,20}发起流程", compact_subject, re.I)
    if match:
        return sanitize_path_part(match.group(1), "").strip("-_，,。 ")
    return ""


def get_item_store_id(item: Any) -> str:
    try:
        parent = getattr(item, "Parent", None)
        if parent is not None:
            return to_str(getattr(parent, "StoreID", ""))
    except Exception:
        pass
    return ""


def count_attachments_and_images(item: Any) -> Tuple[int, bool]:
    attachment_count = 0
    has_image = False
    try:
        for att in item.Attachments:
            attachment_count += 1
            name = safe_lower(getattr(att, "FileName", ""))
            if name.endswith((".jpg", ".jpeg", ".png", ".gif", ".bmp")):
                has_image = True
    except Exception:
        pass
    try:
        if "<img" in to_str(item.HTMLBody).lower():
            has_image = True
    except Exception:
        pass
    return attachment_count, has_image


def get_attachment_names(item: Any) -> List[str]:
    names: List[str] = []
    try:
        for att in item.Attachments:
            name = to_str(getattr(att, "FileName", ""))
            if name:
                names.append(name)
    except Exception:
        pass
    return names


def save_attachments(item: Any, contracts: Sequence[Tuple[str, str]], received: Optional[datetime]) -> Tuple[str, List[Dict[str, Any]]]:
    try:
        if item.Attachments.Count <= 0:
            return "", []
    except Exception:
        return "", []

    contract_part = contracts[0][0] if contracts else "unlinked"
    date_part = received.strftime("%Y-%m-%d") if isinstance(received, datetime) else "unknown-date"
    subject_part = sanitize_path_part(getattr(item, "Subject", ""), "no-subject")
    entry_part = sanitize_path_part(getattr(item, "EntryID", "")[-18:], "no-entry")
    base_target_dir = os.path.join(ATTACHMENT_DIR, sanitize_path_part(contract_part, "contract"), f"{date_part}-{subject_part}-{entry_part}")
    target_dir = base_target_dir
    counter = 1
    while os.path.exists(target_dir):
        target_dir = f"{base_target_dir}_{counter}"
        counter += 1
    os.makedirs(target_dir, exist_ok=True)

    saved_records: List[Dict[str, Any]] = []
    for att in item.Attachments:
        try:
            file_name = sanitize_path_part(getattr(att, "FileName", "attachment"), "attachment")
            path = build_unique_file_path(target_dir, file_name)
            att.SaveAsFile(path)
            saved_records.append({
                "email_entry_id": to_str(getattr(item, "EntryID", "")),
                "contract": contract_part if contract_part != "unlinked" else "",
                "filename": os.path.basename(path),
                "file_path": path,
                "file_size": os.path.getsize(path),
            })
        except Exception:
            continue
    return target_dir if saved_records else "", saved_records


def parse_email_item(item: Any, folder_name: str) -> Optional[Dict[str, Any]]:
    try:
        if item.Class != MAIL_CLASS:
            return None
    except Exception:
        return None
    try:
        subject = to_str(getattr(item, "Subject", ""))
        body = to_str(getattr(item, "Body", ""))
        current_body = extract_current_body(body)
        sender_email = get_sender_smtp(item)
        received_time = coerce_datetime(getattr(item, "ReceivedTime", None)) or get_item_time(item)
        sent_time = coerce_datetime(getattr(item, "SentOn", None))
        attachment_count, has_image = count_attachments_and_images(item)
        attachment_names = get_attachment_names(item)
        conversation_id = to_str(getattr(item, "ConversationID", ""))
        conversation_topic = to_str(getattr(item, "ConversationTopic", ""))
        contracts = extract_contracts_with_type(subject, current_body)
        # Do not save attachments or DB rows during parsing. Sync writes are staged
        # in memory and committed only after the whole job completes.
        attachment_dir, attachment_records = "", []
        client_name = extract_client_name(subject, current_body)
        entry_id = to_str(getattr(item, "EntryID", ""))
        store_id = get_item_store_id(item)

        email = {
            "entry_id": entry_id,
            "store_id": store_id,
            "subject": subject,
            "body": body,
            "html_body": to_str(getattr(item, "HTMLBody", "")),
            "current_body": current_body,
            "sender_name": to_str(getattr(item, "SenderName", "")),
            "sender_email": sender_email,
            "to_recipients": to_str(getattr(item, "To", "")),
            "cc_recipients": to_str(getattr(item, "CC", "")),
            "received_time": received_time,
            "sent_time": sent_time,
            "received_time_text": format_dt(received_time),
            "sent_time_text": format_dt(sent_time),
            "attachment_count": attachment_count,
            "attachment_names": attachment_names,
            "has_image": has_image,
            "conversation_id": conversation_id,
            "topic_key": normalize_topic(conversation_topic or subject),
            "contracts": contracts,
            "contract_source": "direct" if contracts else "",
            "folder": folder_name,
            "attachment_dir": attachment_dir,
            "attachment_records": attachment_records,
            "is_read": int(not bool(getattr(item, "UnRead", False))),
            "importance": getattr(item, "Importance", 1),
            "categories": to_str(getattr(item, "Categories", "")),
            "client_name": client_name,
        }

        return email
    except Exception:
        return None


def compact_text(value: Any) -> str:
    return re.sub(r"\s+", "", to_str(value)).lower()


REPLY_FORWARD_PREFIX_RE = re.compile(r"^\s*(?:(?:re|fw|fwd)\s*(?:[:：\-]|(?=[\s\u4e00-\u9fa5A-Z0-9]))|回复\s*[:：]?|转发\s*[:：]?|答复\s*[:：]?)\s*", re.I)


def strip_reply_forward_prefix(subject: Any) -> str:
    text = to_str(subject)
    previous = None
    while text and previous != text:
        previous = text
        text = REPLY_FORWARD_PREFIX_RE.sub("", text).strip()
    return text


def has_reply_forward_prefix(subject: Any) -> bool:
    return bool(REPLY_FORWARD_PREFIX_RE.match(to_str(subject)))


def is_sales_workflow_base_subject(subject: Any) -> bool:
    """High-confidence subject for the first workflow-opening mail.

    This deliberately ignores sender identity. It matches the real subject
    families seen in the mailbox screenshots:
    - 请开启/开通/启动...合同流程
    - 请开启...合同OCR流程 / M4367-3248OCR2流程开启
    - 【邀请】开启...采购 外部/内部流程
    - 【邀请】发起...采购订单 内部/外部流程
    - ...CQ1093788-M4367-3525--发起流程
    """
    text = compact_text(strip_reply_forward_prefix(subject))
    if not text:
        return False

    # Classic contract-flow titles.
    if "合同流程开启" in text or re.search(r"(?:合同)?OCR\d*流程开启", text, re.I):
        return True
    if re.search(r"(?:请|烦请|麻烦|申请)?(?:开启|开通|启动|启用|发起).{0,80}(?:合同|OCR\d*).{0,12}流程", text, re.I):
        return True

    # Purchasing-flow invitation titles used by Paul-dongpo Li mails.
    if "【邀请】" in text or "[邀请]" in text or "邀请" in text:
        if re.search(r"(?:开启|发起).{0,60}(?:采购|采购订单).{0,60}(?:外部|内部|内外部|外部/内部|内部/外部).{0,20}流程", text):
            return True
        if re.search(r"(?:开启|发起).{0,60}(?:外部|内部|内外部|外部/内部|内部/外部).{0,20}(?:采购|采购订单).{0,20}流程", text):
            return True
        if re.search(r"(?:采购|采购订单).{0,30}(?:开启|发起).{0,60}(?:外部|内部|内外部|外部/内部|内部/外部).{0,20}流程", text):
            return True

    # M/CQ linked title ending with 发起流程.
    if "发起流程" in text and (re.search(r"[mk]4367-\d{4}", text, re.I) or re.search(r"cq\d{5,}", text, re.I)):
        return True

    return False


def is_sales_workflow_opening_body(body: Any) -> bool:
    """High-confidence current-body opening request.

    Used for mails whose subject is a forward/reply but the latest body contains
    the actual request, e.g. ``请加急开启合同流程``.
    """
    text = compact_text(body)
    if not text:
        return False
    return bool(re.search(r"(?:请|烦请|麻烦|申请)?(?:加急)?(?:开启|开通|启动|启用|发起).{0,80}(?:合同|OCR\d*|采购|采购订单).{0,30}流程", text, re.I))


def is_workflow_subject_family(subject: Any) -> bool:
    """Subject-level gate for workflow candidates, with no sender checks."""
    text = compact_text(strip_reply_forward_prefix(subject))
    if not text:
        return False
    if is_sales_workflow_base_subject(subject):
        return True
    if is_check_book_subject(subject) or is_oa_subject(subject):
        return True
    if "合同流程" in text or "合同ocr" in text or re.search(r"ocr\d*流程", text, re.I):
        return True
    if "采购" in text and "流程" in text and ("邀请" in text or "开启" in text or "发起" in text):
        return True
    if "发起流程" in text:
        return True
    return False


def has_contract_id(email: Dict[str, Any]) -> bool:
    if email.get("contracts"):
        return True
    return bool(re.search(r"[MK]4367-\d{4}", to_str(email.get("subject")) + "\n" + to_str(email.get("current_body")), re.I))


def has_business_number(text: Any) -> bool:
    value = to_str(text)
    return bool(
        re.search(r"\bBT[A-Z]?\d{4,}\b", value, re.I)
        or re.search(r"\b50\d{5,}\b", value)
        or re.search(r"\bCQ\d{5,}\b", value, re.I)
    )


def is_check_book_subject(subject: Any) -> bool:
    text = compact_text(strip_reply_forward_prefix(subject))
    raw_text = compact_text(subject)
    # Supports both old subjects like "check-M4367-3569" and real inbox subjects
    # like "RE: check-K4367-3206 - 客户 - 金额".
    return (
        bool(re.search(r"^check[-_：:]*.{0,160}(?:[mk]4367-\d{4}|cq\d{5,}|[一-龥]{2,20})", text, re.I))
        or bool(re.search(r"(?:^|回复:|re:)check[-_：:]*", raw_text, re.I))
        or "bookinsap" in text
    )


def is_oa_subject(subject: Any) -> bool:
    text = compact_text(subject)
    return "yourabborder" in text or "orderacknowledgement" in text or "acknowledgement" in text or bool(re.search(r"(^|[^a-z])oa([^a-z]|$)", text, re.I))


def contains_bt_code(text: Any) -> bool:
    return bool(re.search(r"\b(?:BT[A-Z]?|RTY)\d{4,}\b", to_str(text), re.I))


def contains_so_code(text: Any) -> bool:
    return bool(re.search(r"(?<!\d)50\d{5,}(?:-\d+)*", to_str(text)))


def contains_project_or_cq(text: Any) -> bool:
    value = to_str(text)
    return bool(
        re.search(r"(?<![A-Za-z0-9])[MK]4367-\d{4}(?![A-Za-z0-9])", value, re.I)
        or re.search(r"(?<![A-Za-z0-9])CQ\d{5,}(?![A-Za-z0-9])", value, re.I)
        or re.search(r"(?<![A-Za-z0-9-])\d{4}(?!\d).{0,16}(?:合同)?OCR\d*", value, re.I)
    )


def is_book_request_text(subject: Any, body: Any) -> bool:
    subject_c = compact_text(subject)
    body_c = compact_text(body)
    return bool(
        is_check_book_subject(subject)
        or subject_c.startswith("check")
        or "pleasecheckandbookinsap" in body_c
        or "bookinsap" in body_c
        or "请check" in body_c
        or "请检查并book" in body_c
        or "待aimeebook" in body_c
    )


def is_pm_bt09_request_text(subject: Any, body: Any) -> bool:
    subject_c = compact_text(subject)
    body_c = compact_text(body)
    all_c = subject_c + "\n" + body_c
    # PA -> Aimee intercompany-SO instructions are downstream replies, not PM
    # requests to Shelly.
    if ("dearaimee" in body_c or "hiaimee" in body_c) and ("按附件建" in body_c or "建跨公司调账中间" in body_c):
        return False
    return bool(
        re.search(r"请(?:帮忙|协助|加急)?(?:创建|建|开|开启)bt09", body_c, re.I)
        or re.search(r"(?:创建|建|开启)bt09(?:订单)?", body_c, re.I)
        or "bt09订单" in body_c
        or "合同信息核对无误" in body_c and "bt09" in body_c
        or "请下单" in body_c
        or "请帮忙下单" in body_c
        or "软件下单" in subject_c
        or ("请建工单" in subject_c and ("工单已建" in body_c or "请继续" in body_c or "工单" in body_c))
        or "po如附件" in body_c
        or "请建中间号" in body_c
        or "建中间号" in body_c
        or "调账中间" in body_c
        or ("预付款已到" in all_c and "bt09" in body_c)
    )

def is_pa_so_bt09_reply_text(subject: Any, body: Any, sender_name: Any = "") -> bool:
    subject_c = compact_text(subject)
    body_text = to_str(body)
    body_c = compact_text(body_text)
    sender_c = compact_text(sender_name)
    has_bt = contains_bt_code(subject) or contains_bt_code(body_text)
    has_so = contains_so_code(subject) or contains_so_code(body_text)

    # PM requests to Shelly can include existing SO/BT inventory references. Do
    # not convert those into PA replies just because both numbers are present.
    if is_pm_bt09_request_text(subject, body):
        return False

    # RARO/SalesOrderApprovalRequest replies (including 回复/Approved/价格审批)
    # are NOT PA SO/BT09 replies — they are iProcess approval traffic.
    if "salesorderapprovalrequest" in subject_c or "价格审批" in subject_c:
        return False

    if has_bt and has_so:
        return True
    # PA/Shelly often replies with only a BTC/BTY number after the PM request.
    if has_bt and ("回复" in subject_c or subject_c.startswith("re") or subject_c.startswith("fw") or "下单" in subject_c or "工单" in subject_c or "slot" in subject_c):
        return True
    # SO-only operational replies: direct SO item, SAP number, or intercompany SO.
    if has_so and ("soitem" in body_c or "直接下进so" in body_c or "新建sap" in subject_c or "sap号" in subject_c):
        return True
    if ("dearaimee" in body_c or "hiaimee" in body_c) and has_so and ("rbttoarh" in body_c or "跨公司调账" in body_c or "so" in body_c):
        return True
    if "shelly" in sender_c and has_so and ("按附件建" in body_c or "跨公司调账" in body_c):
        return True
    # RFC table/request replies in this mailbox are PA-level order maintenance
    # evidence unless they contain an actual Order Acknowledgement/receipt.
    if "rfc" in body_c and (has_bt or has_so or re.search(r"cq\d{5,}", body_c, re.I)) and not is_factory_bt_reply_text(subject, body):
        return True
    return False

def is_iprocess_approval_text(subject: Any, body: Any) -> bool:
    subject_c = compact_text(subject)
    body_text = to_str(body)
    body_c = compact_text(body_text)
    combined_c = subject_c + "\n" + body_c
    if "iprocess" in combined_c or "queueid" in combined_c or "authorizationhistory" in combined_c or "display?pagetype" in combined_c:
        return True
    if "salesorderapprovalrequest" in subject_c:
        return True
    if "价格审批" in subject_c:
        return True
    # Icey/iProcess completion mails often contain only one concise line:
    # K4367-3519 - 客户有限公司 - 127,600.00
    project_line = re.search(
        r"(?im)^\s*[MK]4367-\d{4}(?:\s*OCR\d*)?\s*[-－—]\s*.{2,80}?\s*[-－—]\s*[\d,]+(?:\.\d{2})?\s*$",
        body_text,
    )
    if project_line:
        return True
    if re.search(r"[MK]4367-\d{4}(?:\s*OCR\d*)?\s*[-－—]\s*.{2,80}?(?:有限公司|公司|集团|Limited|Ltd\.?|Inc\.?)\s*[-－—]\s*[\d,]+", body_text, re.I):
        return True
    return False


def is_factory_bt_reply_text(subject: Any, body: Any) -> bool:
    subject_c = compact_text(subject)
    body_c = compact_text(body)
    return bool(
        is_oa_subject(subject)
        or "orderacknowledgement" in body_c
        or "orderreceipt" in body_c
        or "thankyouforplacingthisorder" in body_c
        or ("rfc" in subject_c and contains_bt_code(subject + "\n" + to_str(body)) and ("approved" in body_c or "approved" in subject_c))
    )


def ensure_tracking_alias(email: Dict[str, Any]) -> bool:
    """Give valid workflow mail without visible M/CQ/SO/BT a stable review key."""
    if email.get("contracts"):
        return False
    seed = to_str(email.get("conversation_id")) or compact_text(strip_reply_forward_prefix(email.get("subject"))) or compact_text(email.get("current_body"))[:80]
    if not seed:
        return False
    digest = hashlib.sha1(seed.encode("utf-8")).hexdigest()[:8].upper()
    email["contracts"] = [(f"FLOW-{digest}", "subject")]
    email["contract_source"] = "tracking_fallback"
    return True


def strict_subject_evidence(email: Dict[str, Any]) -> Tuple[str, int, List[str]]:
    """Return (level, score, reasons) for workflow candidacy.

    The old gate required a visible M/K/CQ/OCR id before classification.  That
    dropped real inbox cases from the labeled CSV: concise iProcess replies,
    CHECK mails whose project id is only in the body, SO/BT replies, and
    intercompany adjustment mails.  The gate now accepts a mail when it has both
    a workflow action signal and a stable identifier/business number, or when the
    subject itself is a high-confidence workflow-opening title.
    """
    subject = to_str(email.get("subject"))
    current_body = to_str(email.get("current_body"))
    full_body = to_str(email.get("body"))
    combined = subject + "\n" + current_body + "\n" + full_body
    subject_c = compact_text(subject)
    reasons: List[str] = []

    has_any_alias = bool(email.get("contracts"))
    has_id_or_business_number = has_any_alias or has_business_number(combined) or contains_project_or_cq(combined)

    base_ok = is_sales_workflow_base_subject(subject)
    body_opening_ok = is_sales_workflow_opening_body(current_body)
    pm_ok = is_pm_bt09_request_text(subject, current_body)
    pa_ok = is_pa_so_bt09_reply_text(subject, current_body, email.get("sender_name", ""))
    book_ok = is_book_request_text(subject, current_body)
    iprocess_ok = is_iprocess_approval_text(subject, current_body)
    factory_ok = is_factory_bt_reply_text(subject, current_body)

    if base_ok:
        if has_reply_forward_prefix(subject):
            return "confirmed", 100, ["标题属于流程邮件RE/FW"]
        return "confirmed", 120, ["标题命中流程开启根邮件"]

    if body_opening_ok and has_id_or_business_number:
        return "confirmed", 115, ["最新正文命中流程开启请求"]

    if pa_ok and has_id_or_business_number:
        return "confirmed", 118, ["正文命中SO/BT09回复"]

    if factory_ok and (has_id_or_business_number or contains_bt_code(subject + "\n" + current_body)):
        return "confirmed", 116, ["标题/正文命中工厂BT/OA回复"]

    if pm_ok and has_id_or_business_number:
        return "confirmed", 112, ["正文命中PM开启BT09/下单请求"]

    if book_ok and has_id_or_business_number:
        return "confirmed", 110, ["标题/正文命中CHECK/book订单申请"]

    if iprocess_ok and has_id_or_business_number:
        return "confirmed", 108, ["正文命中iProcess审批结果"]

    if is_workflow_subject_family(subject) and has_id_or_business_number:
        return "review", 65, ["标题属于流程邮件家族，但未命中可自动归类强规则"]

    if ("流程" in subject_c) and has_id_or_business_number:
        return "review", 55, ["标题含流程和项目/业务编号，但格式不标准"]

    if not has_id_or_business_number:
        return "ignore", 0, ["未命中项目编号/MK/CQ/OCR/SO/BT"]
    return "ignore", 0, ["标题/正文不属于流程邮件格式"]

def get_candidate_hits(subject: Any, current_body: Any) -> List[str]:
    text_subject = safe_lower(subject)
    text_body = safe_lower(current_body)
    hits = []
    for keyword in CANDIDATE_KEYWORDS:
        kw_l = keyword.lower()
        if kw_l in text_subject or kw_l in text_body:
            hits.append(keyword)
    return unique_list(hits)


def is_candidate_email(email: Dict[str, Any], sender_role_map: Dict[str, str]) -> Tuple[bool, int, List[str]]:
    level, score, reasons = strict_subject_evidence(email)
    if level == "ignore":
        return False, score, reasons
    # Do not use sender as evidence. Sender is kept only for display/opening mail.
    # Raised threshold: only solid evidence (>=65) passes the script gate.
    # Borderline scores (50-64) that have project/business numbers are now
    # forwarded to the LLM as uncertain items instead of being dropped.
    return score >= 65, score, reasons


def score_progress(email: Dict[str, Any], progress_name: str, cfg: Dict[str, Any], sender_role_map: Dict[str, str]) -> Tuple[float, List[str]]:
    subject = to_str(email["subject"])
    body = to_str(email["current_body"])
    score = 0.0
    matched: List[str] = []
    # Sender identity is deliberately excluded from scoring.
    for rule, weight in cfg["subject_rules"]:
        if keyword_rule_hit(rule, subject):
            score += weight
            matched.append(f"标题:{rule}")
    for rule, weight in cfg["body_rules"]:
        if keyword_rule_hit(rule, body):
            score += weight
            matched.append(f"正文:{rule}")
    if cfg["attachment_rule"] != "" and match_attachment_rule(cfg["attachment_rule"], email["attachment_count"]):
        score += cfg["attachment_weight"]
        matched.append(f"附件:{cfg['attachment_rule']}")
    if cfg["image_rule"] != "" and match_image_rule(cfg["image_rule"], email["has_image"]):
        score += cfg["image_weight"]
        matched.append(f"图片:{cfg['image_rule']}")

    text_all = (subject + "\n" + body).lower()
    if progress_name == "销售合同已开启" and is_sales_workflow_base_subject(subject):
        score += 100
        matched.append("强规则:销售合同流程根邮件")
    if progress_name == "BT09邮件已发送" and ("创建bt09" in text_all or "创建 bt09" in text_all or "请创建bt09" in text_all):
        score -= 12
        matched.append("排除:创建BT09")
    if progress_name == "预付款已到" and "bt09" in text_all:
        score -= 8
        matched.append("排除:BT09")
    return round(score, 1), matched


def review_result(score: int, matched: List[str], reason: str, domain_specific: bool = False) -> Dict[str, Any]:
    return {
        "best_progress": "待人工审核",
        "best_score": score,
        "threshold": 100,
        "matched": matched,
        "is_valid": True,
        "needs_review": True,
        "review_reason": reason,
        "domain_specific": domain_specific,
    }


def classify_by_workflow_rules(email: Dict[str, Any], sender_role_map: Dict[str, str]) -> Optional[Dict[str, Any]]:
    """Deterministic classifier for the six labeled workflow states.

    The order matters.  Later-state evidence such as SO/BT numbers and factory
    acknowledgements must beat older thread titles like "请开启...合同流程" so the
    front end reflects the latest actionable state of the project object.
    """
    evidence_level, evidence_score, evidence_reasons = strict_subject_evidence(email)
    if evidence_level == "ignore":
        return None

    subject = to_str(email.get("subject"))
    current = to_str(email.get("current_body"))
    full_body = to_str(email.get("body"))
    sender_name = to_str(email.get("sender_name"))
    all_text = subject + "\n" + current + "\n" + full_body

    def result(progress: str, score: int, matched: List[str]) -> Dict[str, Any]:
        return {
            "best_progress": progress,
            "best_score": score,
            "threshold": 100,
            "matched": evidence_reasons + matched,
            "is_valid": True,
            "needs_review": False,
            "review_reason": "",
        }

    if email.get("missing_root_subject"):
        return review_result(65, evidence_reasons + ["缺少原始流程开启邮件"], "只扫描到RE/FW后续邮件，未扫描到原始流程开启邮件，需人工审核。")

    # Latest-state rules first.
    if is_factory_bt_reply_text(subject, current):
        # Order Acknowledgement / OA 类邮件应归类为工厂反馈OA
        if is_oa_subject(subject) or "orderacknowledgement" in compact_text(current) or "orderreceipt" in compact_text(current) or "thankyouforplacingthisorder" in compact_text(current):
            return result("工厂反馈OA", 130, ["标准流程:工厂OA反馈", "标题/正文:Order Acknowledgement/Order receipt"])
        return result("工厂BT回复", 130, ["标准流程:工厂BT回复", "标题/正文:RFC/工厂下单回复"])

    if is_pm_bt09_request_text(subject, current):
        return result("PM开启BT09", 120, ["标准流程:PM开启BT09/下单", "正文:创建BT09/请下单/建中间号"])

    # iProcess/RARO/价格审批 check BEFORE PA check.
    # RARO mails with SO/BT numbers should NOT be classified as PA回复.
    if is_iprocess_approval_text(subject, current):
        subject_c = compact_text(subject)
        sender_email_l = to_str(email.get("sender_email", "")).lower()
        sender_name_c = compact_text(email.get("sender_name", ""))
        if ("salesorderapprovalrequest" in subject_c or "价格审批" in subject_c) and not ("icey" in sender_email_l or "bing" in sender_email_l or "icey" in sender_name_c):
            return review_result(70, evidence_reasons + ["标准流程:疑似iProcess/RARO但非icey-bing发送"], "iProcess/RARO/价格审批相关邮件非icey-bing chen发送，需人工确认。", domain_specific=True)
        return result("iProcess审批", 112, ["标准流程:iProcess审批", "正文:项目-客户-金额/iProcess/RARO"])

    if is_pa_so_bt09_reply_text(subject, current, sender_name):
        # 运营类往来邮件（发货明细、提货交期、调用现货、BOL issue、downgrade 等）
        # 即使含有 SO/BT 号也不代表 PA 正式回复，应人工审核
        subject_c = compact_text(subject)
        body_c = compact_text(current)
        operational_patterns = [
            "发货明细", "提货交期", "调用.*现货", "closingerror", "downgrade",
            "发货", "交期协调", "提货", "出货", "bol\\b", "bolissue",
            "仓库出货", "协调出货", "出货协调",
        ]
        is_operational = any(
            re.search(pat, (subject_c + "\n" + body_c), re.I)
            for pat in operational_patterns
        )
        if is_operational:
            return review_result(70, evidence_reasons + ["标准流程:疑似PA回复但为运营类往来邮件"], "该邮件为运营类往来（发货明细/提货交期/调用现货/BOL issue/downgrade等），非PA正式SO/BT09回复，需人工审核。", domain_specific=True)
        return result("PA回复SO/BT09", 125, ["标准流程:PA回复SO/BT09", "正文:SO/BT号或跨公司SO"])

    if is_book_request_text(subject, current):
        return result("Book订单申请", 116, ["标准流程:Book订单申请", "标题/正文:CHECK/book in SAP"])

    # The opening state is deliberately after later-state checks because many
    # replies keep the original "请开启...流程" title.
    if (is_sales_workflow_base_subject(subject) and not has_reply_forward_prefix(subject)) or is_sales_workflow_opening_body(current):
        return result("开启流程", 120, ["100分关键字:流程开启标题/正文"])

    if is_sales_workflow_base_subject(subject) and has_reply_forward_prefix(subject):
        return review_result(75, evidence_reasons, "RE/FW流程邮件未能从最新正文确定当前阶段，需人工审核。")

    if evidence_score >= 50:
        return review_result(evidence_score, evidence_reasons, "命中流程邮件格式，但当前邮件证据不足，需人工审核。")
    return None

def classify_email_progress(email: Dict[str, Any], patterns: Dict[str, Dict[str, Any]], sender_role_map: Dict[str, str]) -> Dict[str, Any]:
    deterministic = classify_by_workflow_rules(email, sender_role_map)
    if deterministic:
        return deterministic

    evidence_level, evidence_score, evidence_reasons = strict_subject_evidence(email)
    if evidence_level == "ignore":
        return {
            "best_progress": "",
            "best_score": evidence_score,
            "threshold": 100,
            "matched": evidence_reasons,
            "is_valid": False,
            "needs_review": False,
            "review_reason": "",
        }

    body_lower = safe_lower(email["current_body"])
    sender_lower = safe_lower(email["sender_email"])
    is_revised = "global-online.no-reply@abb.com" in sender_lower and "acknowledgement has been revised" in body_lower
    if is_revised and evidence_score >= 100:
        return {
            "best_progress": "工厂BT回复",
            "best_score": 130,
            "threshold": 100,
            "matched": evidence_reasons + ["特殊规则:revised_ack"],
            "is_valid": True,
            "needs_review": False,
            "review_reason": "",
        }

    best_progress = None
    best_score = -999.0
    best_threshold = 999.0
    best_matched: List[str] = []
    for progress_name, cfg in patterns.items():
        score, matched = score_progress(email, progress_name, cfg, sender_role_map)
        threshold = cfg["threshold"]
        if score > best_score:
            best_progress = progress_name
            best_score = score
            best_threshold = threshold
            best_matched = matched

    # Excel pattern is now only a fallback inside confirmed workflow threads.
    # It cannot pull unrelated contract emails into the normal stages.
    if evidence_score >= 100 and best_progress is not None and best_score >= best_threshold:
        return {
            "best_progress": best_progress,
            "best_score": round(best_score, 1),
            "threshold": best_threshold,
            "matched": evidence_reasons + best_matched,
            "is_valid": True,
            "needs_review": False,
            "review_reason": "",
        }

    return review_result(max(evidence_score, 50), evidence_reasons + best_matched[:3], "合同流程相关性不足100分，已转入人工审核。")


def mark_missing_root_subject(records: List[Dict[str, Any]]) -> int:
    """Mark RE/FW workflow-thread emails that lack a known sales-start root.

    Multi-scan rule: the original sales-start message does not have to be in
    the *current* date range if the same contract already exists in the local
    tracker from a previous successful sync. This allows a daily scan to pick
    up new RE/FW progress without forcing those messages into manual review.
    """
    root_keys = set()
    root_contracts = set()
    known_contracts = set()

    for record in records:
        for contract_id, _ctype in record.get("contracts") or []:
            if contract_id and project_db.get_project_by_contract(contract_id):
                known_contracts.add(contract_id)

        subject = record.get("subject", "")
        current_body = record.get("current_body", "")
        if (is_sales_workflow_base_subject(subject) and not has_reply_forward_prefix(subject)) or is_sales_workflow_opening_body(current_body):
            root_keys.add(compact_text(strip_reply_forward_prefix(subject)))
            for contract_id, _ctype in record.get("contracts") or []:
                root_contracts.add(contract_id)

    marked = 0
    for record in records:
        subject = record.get("subject", "")
        if not (is_workflow_subject_family(subject) and has_reply_forward_prefix(subject)):
            continue
        key = compact_text(strip_reply_forward_prefix(subject))
        contracts = {contract_id for contract_id, _ctype in (record.get("contracts") or [])}
        has_root = key in root_keys or bool(contracts & root_contracts) or bool(contracts & known_contracts)
        if not has_root:
            record["missing_root_subject"] = True
            marked += 1
        else:
            record["missing_root_subject"] = False
    return marked


def backfill_contracts_by_conversation(records: List[Dict[str, Any]]) -> int:
    conv_map: Dict[str, set] = defaultdict(set)
    topic_map: Dict[str, set] = defaultdict(set)
    for record in records:
        if not record["contracts"]:
            continue
        if record["conversation_id"]:
            for contract in record["contracts"]:
                conv_map[record["conversation_id"]].add(contract)
        if record["topic_key"]:
            for contract in record["contracts"]:
                topic_map[record["topic_key"]].add(contract)
    backfilled = 0
    for record in records:
        if record["contracts"]:
            continue
        assigned: List[Tuple[str, str]] = []
        # A single real project can now carry multiple aliases (CQ + M/K + SO).
        # Backfill all identifiers from the conversation/topic instead of
        # requiring exactly one tuple. The project-object resolver will collapse
        # them safely and mark conflicts when multiple existing projects match.
        if record["conversation_id"] and conv_map[record["conversation_id"]]:
            assigned = sort_contract_aliases(conv_map[record["conversation_id"]])
        elif record["topic_key"] and topic_map[record["topic_key"]]:
            assigned = sort_contract_aliases(topic_map[record["topic_key"]])
        if assigned:
            record["contracts"] = assigned
            record["contract_source"] = "conversation"
            backfilled += 1
    return backfilled


def assign_project_objects(records: List[Dict[str, Any]]) -> Dict[str, Dict[str, Any]]:
    """Assign each parsed mail to one project object key using aliases.

    This is the object-centric merge layer.  It collapses these cases:
    - opening mail: CQ1106414 only
    - forwarded/check mail: CQ1106414 + M4367-3569 + 505368900
    - later reply: SO or BT number only, backfilled by conversation/topic

    It does not merge two already-known projects silently.  If aliases point to
    multiple existing project ids, records are routed to the first existing id and
    the project is marked for human review later.
    """

    parent: Dict[str, str] = {}

    def find(x: str) -> str:
        parent.setdefault(x, x)
        if parent[x] != x:
            parent[x] = find(parent[x])
        return parent[x]

    def union(a: str, b: str) -> None:
        ra, rb = find(a), find(b)
        if ra != rb:
            parent[rb] = ra

    def alias_node(alias: Tuple[str, str]) -> str:
        alias_type, alias_value = normalize_contract_alias(alias[0], alias[1])
        return f"A:{alias_type}:{alias_value}"

    def project_node(project_id: str) -> str:
        return f"P:{project_id}"

    # First pass: union aliases found in the same record and connect them to any
    # existing alias-aware project in SQLite.
    for record in records:
        aliases = sort_contract_aliases(record.get("contracts") or [])
        record["contracts"] = aliases
        if not aliases:
            continue
        first_node = alias_node(aliases[0])
        for alias in aliases:
            node = alias_node(alias)
            union(first_node, node)
            existing = project_db.get_project_by_contract(alias[0])
            if existing:
                union(node, project_node(existing["id"]))

    groups: Dict[str, Dict[str, Any]] = defaultdict(lambda: {"aliases": [], "existing_project_ids": []})
    for record in records:
        aliases = sort_contract_aliases(record.get("contracts") or [])
        if not aliases:
            continue
        root = find(alias_node(aliases[0]))
        for alias in aliases:
            if alias not in groups[root]["aliases"]:
                groups[root]["aliases"].append(alias)
            existing = project_db.get_project_by_contract(alias[0])
            if existing and existing["id"] not in groups[root]["existing_project_ids"]:
                groups[root]["existing_project_ids"].append(existing["id"])

    group_infos: Dict[str, Dict[str, Any]] = {}
    for root, group in groups.items():
        aliases = sort_contract_aliases(group["aliases"])
        existing_ids = list(group["existing_project_ids"])
        primary_contract = choose_primary_contract(aliases)
        if existing_ids:
            project_key = existing_ids[0]
            conflict_ids = existing_ids[1:]
        else:
            project_key = primary_contract
            conflict_ids = []
        group_infos[root] = {
            "project_key": project_key,
            "primary_contract": primary_contract,
            "aliases": aliases,
            "db_aliases": aliases_for_db(aliases),
            "conflict_project_ids": conflict_ids,
        }

    for record in records:
        aliases = sort_contract_aliases(record.get("contracts") or [])
        if not aliases:
            continue
        info = group_infos.get(find(alias_node(aliases[0])))
        if not info:
            continue
        record["project_key"] = info["project_key"]
        record["primary_contract"] = info["primary_contract"]
        record["project_aliases"] = info["aliases"]
        record["project_db_aliases"] = info["db_aliases"]
        if info.get("conflict_project_ids"):
            record["project_conflict_ids"] = info["conflict_project_ids"]

    return group_infos


def apply_llm_double_check(
    email: Dict[str, Any],
    rule_classification: Dict[str, Any],
    llm_config: Dict[str, Any],
) -> Dict[str, Any]:
    """Use the local model as a second-pass filter.

    Rules provide the coarse candidate/stage. The model can confirm, downgrade
    to manual review, or ignore. Rule/model conflicts are not auto-resolved; they
    go to manual review, matching the user's requested safety policy.

    Domain-specific reviews (RARO非icey-bing, 运营类往来邮件) are NOT sent to the LLM;
    the domain rule decision is final and requires human review.
    """
    # Domain-specific review decisions skip LLM entirely — human review required.
    if rule_classification.get("domain_specific"):
        return rule_classification

    confirm_threshold = float(llm_config.get("confirm_threshold", 0.75))
    review_threshold = float(llm_config.get("review_threshold", 0.45))
    try:
        ai = llm_review.review_email(email, rule_classification, llm_config)
    except Exception as exc:
        fallback = review_result(50, rule_classification.get("matched", [])[:3], f"本地AI审核失败，需人工审核：{exc}")
        fallback["llm_decision"] = "review"
        fallback["llm_confidence"] = 0.0
        fallback["llm_reason"] = str(exc)
        return fallback

    decision = ai.get("decision") or "review"
    confidence = float(ai.get("confidence") or 0)
    ai_stage = to_str(ai.get("stage"))
    ai_reason = to_str(ai.get("reason"))
    ai_evidence = ai.get("evidence") or []
    if not isinstance(ai_evidence, list):
        ai_evidence = [to_str(ai_evidence)]

    matched = list(rule_classification.get("matched") or [])
    matched.append(f"AI:{decision}/{confidence:.2f}")
    if ai_reason:
        matched.append(f"AI原因:{ai_reason}")
    for ev in ai_evidence[:3]:
        if ev:
            matched.append(f"AI证据:{ev}")

    # Missing root subject is a hard manual-review gate even if the model thinks it is clear.
    if email.get("missing_root_subject"):
        result = review_result(max(int(confidence * 100), 65), matched, "只扫描到RE/FW后续邮件，未扫描到原始销售开启合同邮件；本地AI结果仅作参考。")
    elif decision == "ignored" or confidence < review_threshold:
        result = {
            "best_progress": "",
            "best_score": round(confidence * 100, 1),
            "threshold": 100,
            "matched": matched,
            "is_valid": False,
            "needs_review": False,
            "review_reason": ai_reason or "本地AI判定该邮件不属于标准合同流程。",
        }
    elif decision == "review" or confidence < confirm_threshold or not ai_stage:
        result = review_result(round(confidence * 100), matched, ai_reason or "本地AI认为证据不足，需人工审核。")
    else:
        rule_stage = to_str(rule_classification.get("best_progress"))
        rule_needs_review = bool(rule_classification.get("needs_review")) or rule_stage in {"待人工审核", "待人工确认", ""}
        if rule_stage and not rule_needs_review and ai_stage != rule_stage:
            result = review_result(round(confidence * 100), matched, f"规则判断为「{rule_stage}」，本地AI判断为「{ai_stage}」，阶段冲突需人工审核。")
        else:
            result = {
                "best_progress": ai_stage,
                "best_score": round(confidence * 100, 1),
                "threshold": 100,
                "matched": matched,
                "is_valid": True,
                "needs_review": False,
                "review_reason": "",
            }

    result["llm_decision"] = decision
    result["llm_confidence"] = confidence
    result["llm_reason"] = ai_reason
    result["llm_stage"] = ai_stage
    result["llm_company_name"] = to_str(ai.get("company_name"))
    result["llm_contract_number"] = to_str(ai.get("contract_number"))
    return result


def build_contract_timelines(
    records: List[Dict[str, Any]],
    patterns: Dict[str, Dict[str, Any]],
    sender_role_map: Dict[str, str],
    llm_config: Optional[Dict[str, Any]] = None,
    log: Optional[Callable[[str], None]] = None,
    cancel_check: Optional[Callable[[], bool]] = None,
) -> Tuple[Dict[str, List[Dict[str, Any]]], Dict[str, int]]:
    timelines: Dict[str, List[Dict[str, Any]]] = defaultdict(list)
    llm_stats = {"aiReviewed": 0, "aiConfirmed": 0, "aiReview": 0, "aiIgnored": 0, "aiFailed": 0, "aiSkipped": 0}
    use_llm = bool(llm_config and llm_config.get("enabled"))

    review_candidates = 0
    if use_llm:
        for record in records:
            preview = classify_email_progress(record, patterns, sender_role_map)
            if bool(preview.get("needs_review")):
                review_candidates += 1
        if log:
            log(f"规则可自动确认 {max(len(records) - review_candidates, 0)} 封；仅将 {review_candidates} 封需人工审核候选交给本地AI复核。")

    for index, record in enumerate(records, start=1):
        ensure_not_cancelled(cancel_check)
        rule_classification = classify_email_progress(record, patterns, sender_role_map)
        classification = rule_classification
        should_use_llm = use_llm and bool(rule_classification.get("needs_review"))
        if use_llm and not should_use_llm:
            llm_stats["aiSkipped"] += 1
        if should_use_llm:
            classification = apply_llm_double_check(record, rule_classification, llm_config or {})
            llm_stats["aiReviewed"] += 1
            decision = classification.get("llm_decision") or "review"
            if classification.get("is_valid") and not classification.get("needs_review"):
                llm_stats["aiConfirmed"] += 1
            elif classification.get("needs_review"):
                llm_stats["aiReview"] += 1
            elif decision == "ignored" or not classification.get("is_valid"):
                llm_stats["aiIgnored"] += 1
            if "本地AI审核失败" in to_str(classification.get("review_reason")):
                llm_stats["aiFailed"] += 1
            if log and (llm_stats["aiReviewed"] % 5 == 0 or llm_stats["aiReviewed"] == review_candidates):
                log(f"本地AI复核需人工审核候选：{llm_stats['aiReviewed']}/{review_candidates} 封。")

        record.update({
            "best_progress": classification["best_progress"],
            "best_score": classification["best_score"],
            "threshold": classification["threshold"],
            "matched": classification["matched"],
            "is_valid_progress": classification["is_valid"],
            "needs_review": bool(classification.get("needs_review")),
            "review_reason": classification.get("review_reason", ""),
            "llm_decision": classification.get("llm_decision", ""),
            "llm_confidence": classification.get("llm_confidence", ""),
            "llm_reason": classification.get("llm_reason", ""),
        })
        llm_company_name = to_str(classification.get("llm_company_name"))
        if llm_company_name:
            record["client_name"] = llm_company_name

        if not record.get("contracts"):
            continue
        # AI ignored candidates are not persisted into project timelines.
        if use_llm and not record.get("is_valid_progress") and not record.get("needs_review"):
            continue

        project_key = record.get("project_key") or record.get("primary_contract") or choose_primary_contract(record.get("contracts") or [])
        primary_contract = record.get("primary_contract") or choose_primary_contract(record.get("contracts") or [])
        primary_alias_type = alias_type_for_contract(primary_contract, "")

        # Revised acknowledgement rule historically guarded OCR duplicates. Keep
        # the guard, but do it at object level instead of creating one row per
        # identifier.
        if "特殊规则:revised_ack" in record["matched"] and primary_alias_type != "PROJECT_NUMBER":
            continue

        timelines[project_key].append({
            **record,
            "contract_id": primary_contract,
            "contract_type": "object",
            "project_key": project_key,
            "contract_aliases": record.get("project_aliases") or record.get("contracts") or [],
            "db_aliases": record.get("project_db_aliases") or aliases_for_db(record.get("contracts") or []),
        })
    return timelines, llm_stats


def resolve_current_progress(timelines: Dict[str, List[Dict[str, Any]]]) -> Dict[str, Dict[str, Any]]:
    resolved: Dict[str, Dict[str, Any]] = {}
    for project_key, emails in timelines.items():
        emails_sorted = sorted(emails, key=lambda x: x["received_time"] if x.get("received_time") else datetime.min, reverse=True)
        selected = None
        fallback_latest = emails_sorted[0] if emails_sorted else None
        for email in emails_sorted:
            if email.get("is_valid_progress"):
                selected = email
                break
        source = selected or fallback_latest
        if not source:
            continue
        progress = source.get("best_progress") if selected else "待人工确认"
        stage_dates: Dict[str, str] = {}
        all_aliases: List[Tuple[str, str]] = []
        events: List[Dict[str, Any]] = []
        conflict_ids: List[str] = []

        for email in emails_sorted:
            for alias in email.get("contract_aliases") or email.get("contracts") or []:
                if alias not in all_aliases:
                    all_aliases.append(alias)
            for conflict_id in email.get("project_conflict_ids") or []:
                if conflict_id not in conflict_ids:
                    conflict_ids.append(conflict_id)
            if email.get("is_valid_progress"):
                stage = PROGRESS_TO_STAGE.get(to_str(email.get("best_progress")))
                if stage:
                    current_date = date_only(email.get("received_time"))
                    if stage not in stage_dates or current_date > stage_dates[stage]:
                        stage_dates[stage] = current_date
            events.append({
                "event_type": event_type_from_progress(email.get("best_progress")),
                "event_time": email.get("received_time"),
                "sender": email.get("sender_email") or email.get("sender_name") or "",
                "subject": email.get("subject", ""),
                "source_email_id": email.get("entry_id", ""),
                "source_attachment_id": "",
                "confidence": to_float(email.get("best_score"), 0.0) / 100.0,
                "extracted_fields": {
                    "primary_contract": email.get("contract_id", ""),
                    "aliases": aliases_for_db(email.get("contract_aliases") or email.get("contracts") or []),
                    "progress": email.get("best_progress", ""),
                    "stage": PROGRESS_TO_STAGE.get(to_str(email.get("best_progress")), ""),
                    "score": email.get("best_score", ""),
                    "matched": email.get("matched", []),
                    "client_name": email.get("client_name", ""),
                    "folder": email.get("folder", ""),
                    "review_reason": email.get("review_reason", ""),
                },
            })

        all_aliases = sort_contract_aliases(all_aliases)
        primary_contract = source.get("contract_id") or choose_primary_contract(all_aliases) or project_key
        stage = PROGRESS_TO_STAGE.get(to_str(progress), REVIEW_STAGE_ID if source.get("needs_review") else "sales-contract")
        needs_review = bool(source.get("needs_review")) or stage == REVIEW_STAGE_ID or bool(conflict_ids)
        review_reason = source.get("review_reason", "")
        if conflict_ids:
            conflict_text = "同一组项目别名命中了多个已有项目：" + ", ".join(conflict_ids) + "；已保留第一个匹配项目，需人工确认是否合并。"
            review_reason = (review_reason + "；" + conflict_text).strip("；") if review_reason else conflict_text

        # Collect LLM review summary for this project from the source email.
        llm_reviewed = bool(source.get("llm_decision"))
        llm_summary = ""
        if llm_reviewed:
            llm_decision = source.get("llm_decision", "")
            llm_confidence = source.get("llm_confidence", "")
            llm_reason = source.get("llm_reason", "")
            llm_stage = source.get("llm_stage", "")
            parts = []
            if llm_decision:
                parts.append(f"AI判定: {llm_decision}")
            if llm_stage:
                parts.append(f"阶段: {llm_stage}")
            if isinstance(llm_confidence, (int, float)):
                parts.append(f"置信度: {round(float(llm_confidence) * 100)}%")
            if llm_reason:
                parts.append(f"原因: {llm_reason}")
            llm_summary = "；".join(parts)

        resolved[project_key] = {
            "project_key": project_key,
            "contract_id": primary_contract,
            "contract_type": source.get("contract_type", "object"),
            "aliases": aliases_for_db(all_aliases),
            "raw_aliases": all_aliases,
            "events": events,
            "date": source.get("received_time"),
            "progress": progress,
            "stage": stage,
            "score": source.get("best_score"),
            "threshold": source.get("threshold"),
            "subject": source.get("subject", ""),
            "sender": source.get("sender_email", ""),
            "email_entry_id": source.get("entry_id", ""),
            "email_store_id": source.get("store_id", ""),
            "email_folder": source.get("folder", ""),
            "matched": "; ".join(source.get("matched", [])),
            "contract_source": source.get("contract_source", ""),
            "client_name": source.get("client_name", ""),
            "attachment_dir": source.get("attachment_dir", ""),
            "stageDates": stage_dates,
            "needs_review": needs_review,
            "review_reason": review_reason,
            "conflict_project_ids": conflict_ids,
            "llmReviewed": llm_reviewed,
            "llmSummary": llm_summary,
        }
    return resolved

def apply_sync_to_projects(resolved_contracts: Dict[str, Dict[str, Any]]) -> Dict[str, int]:
    """Merge sync results into project objects without corrupting history.

    Object-centric policy:
    - ``project_key`` is the internal object id when an existing alias resolves,
      otherwise the best visible identifier in the alias group.
    - ``contract`` remains the front-end display field and may upgrade from CQ to
      M/K when that stronger alias appears.
    - All identifiers are saved as aliases, so future CQ/M/SO/BT mails attach to
      the same object.
    - Manual edits still win over email classification.
    """
    counts = {"created": 0, "updated": 0, "review": 0, "skipped": 0}
    for object_key, info in resolved_contracts.items():
        contract_id = info.get("contract_id") or object_key
        aliases = info.get("aliases") or []

        existing = project_db.get_project(str(info.get("project_key") or object_key))
        if not existing:
            existing, alias_conflicts = project_db.get_project_by_any_alias(aliases or [contract_id])
        else:
            alias_conflicts = []
        if not existing:
            existing = project_db.get_project_by_contract(contract_id)
            alias_conflicts = []

        new_stage = info["stage"]
        existing_stage = existing.get("stage") if existing else ""
        existing_index = stage_index(existing_stage)
        new_index = stage_index(new_stage)
        needs_review = bool(info.get("needs_review")) or new_stage == REVIEW_STAGE_ID
        review_reason = info.get("review_reason", "") or ("需要人工审核。" if needs_review else "")
        if alias_conflicts:
            needs_review = True
            conflict_text = "项目别名与其他项目冲突：" + ", ".join(alias_conflicts)
            review_reason = (review_reason + "；" + conflict_text).strip("；") if review_reason else conflict_text

        new_email_dt = coerce_datetime(info.get("date"))
        existing_latest_dt = coerce_datetime(existing.get("latestEmailTime")) if existing else None
        new_email_is_newer = bool(new_email_dt and (not existing_latest_dt or new_email_dt > existing_latest_dt))
        new_email_is_same_or_newer = bool(new_email_dt and (not existing_latest_dt or new_email_dt >= existing_latest_dt))

        stage_to_save = new_stage
        update_stage = True
        update_latest_email_fields = True
        stale_scan_result = False

        if existing:
            stage_to_save = existing_stage or new_stage

            if existing.get("manualOverride"):
                stage_to_save = existing_stage or new_stage
                update_stage = False
                update_latest_email_fields = new_email_is_newer
                if new_stage != existing_stage:
                    needs_review = True
                    review_reason = "邮件识别阶段与人工阶段不一致；已保留人工修改。"
            elif new_email_is_newer or not existing_latest_dt:
                # User-facing policy: the card always shows the latest valid
                # workflow email, not the highest historical stage.  Older code
                # blocked regressions; that hid newer real events in old threads.
                stage_to_save = new_stage
                update_stage = True
                update_latest_email_fields = True
            elif new_email_is_same_or_newer and new_index > existing_index:
                stage_to_save = new_stage
                update_stage = True
                update_latest_email_fields = True
            else:
                stage_to_save = existing_stage or new_stage
                update_stage = False
                update_latest_email_fields = False
                stale_scan_result = True

            if stale_scan_result and not needs_review:
                existing_stage_dates = existing.get("stageDates", {}) if existing else {}
                merged_stage_dates = {**existing_stage_dates, **(info.get("stageDates") or {})}
                if merged_stage_dates != existing_stage_dates:
                    existing["stageDates"] = merged_stage_dates
                    project_db.upsert_project(existing)
                # Even stale rescans can reveal aliases/events; keep those.
                project_db.save_project_aliases(existing["id"], aliases, confidence=to_float(info.get("score"), 0.0) / 100.0)
                for event in info.get("events") or []:
                    project_db.record_project_event(
                        existing["id"],
                        event_type=event.get("event_type", "EMAIL_EVIDENCE"),
                        event_time=event.get("event_time", ""),
                        sender=event.get("sender", ""),
                        subject=event.get("subject", ""),
                        source_email_id=event.get("source_email_id", ""),
                        source_attachment_id=event.get("source_attachment_id", ""),
                        extracted_fields=event.get("extracted_fields", {}),
                        confidence=event.get("confidence", 0.0),
                    )
                counts["skipped"] += 1
                continue

        existing_stage_dates = existing.get("stageDates", {}) if existing else {}
        stage_dates = {**existing_stage_dates, **(info.get("stageDates") or {})}
        if stage_to_save not in stage_dates:
            stage_dates[stage_to_save] = date_only(info.get("date")) or date_only(datetime.now())

        latest_email_time = format_dt(info.get("date"))
        keep_existing_latest = existing is not None and not update_latest_email_fields
        existing_needs_review = bool(existing.get("needsReview")) if existing else False

        project_id = existing.get("id") if existing else (info.get("project_key") or object_key or contract_id)
        project = {
            "id": project_id,
            "contract": existing.get("contract") if existing and (existing.get("manualOverride") or keep_existing_latest) and existing.get("contract") else contract_id,
            "name": existing.get("name") if existing and (existing.get("manualOverride") or keep_existing_latest) else info.get("subject") or contract_id,
            "client": existing.get("client") if existing and (existing.get("manualOverride") or keep_existing_latest) and existing.get("client") else info.get("client_name") or (existing.get("client") if existing else ""),
            "amount": existing.get("amount") if existing else "",
            "type": existing.get("type") if existing else "standard",
            "stage": stage_to_save,
            "date": date_only(info.get("date")) if (not existing or update_stage or update_latest_email_fields) else (existing.get("date") if existing else date_only(datetime.now())),
            "notes": existing.get("notes") if existing and (existing.get("manualOverride") or keep_existing_latest) else f"邮件识别：{info.get('progress', '')}",
            "favorite": existing.get("favorite") if existing else False,
            "suspended": existing.get("suspended") if existing else False,
            "archived": existing.get("archived") if existing else False,
            "archivedAt": existing.get("archivedAt") if existing else "",
            "archivedFromStage": existing.get("archivedFromStage") if existing else "",
            "stageDates": stage_dates,
            "currentProgress": existing.get("currentProgress") if keep_existing_latest else info.get("progress", ""),
            "latestEmailTime": existing.get("latestEmailTime") if keep_existing_latest else latest_email_time,
            "latestEmailSubject": existing.get("latestEmailSubject") if keep_existing_latest else info.get("subject", ""),
            "latestSender": existing.get("latestSender") if keep_existing_latest else info.get("sender", ""),
            "needsReview": needs_review or existing_needs_review,
            "reviewReason": review_reason or (existing.get("reviewReason") if existing else ""),
            "manualOverride": existing.get("manualOverride") if existing else False,
            "source": "outlook-sync",
            "latestAttachmentDir": existing.get("latestAttachmentDir") if keep_existing_latest else info.get("attachment_dir", ""),
            "latestEmailEntryId": existing.get("latestEmailEntryId") if keep_existing_latest else info.get("email_entry_id", ""),
            "latestEmailStoreId": existing.get("latestEmailStoreId") if keep_existing_latest else info.get("email_store_id", ""),
            "latestEmailFolder": existing.get("latestEmailFolder") if keep_existing_latest else info.get("email_folder", ""),
            "llmReviewed": info.get("llmReviewed", False) if update_latest_email_fields else (existing.get("llmReviewed") if existing else False),
            "llmSummary": info.get("llmSummary", "") if update_latest_email_fields else (existing.get("llmSummary") if existing else ""),
            "createdAt": existing.get("createdAt") if existing else "",
        }
        saved_project = project_db.upsert_project(project)
        saved_project_id = saved_project.get("id") or project_id

        alias_conflicts = project_db.save_project_aliases(
            saved_project_id,
            aliases or [{"alias_type": project_db.infer_alias_type(contract_id), "alias_value": contract_id}],
            confidence=to_float(info.get("score"), 0.0) / 100.0,
        )
        event_ids: List[str] = []
        for event in info.get("events") or []:
            event_id = project_db.record_project_event(
                saved_project_id,
                event_type=event.get("event_type", "EMAIL_EVIDENCE"),
                event_time=event.get("event_time", ""),
                sender=event.get("sender", ""),
                subject=event.get("subject", ""),
                source_email_id=event.get("source_email_id", ""),
                source_attachment_id=event.get("source_attachment_id", ""),
                extracted_fields=event.get("extracted_fields", {}),
                confidence=event.get("confidence", 0.0),
            )
            if event_id:
                event_ids.append(event_id)

        if alias_conflicts:
            saved_project["needsReview"] = True
            conflict_text = "项目别名保存时发现冲突：" + ", ".join(alias_conflicts)
            saved_project["reviewReason"] = (to_str(saved_project.get("reviewReason")) + "；" + conflict_text).strip("；")
            project_db.upsert_project(saved_project)
            needs_review = True

        if not existing:
            counts["created"] += 1
            if saved_project.get("needsReview") or needs_review:
                counts["review"] += 1
        elif needs_review:
            counts["review"] += 1
        else:
            counts["updated"] += 1
    return counts


def persist_synced_mail(records: List[Dict[str, Any]]) -> None:
    for email in records:
        contract = email.get("primary_contract") or choose_primary_contract(email.get("contracts") or [])
        project_id = email.get("project_key") or contract
        project_db.save_email_record({
            "entry_id": email.get("entry_id"),
            "store_id": email.get("store_id"),
            "project_id": project_id,
            "contract": contract,
            "subject": email.get("subject"),
            "sender_name": email.get("sender_name"),
            "sender_email": email.get("sender_email"),
            "to_recipients": email.get("to_recipients"),
            "cc_recipients": email.get("cc_recipients"),
            "received_time": email.get("received_time_text"),
            "sent_time": email.get("sent_time_text"),
            "body": email.get("body"),
            "html_body": email.get("html_body"),
            "folder": email.get("folder"),
            "attachment_dir": email.get("attachment_dir"),
            "has_attachments": bool(email.get("attachment_records")),
            "is_read": email.get("is_read"),
            "importance": email.get("importance"),
            "categories": email.get("categories"),
        })
    attachments: List[Dict[str, Any]] = []
    for email in records:
        contract = email.get("primary_contract") or choose_primary_contract(email.get("contracts") or [])
        project_id = email.get("project_key") or contract
        for attachment in email.get("attachment_records") or []:
            attachment = dict(attachment)
            attachment["project_id"] = project_id
            attachment["contract"] = contract or attachment.get("contract", "")
            attachments.append(attachment)
    if attachments:
        project_db.save_attachment_records(attachments)


def sync_outlook(
    payload: Dict[str, Any],
    log: Optional[Callable[[str], None]] = None,
    cancel_check: Optional[Callable[[], bool]] = None,
) -> Dict[str, Any]:
    saved_attachment_dirs: List[str] = []

    def cleanup_current_run() -> None:
        cleanup_paths(saved_attachment_dirs)

    def check_cancel() -> None:
        ensure_not_cancelled(cancel_check, cleanup_current_run)

    def emit(message: str) -> None:
        if log:
            log(message)

    try:
        mailbox = to_str(payload.get("mailbox"))
        folder_path = to_str(payload.get("folder_path") or payload.get("folderPath"))
        include_subfolders = bool(payload.get("include_subfolders", payload.get("includeSubfolders", True)))
        start_dt = parse_date(to_str(payload.get("start_date") or payload.get("startDate")))
        end_dt = parse_date(to_str(payload.get("end_date") or payload.get("endDate")), end_of_day=True)
        if end_dt < start_dt:
            raise ValueError("结束日期不能早于开始日期。")

        project_db.init_db()
        emit("本次执行完整扫描：系统会扫描你选择的整个时间段，不跳过历史区间。")
        check_cancel()
        emit("读取 Email Pattern 配置...")
        patterns, sender_role_map = load_config_from_pattern_excel(DEFAULT_PATTERN_PATH)
        emit(f"规则加载完成：{len(patterns)} 个进度规则；流程判断不再使用发件人作为依据。")

        use_ai_review = bool(payload.get("use_ai_review", payload.get("useAiReview", True)))
        llm_config: Optional[Dict[str, Any]] = None
        if use_ai_review:
            llm_config = llm_review.load_llm_config()
            llm_config["enabled"] = True
            emit(f"启用AI二次审核：{llm_config.get('provider')} / {llm_config.get('model')} @ {llm_config.get('base_url')}。")
            health = llm_review.health_check(llm_config)
            if not health.get("ok"):
                raise RuntimeError(f"AI服务不可用：{health.get('error')}。请确认 API 地址和 api_key 配置正确。")
            emit("AI服务连接正常。")
        else:
            emit("未启用AI二次审核，本次仅使用规则识别。")

        check_cancel()
        emit("连接 Windows Outlook...")
        outlook = connect_outlook()
        namespace = outlook.GetNamespace("MAPI")

        check_cancel()
        emit(f"定位文件夹：{mailbox} / {folder_path or '(根目录)'}")
        target_folder = get_folder_by_root_path(namespace, mailbox, folder_path)

        emit(f"读取邮件：{start_dt.strftime('%Y-%m-%d')} 至 {end_dt.strftime('%Y-%m-%d')}。")
        records: List[Dict[str, Any]] = []
        raw_count = 0
        candidate_count = 0
        ignored_count = 0
        folder_count = 0

        for folder, folder_name in iter_folders(target_folder, include_subfolders=include_subfolders):
            check_cancel()
            folder_count += 1
            folder_total = get_folder_item_count(folder)
            before_raw = raw_count
            if folder_total >= 0:
                emit(f"扫描文件夹：{folder_name}（总项目 {folder_total}）")
            else:
                emit(f"扫描文件夹：{folder_name}")
            for item in get_items_in_date_range(folder, start_dt, end_dt):
                check_cancel()
                raw_count += 1
                email = parse_email_item(item, folder_name)
                if not email:
                    continue
                candidate, candidate_score, candidate_reasons = is_candidate_email(email, sender_role_map)
                if not candidate:
                    ignored_count += 1
                    continue
                email["candidate_score"] = candidate_score
                email["candidate_reasons"] = candidate_reasons
                ensure_tracking_alias(email)

                # Save candidate attachments only. They are deleted if the job is cancelled/failed.
                attachment_dir, attachment_records = save_attachments(item, email.get("contracts") or [], email.get("received_time"))
                email["attachment_dir"] = attachment_dir
                email["attachment_records"] = attachment_records
                if attachment_dir:
                    saved_attachment_dirs.append(attachment_dir)

                records.append(email)
                candidate_count += 1
                if candidate_count % 20 == 0:
                    emit(f"已识别流程候选邮件：{candidate_count} 封。")
            folder_raw = raw_count - before_raw
            if folder_raw == 0 and folder_total > 0:
                emit("该文件夹存在邮件，但没有邮件落在本次日期范围内。")

        emit(f"邮件抓取完成：扫描 {folder_count} 个文件夹，读取 {raw_count} 封邮件，流程候选 {candidate_count} 封，忽略 {ignored_count} 封。")
        if not records:
            return {
                "ok": True,
                "message": "未找到符合严格流程条件的邮件。",
                "counts": {"folders": folder_count, "rawEmails": raw_count, "candidateEmails": 0, "ignoredEmails": ignored_count, "contracts": 0, "created": 0, "updated": 0, "review": 0},
            }

        check_cancel()
        backfilled = backfill_contracts_by_conversation(records)
        root_missing = mark_missing_root_subject(records)
        object_groups = assign_project_objects(records)
        emit(f"合同号回填完成：{backfilled} 封邮件通过会话回填；{root_missing} 封RE/FW缺少原始销售开启邮件，转入人工审核。")
        emit(f"项目对象归并完成：{len(object_groups)} 个项目对象；CQ/M/SO/BT 将作为别名持续跟踪。")

        check_cancel()
        emit("建立项目对象时间线并判定最新进度...")
        timelines, llm_stats = build_contract_timelines(
            records,
            patterns,
            sender_role_map,
            llm_config=llm_config,
            log=emit,
            cancel_check=cancel_check,
        )
        resolved = resolve_current_progress(timelines)
        emit(f"项目进度判定完成：{len(resolved)} 个项目对象。")

        # Do not keep attachments or email DB rows for candidates that the local AI rejected.
        records_to_persist = [r for r in records if r.get("is_valid_progress") or r.get("needs_review")]
        ai_ignored_records = [r for r in records if r not in records_to_persist]
        ignored_attachment_dirs = [r.get("attachment_dir") for r in ai_ignored_records if r.get("attachment_dir")]
        if ignored_attachment_dirs:
            cleanup_paths(ignored_attachment_dirs)
            saved_attachment_dirs = [p for p in saved_attachment_dirs if p not in ignored_attachment_dirs]

        check_cancel()
        emit("准备写入项目库...")
        persist_synced_mail(records_to_persist)
        counts = apply_sync_to_projects(resolved)
        emit(f"写入项目库完成：新增 {counts['created']}，更新 {counts['updated']}，需人工审核 {counts['review']}。")

        return {
            "ok": True,
            "message": "同步完成。",
            "counts": {
                "folders": folder_count,
                "rawEmails": raw_count,
                "candidateEmails": candidate_count,
                "ignoredEmails": ignored_count + len(ai_ignored_records),
                "backfilled": backfilled,
                "rootMissing": root_missing,
                "contracts": len(resolved),
                **llm_stats,
                **counts,
            },
        }
    except SyncCancelled:
        cleanup_current_run()
        raise
    except Exception:
        cleanup_current_run()
        raise


def default_date_range() -> Dict[str, str]:
    today = datetime.now().date()
    return {
        "startDate": (today - timedelta(days=7)).strftime("%Y-%m-%d"),
        "endDate": today.strftime("%Y-%m-%d"),
    }