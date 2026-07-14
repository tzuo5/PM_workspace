# -*- coding: utf-8 -*-
"""Read-only Contract / CQP / TA comparison with traceable PDF evidence.

The source PDFs are never edited. Every review item carries its source page and,
where possible, exact word rectangles so the browser can navigate and highlight
what a reviewer must compare.
"""

from __future__ import annotations

import csv
import os
import re
import unicodedata
from difflib import SequenceMatcher
from collections import defaultdict
from dataclasses import dataclass
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple

from services.contract_llm_review import run_llm_contract_review
from services.pdf_evidence import (
    ParsedPDF,
    ParsedPage,
    evidence_entry,
    locate_all_evidence,
    locate_evidence,
    match_key,
    normalize_for_match,
    parse_pdf_with_evidence,
)

CATEGORY_CUSTOMER = "customer_information"
CATEGORY_PRODUCT = "product_information"
CATEGORY_OTHER = "other_information"

CATEGORY_LABELS = {
    CATEGORY_CUSTOMER: "客户信息",
    CATEGORY_PRODUCT: "产品信息",
    CATEGORY_OTHER: "其他信息",
}

SELLER_PREFIX_MAP = {
    "M": "ABB（上海）机器人投资有限公司",
    "K": "ABB机器人（珠海）有限公司",
}

MODEL_PATTERNS: List[Tuple[str, re.Pattern[str]]] = [
    ("IRB 1200-7/0.7 Gen2", re.compile(r"IRB\s*1200\s*-\s*7\s*/\s*0[.]7\s*Gen\s*2", re.I)),
    ("IRB 1200-7/0.9 LPS", re.compile(r"IRB\s*1200\s*-\s*7\s*/\s*0[.]9\s*(?:LPS|Lite\s*[+＋])", re.I)),
    ("IRB 1100-4/0.58", re.compile(r"IRB\s*1100\s*-\s*4\s*/\s*0[.]58", re.I)),
]

GENERIC_MODEL_PATTERN = re.compile(
    r"\bIRB\s*\d{3,4}(?:\s*-\s*[\d.]+\s*/\s*[\d.]+)?"
    r"(?:\s*(?:Gen\s*\d+|LPS|Lite\s*[+＋]?))?\b",
    re.I,
)

CHECKED_MARKERS = ("☒", "☑", "■", "●", "✓", "√", "[x]", "[X]")
UNCHECKED_MARKERS = ("☐", "□", "○", "[ ]")
AMOUNT_ROUNDING_TOLERANCE = 1.0
EXPECTED_CN_ROBOT_VAT = 0.13

# CQP configuration codes that are commercial metadata and are not required to
# be repeated in the technical agreement.
CQP_ONLY_ALLOWED_CODES = {"448-125"}

CONFIG_FEATURES: Dict[str, Sequence[str]] = {
    "3300-122": ("IRB 1200-7/0.7 Gen2",),
    "3300-121": ("IRB 1200-7/0.9 LPS", "IRB 1200-7/0.9 Lite+"),
    "3300-2": ("IRB 1100-4/0.58",),
    "209-202": ("ABB Graphite White", "标准石墨白"),
    "3350-400": ("Base 40", "IP40"),
    "3309-2": ("From side of base", "底座侧面出线"),
    "3000-105": ("OmniCore E10",),
    "3004-1": ("Max 45deg", "45°C", "45deg"),
    "3007-1": ("220-230 V AC", "220-230V AC"),
    "3007-2": ("110-230 V AC", "110-230V AC"),
    "3013-4": ("Embedded wired WAN", "连接服务"),
    "3016-1": ("FlexPendant 3m", "附带3 米电缆", "附带 3 米电缆"),
    "3043-11": ("SafeMove Standard",),
    "3044-1": ("3 modes Keyless", "3 档模式开关"),
    "3200-2": ("Length: 7m", "电缆长度7m", "电缆长度 7m"),
    "3203-5": ("CN mains cable, 3m",),
    "3120-2": ("Essential app package", "基础功能包"),
    "3151-1": ("Program package", "独立应用程序包"),
    "3303-1": ("Parallel & Air",),
    "438-1": ("Standard Warranty", "标准质保"),
    "438-102": ("Lite Warranty", "Lite 标准质保"),
}


@dataclass
class DocumentSet:
    contract_physical: Optional[ParsedPDF]
    contract_body: Optional[ParsedPDF]
    cqp: Optional[ParsedPDF]
    ta: Optional[ParsedPDF]
    ta_embedded: bool = False


def parse_pdf(filepath: str) -> ParsedPDF:
    return parse_pdf_with_evidence(filepath)


def _slice_pdf(parsed: ParsedPDF, pages: Iterable[ParsedPage], suffix: str = "") -> ParsedPDF:
    selected = list(pages)
    return ParsedPDF(
        filepath=parsed.filepath + suffix,
        pages=selected,
        full_text="\n".join(page.text for page in selected),
    )


def _find_ta_start(parsed: ParsedPDF) -> Optional[int]:
    """Return the physical page where the embedded TA actually begins.

    A contract table of contents and the separator page also contain the words
    “技术协议”, so generic keyword matching splits the contract too early.  A
    TA start must carry a strong cover/header marker.
    """
    for page in parsed.pages:
        key = match_key(page.text)
        strong_marker = (
            "technicalagreement" in key
            or "技术协议书" in key
            or "docno.3.02.f03" in key
            or "docno3.02.f03" in key
        )
        if strong_marker:
            return page.page_num
    return None


def _resolve_documents(
    parsed_by_path: Dict[str, ParsedPDF],
    file_roles: Optional[Dict[str, str]],
) -> DocumentSet:
    contract: Optional[ParsedPDF] = None
    cqp: Optional[ParsedPDF] = None
    ta: Optional[ParsedPDF] = None

    if file_roles:
        for role, path in file_roles.items():
            parsed = parsed_by_path.get(os.path.abspath(path))
            if role == "contract":
                contract = parsed
            elif role == "cqp":
                cqp = parsed
            elif role == "ta":
                ta = parsed

    # Content-based fallback, used only when a named role was not supplied.
    for parsed in parsed_by_path.values():
        text_key = match_key(parsed.full_text)
        if contract is None and "销售合同" in parsed.full_text and re.search(r"[MK]\s*\d{4}\s*-\s*\d{4}", parsed.full_text):
            contract = parsed
        if cqp is None and re.search(r"CQ\d{7}", parsed.full_text) and ("报价" in parsed.full_text or "quotation" in text_key):
            cqp = parsed
        if ta is None and "技术协议" in parsed.full_text and "销售合同" not in parsed.full_text:
            ta = parsed

    contract_body = contract
    ta_embedded = False
    if contract:
        ta_start = _find_ta_start(contract)
        if ta_start:
            body_pages = [page for page in contract.pages if page.page_num < ta_start]
            ta_pages = [page for page in contract.pages if page.page_num >= ta_start]
            contract_body = _slice_pdf(contract, body_pages, "#contract")
            if ta is None and ta_pages:
                ta = _slice_pdf(contract, ta_pages, "#ta")
                ta_embedded = True

    return DocumentSet(contract, contract_body, cqp, ta, ta_embedded)


# ---------------------------------------------------------------------------
# Text helpers
# ---------------------------------------------------------------------------


def _page_text(parsed: Optional[ParsedPDF], page_num: int) -> str:
    if parsed is None:
        return ""
    page = next((item for item in parsed.pages if item.page_num == page_num), None)
    return page.text if page else ""


def _clean_inline(value: str) -> str:
    value = unicodedata.normalize("NFKC", value or "")
    value = re.sub(r"[\uf000-\uf8ff]", "", value)
    value = re.sub(r"\s+", " ", value).strip(" _\t\r\n")
    return value


def _clean_entity(value: str) -> str:
    value = _clean_inline(value)
    value = re.sub(r"\s+", "", value)
    return value.strip("_：:")


def _compact(value: str) -> str:
    return re.sub(r"[^a-z0-9\u3400-\u9fff]+", "", normalize_for_match(value))


def _contains_equivalent(left: str, right: str) -> bool:
    lkey, rkey = _compact(left), _compact(right)
    return bool(lkey and rkey and (lkey == rkey or lkey in rkey or rkey in lkey))


def _match_first(text: str, patterns: Sequence[str], flags: int = re.I | re.S) -> str:
    for pattern in patterns:
        match = re.search(pattern, text, flags)
        if match:
            return _clean_inline(match.group(1))
    return ""


def _first_line_value(text: str, label: str) -> str:
    """Read the value on the same visual text line as ``label``.

    CQP pages often contain two columns on one extracted line.  The helper
    returns only the text between the requested label and the next known field
    label, preventing customer data from swallowing the ABB seller column.
    """
    stop_labels = (
        "ABB 公司名称", "ABB单位名称", "ABB 单位名称", "地址 / 街道", "地址/街道",
        "联络人", "电话号码", "联系人电子邮箱", "电子邮件", "邮政编码", "城市",
        "国家或地区", "项目名称", "报价编号", "报价单编号", "日期", "报价修订版本",
        "初始编号", "询价日期", "报价日期", "负责人", "页码",
    )
    for raw_line in (text or "").splitlines():
        line = _clean_inline(raw_line)
        pos = line.find(label)
        if pos < 0:
            continue
        value = line[pos + len(label):].lstrip(" :：")
        cut = len(value)
        for stop in stop_labels:
            idx = value.find(stop)
            if idx > 0:
                cut = min(cut, idx)
        value = value[:cut].strip()
        if value:
            return value
    return ""


def _parse_money(value: str) -> float:
    try:
        return float(value.replace(",", "").strip())
    except (TypeError, ValueError):
        return 0.0


def _canonical_model(value: str) -> str:
    normalized = unicodedata.normalize("NFKC", value or "")
    for canonical, pattern in MODEL_PATTERNS:
        if pattern.search(normalized):
            return canonical
    match = GENERIC_MODEL_PATTERN.search(normalized)
    if not match:
        return _clean_inline(value)
    raw = _clean_inline(match.group(0))
    parsed = re.match(
        r"IRB\s*(\d{3,4})(?:\s*-\s*([\d.]+)\s*/\s*([\d.]+))?(?:\s*(.*))?$",
        raw,
        re.I,
    )
    if not parsed:
        return raw
    base = f"IRB {parsed.group(1)}"
    if parsed.group(2) and parsed.group(3):
        base += f"-{parsed.group(2)}/{parsed.group(3)}"
    suffix = _clean_inline(parsed.group(4) or "")
    if re.search(r"Lite\s*[+＋]?", suffix, re.I):
        suffix = "LPS"
    elif re.search(r"Gen\s*(\d+)", suffix, re.I):
        suffix = "Gen" + re.search(r"Gen\s*(\d+)", suffix, re.I).group(1)
    elif suffix.upper() == "LPS":
        suffix = "LPS"
    return f"{base} {suffix}".strip()


def _models_in_text(text: str) -> List[str]:
    found: List[str] = []
    for match in GENERIC_MODEL_PATTERN.finditer(unicodedata.normalize("NFKC", text or "")):
        model = _canonical_model(match.group(0))
        if model and model not in found:
            found.append(model)
    return found


def _extract_model_quantities(parsed: Optional[ParsedPDF]) -> Dict[str, int]:
    if parsed is None:
        return {}
    candidates: Dict[str, List[int]] = defaultdict(list)
    for page in parsed.pages:
        lines = [unicodedata.normalize("NFKC", line).strip() for line in page.text.splitlines()]
        for index, line in enumerate(lines):
            for model_match in GENERIC_MODEL_PATTERN.finditer(line):
                model = _canonical_model(model_match.group(0))
                before = line[:model_match.start()]
                after = line[model_match.end():]
                before_matches = re.findall(r"(\d+)\s*(?:台|套|pcs?|units?)\s*$", before, re.I)
                if before_matches:
                    candidates[model].append(int(before_matches[-1]))
                after_numbers = re.findall(r"(?<![\d.])(\d+)(?![\d.])", after)
                if after_numbers:
                    number = int(after_numbers[0])
                    if 0 < number < 1000:
                        candidates[model].append(number)
                for nearby in lines[index + 1:index + 4]:
                    unit_match = re.search(r"(?:数量[:：]?\s*)?(\d+)\s*(?:台|套|pcs?|units?)\b", nearby, re.I)
                    if unit_match:
                        candidates[model].append(int(unit_match.group(1)))
                        break
                    if re.fullmatch(r"\d+", nearby):
                        number = int(nearby)
                        if 0 < number < 1000:
                            candidates[model].append(number)
                            break
    result: Dict[str, int] = {}
    for model, values in candidates.items():
        sensible = [value for value in values if 0 < value < 1000]
        if sensible:
            result[model] = max(set(sensible), key=lambda value: (sensible.count(value), value))
    return result


def _extract_cqp_products(parsed: Optional[ParsedPDF]) -> List[Dict[str, Any]]:
    """Extract CQP product rows without assuming a fixed robot model list."""
    if parsed is None:
        return []
    rows: List[Dict[str, Any]] = []
    for page in parsed.pages:
        lines = [unicodedata.normalize("NFKC", line).strip() for line in page.text.splitlines()]
        for index, line in enumerate(lines):
            for model_match in GENERIC_MODEL_PATTERN.finditer(line):
                model = _canonical_model(model_match.group(0))
                tail = line[model_match.end():]
                numbers = re.findall(r"(?<![A-Za-z0-9])([\d,]+[.]\d+|\d+)(?![A-Za-z0-9])", tail)
                if len(numbers) < 3:
                    continue
                qty = int(numbers[0].replace(",", ""))
                if not 0 < qty < 1000:
                    continue
                item_code = ""
                for nearby in lines[index:index + 4]:
                    code_match = re.search(r"(3HAC[\d-]+)", nearby, re.I)
                    if code_match:
                        item_code = code_match.group(1).upper()
                        break
                rows.append({
                    "model": model,
                    "item_code": item_code,
                    "qty": qty,
                    "unit_price": _parse_money(numbers[1]),
                    "line_total": _parse_money(numbers[2]),
                    "page": page.page_num,
                })
    if not rows:
        quantities = _extract_model_quantities(parsed)
        return [
            {"model": model, "qty": qty, "item_code": "", "unit_price": 0.0, "line_total": 0.0, "page": 0}
            for model, qty in quantities.items()
        ]

    merged: Dict[str, Dict[str, Any]] = {}
    for row in rows:
        model = row["model"]
        if model not in merged:
            merged[model] = dict(row)
            continue
        existing = merged[model]
        existing["qty"] += row["qty"]
        existing["line_total"] = round(existing.get("line_total", 0.0) + row.get("line_total", 0.0), 2)
        if not existing.get("item_code"):
            existing["item_code"] = row.get("item_code", "")
    return list(merged.values())


def _extract_delivery_schedule(parsed: Optional[ParsedPDF]) -> List[Dict[str, Any]]:
    if parsed is None:
        return []
    schedule: List[Dict[str, Any]] = []
    for page in parsed.pages:
        text = unicodedata.normalize("NFKC", page.text)
        for model_match in GENERIC_MODEL_PATTERN.finditer(text):
            model = _canonical_model(model_match.group(0))
            before = text[max(0, model_match.start() - 100):model_match.start()]
            after = text[model_match.end():model_match.end() + 350]
            qty_match = re.search(r"(\d+)\s*(?:台|套)\s*$", before)
            weeks_match = re.search(r"(?:发货时间|交货时间|交期)?[^\n]{0,120}?(\d+)\s*周", after, re.I)
            if not weeks_match:
                continue
            condition_text = after[:weeks_match.start()]
            schedule.append({
                "model": model,
                "qty": int(qty_match.group(1)) if qty_match else 0,
                "weeks": int(weeks_match.group(1)),
                "condition": _clean_inline(condition_text),
                "page": page.page_num,
            })
    unique: Dict[str, Dict[str, Any]] = {}
    for entry in schedule:
        unique[entry["model"]] = entry
    return list(unique.values())


def _extract_payment_terms(text: str) -> Dict[str, Any]:
    source = (text or "").replace("\r\n", "\n").replace("\r", "\n")
    start_match = re.search(r"(?:付款条件|付款方式)", source)
    section_raw = source[start_match.start():] if start_match else source
    end_match = re.search(r"\n\s*(?:附件三|诚信条款|质量保证|质保条款|交货时间|交货条款)\b", section_raw[1:])
    if end_match:
        section_raw = section_raw[: end_match.start() + 1]
    section_raw = section_raw[:4000].strip()
    section = unicodedata.normalize("NFKC", section_raw)

    chinese_numbers = {
        "十": 10, "二十": 20, "三十": 30, "四十": 40, "五十": 50,
        "六十": 60, "七十": 70, "八十": 80, "九十": 90, "一百": 100,
    }

    def clause_percentages(clause: str) -> List[float]:
        arabic = [float(value) for value in re.findall(r"(\d+(?:[.]\d+)?)\s*%", clause)]
        if arabic:
            return arabic
        return [float(chinese_numbers[value]) for value in re.findall(r"百分之([一二三四五六七八九十百]+)", clause) if value in chinese_numbers]

    clauses = [
        _clean_inline(part)
        for part in re.split(r"(?=(?:^|\n)\s*(?:\d+|[一二三四五六七八九十]+)[）).、])", section)
        if _clean_inline(part)
    ]
    if len(clauses) <= 1:
        clauses = [_clean_inline(line) for line in section.splitlines() if _clean_inline(line)]

    percentage_values: List[float] = []
    installments: List[Dict[str, Any]] = []
    for clause in clauses:
        values = clause_percentages(clause)
        percentage_values.extend(values)
        for percent in values:
            installments.append({
                "percent": int(percent) if percent.is_integer() else percent,
                "text": clause,
                "trigger": _clean_inline(_match_first(clause, [r"(?:在|于)?(.{0,80}?(?:后|前|时|之日))"])),
                "method": "银行承兑汇票" if "承兑" in clause else ("电汇" if "电汇" in clause else ""),
            })
    if not percentage_values:
        percentage_values = clause_percentages(section)

    bank = _match_first(section, [r"用户银行[:：]\s*([^\n]+)", r"结算银行[:：]?\s*([^\n]+)", r"开户银行[:：]?\s*([^\n]+)"])
    account_name = _match_first(section, [r"(?:账户名称|户名|名称)[:：]\s*([^\n]+)"])
    account_number = re.sub(r"\s+", "", _match_first(section, [r"(?:账户号码|帐号|账号)[:：]\s*([\d\s-]+)"]))
    return {
        "installments": installments,
        "percentages": percentage_values,
        "raw": section_raw,
        "bank": bank,
        "account_name": account_name,
        "account_number": account_number,
        "complete": bool(section_raw and percentage_values and abs(sum(percentage_values) - 100) < 0.01),
    }


def _extract_configurations(parsed: Optional[ParsedPDF], source_type: str) -> List[Dict[str, Any]]:
    """Extract configuration codes and descriptions, bound to the nearest robot model."""
    if parsed is None:
        return []
    code_pattern = re.compile(r"(?<![\d.])(\d{3,4}-\d{1,4})(?![\d.])")
    configs: List[Dict[str, Any]] = []
    current_model = ""
    for page in parsed.pages:
        lines = [unicodedata.normalize("NFKC", line).strip() for line in page.text.splitlines() if line.strip()]
        page_models = _models_in_text("\n".join(lines))
        if len(page_models) == 1:
            current_model = page_models[0]
        for index, line in enumerate(lines):
            line_model_matches = list(GENERIC_MODEL_PATTERN.finditer(line))
            if line_model_matches:
                current_model = _canonical_model(line_model_matches[0].group(0))
            for match in code_pattern.finditer(line):
                if any(model.start() <= match.start() < model.end() for model in line_model_matches):
                    continue
                code = match.group(1)
                description = line[match.end():].strip(" :：-—")
                if not description and index + 1 < len(lines):
                    description = lines[index + 1]
                if code.startswith("3300-"):
                    detected = _canonical_model(description)
                    if detected and detected.startswith("IRB "):
                        current_model = detected
                if not current_model:
                    continue
                configs.append({
                    "model": current_model,
                    "code": code,
                    "description": _clean_inline(description),
                    "page": page.page_num,
                    "source": source_type,
                })
    unique: List[Dict[str, Any]] = []
    seen = set()
    for config in configs:
        key = (config.get("model"), config.get("code"))
        if key not in seen:
            unique.append(config)
            seen.add(key)
    return unique


def _model_section_text(parsed: Optional[ParsedPDF], model: str) -> str:
    if parsed is None:
        return ""
    pages = []
    for page in parsed.pages:
        models = _models_in_text(page.text)
        if model in models:
            pages.append(page.text)
    return "\n".join(pages)


def _extract_warranty_codes_by_model(configs: Sequence[Dict[str, Any]]) -> Dict[str, str]:
    result: Dict[str, str] = {}
    for config in configs:
        code = config.get("code", "")
        model = config.get("model", "")
        description = config.get("description", "")
        if model and (code.startswith("438-") or re.search(r"warranty|质保|保修", description, re.I)):
            result[model] = code
    return result


def _extract_named_clause(text: str, clause_number: str, max_chars: int = 4000) -> str:
    normalized = unicodedata.normalize("NFKC", text or "").replace("\r\n", "\n").replace("\r", "\n")
    start = re.search(rf"(?m)^\s*{re.escape(clause_number)}\s*[.、]?\s*", normalized)
    if not start:
        start = re.search(rf"(?<!\d){re.escape(clause_number)}\s*[.、]?\s*", normalized)
    if not start:
        return ""
    major, _, minor = clause_number.partition(".")
    next_minor = str(int(minor) + 1) if minor.isdigit() else ""
    tail = normalized[start.start(): start.start() + max_chars]
    if next_minor:
        end = re.search(rf"(?m)^\s*{re.escape(major)}[.]\s*{re.escape(next_minor)}\b", tail[1:])
        if end:
            tail = tail[: end.start() + 1]
    return tail.strip()


def _extract_warranty_clause(text: str) -> Dict[str, Any]:
    raw = _extract_named_clause(text, "5.2")
    if not raw:
        marker = re.search(r"(?:质量保证|质保条款|保修期)", text or "", re.I)
        raw = (text[marker.start(): marker.start() + 3000] if marker else "").strip()
    normalized = unicodedata.normalize("NFKC", raw)
    chinese_months = {"十二": 12, "十五": 15, "十八": 18, "二十四": 24, "三十六": 36}
    converted = normalized
    for word, number in chinese_months.items():
        converted = converted.replace(word, str(number))
    periods: List[Dict[str, Any]] = []
    pair_pattern = re.compile(
        r"(\d{1,3})\s*[（(]?\s*\d*\s*[）)]?\s*个?月.{0,180}?(\d{1,3})\s*[（(]?\s*\d*\s*[）)]?\s*个?月",
        re.S,
    )
    for match in pair_pattern.finditer(converted):
        first, second = int(match.group(1)), int(match.group(2))
        context = _clean_inline(converted[max(0, match.start() - 100): match.end() + 80])
        periods.append({"period": f"{first}/{second}", "first_months": first, "second_months": second, "context": context})
    for match in re.finditer(r"(?<!\d)(\d{1,3})\s*/\s*(\d{1,3})(?!\d)", converted):
        period = f"{int(match.group(1))}/{int(match.group(2))}"
        if not any(item["period"] == period for item in periods):
            periods.append({
                "period": period,
                "first_months": int(match.group(1)),
                "second_months": int(match.group(2)),
                "context": _clean_inline(converted[max(0, match.start() - 100): match.end() + 80]),
            })
    primary = periods[0]["period"] if periods else ""
    return {"raw": raw, "periods": periods, "primary_period": primary, "classification": _warranty_class_from_text(raw, periods)}


def _warranty_class_from_text(text: str, periods: Optional[Sequence[Dict[str, Any]]] = None) -> str:
    normalized = match_key(text or "")
    if any(term in normalized for term in ("extendedwarranty", "延长质保", "延保", "延长保修")):
        return "Extended Warranty"
    periods = list(periods or [])
    if any(int(item.get("first_months", 0)) > 18 or int(item.get("second_months", 0)) > 12 for item in periods):
        return "Extended Warranty"
    if any(term in normalized for term in ("standardwarranty", "标准质保", "标准保修", "litewarranty")):
        return "Standard Warranty"
    if periods:
        return "Standard Warranty"
    return "Unknown"


def _warranty_for_model(warranty: Dict[str, Any], model: str) -> Dict[str, Any]:
    periods = warranty.get("periods", [])
    model_key = _compact(model)
    for item in periods:
        if model_key and model_key in _compact(item.get("context", "")):
            return item
    if "LPS" in model.upper():
        for item in periods:
            if "lps" in match_key(item.get("context", "")) or "lite" in match_key(item.get("context", "")):
                return item
    return periods[0] if periods else {"period": "", "first_months": 0, "second_months": 0, "context": ""}


def _warranty_config_details(configs: Sequence[Dict[str, Any]]) -> Dict[str, Dict[str, str]]:
    result: Dict[str, Dict[str, str]] = {}
    for config in configs:
        code = str(config.get("code", ""))
        description = str(config.get("description", ""))
        if code.startswith("438-") or re.search(r"warranty|质保|保修", description, re.I):
            result[str(config.get("model", ""))] = {
                "code": code,
                "description": description,
                "classification": _warranty_class_from_text(description),
            }
    return result


def _detect_contract_incoterm(page_text: str) -> Dict[str, Any]:
    lines: List[Dict[str, str]] = []
    selected: List[str] = []
    for raw_line in unicodedata.normalize("NFKC", page_text or "").splitlines():
        line = _clean_inline(raw_line)
        if "到货价" not in line and "出厂价" not in line:
            continue
        term = "DDP" if "到货价" in line else "EXW"
        checked = any(marker in line for marker in CHECKED_MARKERS)
        unchecked = any(marker in line for marker in UNCHECKED_MARKERS)
        state = "checked" if checked and not unchecked else ("unchecked" if unchecked and not checked else "unknown")
        lines.append({"term": term, "state": state, "text": line})
        if state == "checked":
            selected.append(term)
    return {
        "selected": selected[0] if len(set(selected)) == 1 else "",
        "selected_terms": sorted(set(selected)),
        "conflict": len(set(selected)) > 1,
        "lines": lines,
        "candidates": sorted(set(item["term"] for item in lines)),
    }


def _normalize_incoterm(value: str) -> str:
    normalized = match_key(value or "")
    if "ddp" in normalized or "到货" in normalized:
        return "DDP"
    if "exw" in normalized or "出厂" in normalized:
        return "EXW"
    return ""


def _infer_incoterm(contract: Dict[str, Any], cqp: Dict[str, Any]) -> Dict[str, Any]:
    detection = contract.get("incoterm_detection", {})
    selected = detection.get("selected", "")
    cqp_term = _normalize_incoterm(cqp.get("delivery_term", ""))
    location = _clean_inline(contract.get("delivery_location", ""))
    conflict = bool(detection.get("conflict"))
    status, severity, conclusion, reason = "UNDETERMINED", "blocker", "", "证据不足，无法确认DDP或EXW。"

    if conflict:
        reason = "合同中DDP与EXW均被识别为已勾选，合同证据互相矛盾。"
    elif selected:
        conclusion = selected
        if cqp_term and cqp_term != selected:
            status, severity = "MISMATCH", "blocker"
            reason = f"合同明确选择{selected}，但CQP为{cqp_term}。"
        elif selected == "DDP" and not location:
            status, severity = "UNDETERMINED", "blocker"
            reason = "合同选择DDP，但未提取到明确交付地点。"
        elif selected == "EXW" and location:
            status, severity = "WARNING", "warning"
            reason = "合同与CQP均为EXW，但合同仍填写了交付地点；需确认该地点不会被误用为Ship-to。"
        elif cqp_term:
            status, severity = "PASS", "info"
            reason = f"合同勾选与CQP均为{selected}。"
        else:
            status, severity = "WARNING", "warning"
            reason = f"合同明确选择{selected}，但CQP贸易术语未提取，需备注复核。"
    elif cqp_term == "EXW" and not location:
        conclusion, status, severity = "EXW", "WARNING", "warning"
        reason = "合同勾选未识别，但交付地点为空且CQP为EXW，可合理推定为EXW。"
    elif cqp_term == "DDP" and location:
        conclusion, status, severity = "DDP", "WARNING", "warning"
        reason = "合同勾选未识别，但存在明确交付地点且CQP为DDP，可合理推定为DDP。"
    elif cqp_term == "EXW" and location:
        conclusion, status, severity = "", "MISMATCH", "blocker"
        reason = "CQP为EXW，但合同存在明确交付地点，证据冲突。"
    elif cqp_term == "DDP" and not location:
        conclusion, status, severity = "", "UNDETERMINED", "blocker"
        reason = "CQP为DDP，但合同交付地点缺失，无法确认完整named place。"

    buyer_name = contract.get("buyer_name", "")
    buyer_address = contract.get("buyer_address", "")
    if conclusion == "EXW":
        ship_to_name = buyer_name
        ship_to_address = buyer_address
        incoterm_2 = "Shanghai"
        ship_to_rule = "EXW下BT09 Ship-to使用Buyer name/address；Incoterm 2填卖方起运地Shanghai。"
    elif conclusion == "DDP":
        ship_to_name = contract.get("ship_to_name") or contract.get("end_customer_name") or buyer_name
        ship_to_address = location
        incoterm_2 = location
        ship_to_rule = "DDP下Ship-to与Incoterm 2使用合同明确的客户目的地；不得默认使用安装地点。"
    else:
        ship_to_name = ship_to_address = incoterm_2 = ""
        ship_to_rule = "Incoterm未确认，暂不生成Ship-to。"
    return {
        "conclusion": conclusion or "UNDETERMINED",
        "status": status,
        "severity": severity,
        "reason": reason,
        "contract_selected": selected,
        "cqp_term": cqp_term,
        "delivery_location": location,
        "ship_to_name": ship_to_name,
        "ship_to_address": ship_to_address,
        "incoterm_2": incoterm_2,
        "ship_to_rule": ship_to_rule,
    }


def _payment_terms_consistency(contract_terms: Dict[str, Any], cqp_terms: Dict[str, Any]) -> Tuple[str, str, str]:
    if not contract_terms.get("raw") or not contract_terms.get("complete"):
        return "MISMATCH", "blocker", "合同附件二付款条件缺失、无法完整提取或比例合计不为100%。"
    if not cqp_terms.get("raw"):
        return "WARNING", "warning", "合同付款条件已提取；CQP付款条件未提取。BT09仍应原文照抄合同附件二。"
    contract_percentages = sorted(contract_terms.get("percentages", []))
    cqp_percentages = sorted(cqp_terms.get("percentages", []))
    if contract_percentages != cqp_percentages:
        return "MISMATCH", "blocker", "合同与CQP付款比例不一致。"

    def signatures(terms: Dict[str, Any]) -> List[Tuple[Any, Tuple[str, ...], str]]:
        keywords = ("签订合同", "合同生效", "预付款", "发货前", "发货后", "交付", "验收", "开票", "工作日")
        output = []
        for entry in terms.get("installments", []):
            text = str(entry.get("text", ""))
            events = tuple(keyword for keyword in keywords if keyword in text)
            output.append((entry.get("percent"), events, str(entry.get("method", ""))))
        return sorted(output, key=lambda item: (float(item[0] or 0), item[1], item[2]))

    contract_signatures, cqp_signatures = signatures(contract_terms), signatures(cqp_terms)
    if contract_signatures and cqp_signatures and contract_signatures != cqp_signatures:
        return "MISMATCH", "blocker", "合同与CQP付款触发条件或付款方式不一致。"
    return "PASS", "info", "合同与CQP付款核心条件一致；BT09付款条件应使用合同附件二原文。"


def _address_match(left: str, right: str) -> bool:
    left_key, right_key = _compact(left), _compact(right)
    if not left_key or not right_key:
        return False
    if left_key in right_key or right_key in left_key:
        return True
    left_digits = set(re.findall(r"\d+", left_key))
    right_digits = set(re.findall(r"\d+", right_key))
    digit_ok = not left_digits or not right_digits or bool(left_digits & right_digits)
    return digit_ok and SequenceMatcher(None, left_key, right_key).ratio() >= 0.58


def _amount_status(left: float, right: float) -> Tuple[str, str, float]:
    if not left or not right:
        return "UNDETERMINED", "blocker", round(abs((left or 0) - (right or 0)), 2)
    diff = round(abs(left - right), 2)
    if diff == 0:
        return "PASS", "info", diff
    if diff < AMOUNT_ROUNDING_TOLERANCE:
        return "WARNING", "warning", diff
    return "MISMATCH", "blocker", diff


def _config_semantic_match(config: Dict[str, Any], ta_configs: Sequence[Dict[str, Any]], ta_section: str) -> Tuple[bool, str]:
    model, code = config.get("model", ""), config.get("code", "")
    exact = any(item.get("model") == model and item.get("code") == code for item in ta_configs)
    if exact:
        return True, "相同代码"
    terms = list(CONFIG_FEATURES.get(code, ()))
    if config.get("description"):
        terms.append(str(config.get("description")))
    for term in terms:
        term_key = match_key(term)
        if term_key and len(term_key) >= 4 and term_key in match_key(ta_section):
            return True, "等价描述"
    return False, "未匹配"


# ---------------------------------------------------------------------------
# Structured extraction
# ---------------------------------------------------------------------------


def extract_contract(parsed: Optional[ParsedPDF]) -> Dict[str, Any]:
    if parsed is None:
        return {}
    text = unicodedata.normalize("NFKC", parsed.full_text)
    first_pages = "\n".join(page.text for page in parsed.pages[:3])
    page2 = unicodedata.normalize("NFKC", _page_text(parsed, 2))
    signature_text = "\n".join(page.text for page in parsed.pages[-3:])

    contract_match = re.search(r"\b([MK])\s*(\d{4})\s*-\s*(\d{4})\b", text, re.I)
    contract_number = "" if not contract_match else f"{contract_match.group(1).upper()}{contract_match.group(2)}-{contract_match.group(3)}"
    buyer = _match_first(first_pages, [r"买方[:：]\s*(.*?)\s*地址[:：]", r"甲方[（(]买方[）)][:：]\s*([^\n]+)"], re.S)
    seller = _match_first(first_pages, [r"卖方[:：]\s*(.*?)\s*地址[:：]", r"乙方[（(]卖方[）)][:：]\s*([^\n]+)"], re.S)
    buyer_address = _match_first(first_pages, [r"买方[:：].*?地址[:：]\s*(.*?)\s*卖方[:：]", r"买方地址[:：]\s*([^\n]+)"], re.S)
    seller_address = _match_first(first_pages, [r"卖方[:：].*?地址[:：]\s*(.*?)(?:目录|合同编号|$)", r"卖方地址[:：]\s*([^\n]+)"], re.S)
    quantities = _extract_model_quantities(parsed)
    delivery_schedule = _extract_delivery_schedule(parsed)

    untaxed_text = _match_first(text, [r"不含增值税(?:总额|金额)?(?:为)?[:：]?\s*(?:CNY|RMB|人民币)?\s*([\d,]+[.]\d{1,2})"])
    gross_text = _match_first(text, [r"含增值税(?:总额|金额)?(?:为)?[:：]?\s*(?:CNY|RMB|人民币)?\s*([\d,]+[.]\d{1,2})", r"合同价格的含增值税总额为[:：]?.*?([\d,]+[.]\d{1,2})"])
    vat_text = _match_first(text, [r"增值税(?:税率)?[:：]?\s*(\d{1,2}(?:[.]\d+)?)\s*%"])

    annex2_match = re.search(r"附件二[^\n]{0,80}(?:付款方式|付款条件)", text)
    payment_source = text[annex2_match.start():] if annex2_match else text
    annex3 = re.search(r"\n\s*附件三\b", payment_source[1:])
    if annex3:
        payment_source = payment_source[:annex3.start() + 1]
    warranty = _extract_warranty_clause(text)
    incoterm_detection = _detect_contract_incoterm(page2 or text)

    placeholder_text = re.sub(r"\s+", "", signature_text + "\n" + text)
    known_placeholders = {"@@@Chop_ABB", "@@@Chop_Customer", "@@@Sign_ABBPerson", "@@@Sign_CustomerPerson"}
    placeholders = sorted(token for token in known_placeholders if token in placeholder_text)
    blank_dates = bool(re.search(r"日期[:：]\s*日期[:：]", signature_text) or re.search(r"日期[:：]\s*(?:\n|$)", signature_text))

    delivery_trigger = ""
    for phrase in ("合同生效且收到预付款后", "合同生效并收到预付款后", "合同生效且预付款到账后", "预付款到账后"):
        if phrase in text:
            delivery_trigger = phrase
            break

    return {
        "contract_number": contract_number,
        "cqp_reference": _match_first(text, [r"(?:单价信息|报价单|CQP)[:：]?\s*(CQ\d{7})", r"\b(CQ\d{7})\b"]),
        "buyer_name": _clean_entity(buyer),
        "buyer_address": _clean_entity(buyer_address),
        "seller_name": _clean_entity(seller),
        "seller_address": _clean_entity(seller_address),
        "project_name": _match_first(text, [r"买方基于(.{2,80}?)项目需求", r"项目名称[:：]\s*([^\n]+)"]),
        "end_customer_name": _clean_entity(_match_first(text, [r"最终用户(?:名称)?[:：]\s*([^\n]+)"])),
        "end_customer_address": _clean_entity(_match_first(text, [r"最终用户地址[:：]\s*([^\n]+)"])),
        "installation_location": _clean_entity(_match_first(text, [r"(?:设备)?安装地点[:：]_?\s*([^\n]+)"])),
        "delivery_location": _clean_entity(_match_first(text, [r"交付地点[:：]_?\s*([^\n]+)", r"交货地点[:：]_?\s*([^\n]+)"])),
        "ship_to_name": _clean_entity(_match_first(text, [r"(?:Ship[- ]?to|收货方)(?:名称)?[:：]\s*([^\n]+)"])),
        "ship_to_address": _clean_entity(_match_first(text, [r"(?:Ship[- ]?to|收货方)地址[:：]\s*([^\n]+)"])),
        "sales_person": _clean_inline(_match_first(text, [r"(?:销售人员|销售负责人|Sales)[:：]\s*([^\n]+)"])),
        "pm": _clean_inline(_match_first(text, [r"(?:项目经理|PM)[:：]\s*([^\n]+)"])),
        "products": [{"model": model, "qty": qty} for model, qty in quantities.items()],
        "total_qty": sum(quantities.values()),
        "delivery_schedule": delivery_schedule,
        "delivery_trigger": delivery_trigger,
        "split_delivery": bool(re.search(r"允许分批(?:装运|发货)", text)),
        "incoterm_detection": incoterm_detection,
        "incoterm_options": incoterm_detection.get("candidates", []),
        "incoterm_selected": incoterm_detection.get("selected", ""),
        "untaxed_amount": _parse_money(untaxed_text),
        "vat_rate": float(vat_text) / 100 if vat_text else 0.0,
        "tax_included_amount": _parse_money(gross_text),
        "amount_uppercase_untaxed": _match_first(text, [r"不含增值税.*?大写[:：]?(?:总计人民币)?[:：]?([^\n]+)"], re.S),
        "amount_uppercase_gross": _match_first(text, [r"含增值税.*?大写[:：]?(?:总计人民币)?[:：]?([^\n]+)"], re.S),
        "payment_terms": _extract_payment_terms(payment_source),
        "warranty": warranty,
        "signature_placeholders": placeholders,
        "blank_signature_dates": blank_dates,
        "attachments": {
            "ta": bool(re.search(r"附件一[^\n]{0,80}技术协议", text)),
            "payment": bool(re.search(r"附件二[^\n]{0,80}(?:付款方式|付款条件)", text)),
            "integrity": bool(re.search(r"附件三[^\n]{0,80}诚信条款", text)),
        },
        "file_priority": "本合同优先于附件" if re.search(r"本合同.*?优先于.*?附件|文件优先性", text, re.S) else "",
    }


def extract_cqp(parsed: Optional[ParsedPDF]) -> Dict[str, Any]:
    if parsed is None:
        return {}
    text = unicodedata.normalize("NFKC", parsed.full_text)
    products = _extract_cqp_products(parsed)
    configs = _extract_configurations(parsed, "cqp")

    cqp_number_match = re.search(r"(?:报价单编号|报价编号)[:：]?\s*(CQ\d{7})", text)
    if not cqp_number_match:
        cqp_number_match = re.search(r"\b(CQ\d{7})\b", text)
    version = _match_first(text, [r"报价修订版本[:：]?\s*([A-Za-z0-9._-]+)", r"版本号[:：]?\s*([A-Za-z0-9._-]+)"])
    customer = _first_line_value(text, "客户") or _match_first(text, [r"客户(?:名称)?[:：]\s*([^\n]+)"])
    customer_address = _first_line_value(text, "联络地址") or _match_first(text, [r"客户地址[:：]\s*([^\n]+)"])
    seller = _match_first(text, [r"ABB\s*公司名称[:：]?\s*([^\n]+)", r"ABB\s*单位名称[:：]?\s*([^\n]+)"])
    project = _first_line_value(text, "项目名称") or _match_first(text, [r"项目名称[:：]?\s*([^\n]+)"])

    money_values = [_parse_money(value) for value in re.findall(r"CNY\s*([\d,]+[.]\d{1,2})", text, re.I)]
    untaxed = _parse_money(_match_first(text, [r"(?:未税|不含税)(?:总额|金额)?[:：]?\s*(?:CNY|RMB)?\s*([\d,]+[.]\d{1,2})"]))
    tax_amount = _parse_money(_match_first(text, [r"(?:增值税额|税额)[:：]?\s*(?:CNY|RMB)?\s*([\d,]+[.]\d{1,2})"]))
    gross = _parse_money(_match_first(text, [r"(?:含税总额|含税金额)[:：]?\s*(?:CNY|RMB)?\s*([\d,]+[.]\d{1,2})"]))
    if not untaxed and money_values:
        untaxed = money_values[0]
    if not tax_amount and len(money_values) >= 2:
        tax_amount = money_values[1]
    if not gross and len(money_values) >= 3:
        gross = money_values[2]

    payment_marker = re.search(r"付款条件", text)
    payment_text = text[payment_marker.start():] if payment_marker else ""
    delivery_time = _match_first(text, [r"交货时间[:：]?\s*([^\n]+)", r"交期[:：]?\s*([^\n]+)"])
    delivery_term = _match_first(text, [r"交货条款[:：]?\s*([^\n]+)", r"贸易术语[:：]?\s*([^\n]+)"])
    warranty_terms = _match_first(text, [r"质量保证[:：]?\s*([^\n]+)", r"质保[:：]?\s*([^\n]+)"])
    weeks_match = re.search(r"(\d+)\s*周", delivery_time)
    vat_match = re.search(r"增值税(?:率)?\s*(\d{1,2}(?:[.]\d+)?)\s*%", text)

    return {
        "cqp_number": cqp_number_match.group(1) if cqp_number_match else "",
        "version": version,
        "quote_date": _match_first(text, [r"报价日期[:：]?\s*([0-9]{4}[-/.][0-9]{1,2}[-/.][0-9]{1,2})"]),
        "document_date": _match_first(text, [r"日期[:：]?\s*([0-9]{4}[-/.][0-9]{1,2}[-/.][0-9]{1,2})"]),
        "validity": _match_first(text, [r"报价有效期限[:：]?\s*([^\n]+)"]),
        "project_name": _clean_inline(project),
        "customer_name": _clean_inline(customer),
        "customer_address": _clean_inline(customer_address),
        "end_user": _clean_entity(_match_first(text, [r"最终用户(?:名称)?[:：]\s*([^\n]+)"])),
        "end_user_address": _clean_entity(_match_first(text, [r"最终用户地址[:：]\s*([^\n]+)"])),
        "ship_to_name": _clean_entity(_match_first(text, [r"(?:Ship[- ]?to|收货方)(?:名称)?[:：]\s*([^\n]+)"])),
        "ship_to_address": _clean_entity(_match_first(text, [r"(?:Ship[- ]?to|收货方)地址[:：]\s*([^\n]+)"])),
        "seller_name": _clean_entity(seller),
        "contact_name": _first_line_value(text, "联络人"),
        "contact_email": _first_line_value(text, "联系人电子邮箱"),
        "sales_person": _clean_inline(_match_first(text, [r"(?:负责人|销售人员|Sales)[:：]\s*([^\n]+)"])),
        "products": products,
        "total_qty": sum(int(product.get("qty", 0)) for product in products),
        "untaxed_total": untaxed,
        "tax_amount": tax_amount,
        "tax_included_total": gross,
        "vat_rate": float(vat_match.group(1)) / 100 if vat_match else 0.0,
        "payment_terms": _extract_payment_terms(payment_text),
        "delivery_time": _clean_inline(delivery_time),
        "delivery_weeks": int(weeks_match.group(1)) if weeks_match else 0,
        "delivery_trigger": "预付款到账" if "预付款" in delivery_time else ("合同生效" if "合同生效" in delivery_time else ""),
        "delivery_term": _clean_inline(delivery_term),
        "warranty_terms": _clean_inline(warranty_terms),
        "configurations": configs,
        "warranty_codes_by_model": _extract_warranty_codes_by_model(configs),
        "warranty_details_by_model": _warranty_config_details(configs),
        "line_rounding": [
            {
                "model": product["model"],
                "shown_calculation": round(product["unit_price"] * product["qty"], 2),
                "line_total": product["line_total"],
                "difference": round(product["line_total"] - product["unit_price"] * product["qty"], 2),
            }
            for product in products
            if product.get("unit_price") and product.get("line_total")
        ],
    }


def extract_ta(parsed: Optional[ParsedPDF]) -> Dict[str, Any]:
    if parsed is None:
        return {}
    text = unicodedata.normalize("NFKC", parsed.full_text)
    configs = _extract_configurations(parsed, "ta")
    quantities = _extract_model_quantities(parsed)
    compact_text = re.sub(r"\s+", "", text)
    known_placeholders = {"@@@Chop_ABB", "@@@Chop_Customer", "@@@Sign_ABBPerson", "@@@Sign_CustomerPerson"}
    placeholders = sorted(token for token in known_placeholders if token in compact_text)
    contract_match = re.search(r"合同编号[:：]?\s*([MK]\s*\d{4}\s*-\s*\d{4})", text, re.I)
    contract_number = re.sub(r"\s+", "", contract_match.group(1)).upper() if contract_match else ""
    return {
        "contract_number": contract_number,
        "buyer_name": _clean_entity(_match_first(text, [r"甲方[（(]买方[）)][:：]\s*([^\n]+)", r"买方[:：]\s*([^\n]+)"])),
        "buyer_address": _clean_entity(_match_first(text, [r"买方地址[:：]\s*([^\n]+)"])),
        "seller_name": _clean_entity(_match_first(text, [r"卖方[（(]乙方[）)][:：]\s*([^\n]+)", r"乙方[（(]卖方[）)][:：]\s*([^\n]+)"])),
        "products": [{"model": model, "qty": qty} for model, qty in quantities.items()],
        "total_qty": sum(quantities.values()),
        "configurations": configs,
        "warranty_codes_by_model": _extract_warranty_codes_by_model(configs),
        "warranty_details_by_model": _warranty_config_details(configs),
        "lps_name_in_supply": "LPS" if re.search(r"IRB\s*\d{3,4}[^\n]{0,40}\bLPS\b", text, re.I) else "",
        "lps_name_in_parameters": "Lite+" if re.search(r"IRB\s*\d{3,4}[^\n]{0,40}Lite\s*[+＋]", text, re.I) else "",
        "repeatability_gen2": _match_first(text, [r"IRB\s*\d{3,4}[^\n]{0,80}Gen\s*2.*?位置重复精度[:：]\s*([^\n]+)"], re.S),
        "signature_placeholders": placeholders,
        "blank_signature_dates": bool(re.search(r"日期[:：]\s+日期[:：]", text) or re.search(r"日期[:：]\s*(?:\n|$)", text)),
        "responsibilities": {
            "buyer_integration": bool(re.search(r"买方.*?负责.*?系统集成", text, re.S)),
            "buyer_installation": bool(re.search(r"买方.*?负责.*?(?:卸货|起吊|就位|现场安装)", text, re.S)),
            "seller_not_integration": bool(re.search(r"卖方.*?不承担.*?系统集成", text, re.S)),
        },
    }


# ---------------------------------------------------------------------------
# Evidence helpers
# ---------------------------------------------------------------------------


def _evidence(
    parsed: Optional[ParsedPDF],
    document_type: str,
    label: str,
    terms: Sequence[str],
    *,
    page_hint: Optional[int] = None,
    quote: str = "",
) -> Dict[str, Any]:
    return locate_evidence(parsed, document_type, label, terms, page_hint=page_hint, quote=quote)


def _evidence_many(parsed: Optional[ParsedPDF], document_type: str, label: str, terms: Sequence[str]) -> List[Dict[str, Any]]:
    return locate_all_evidence(parsed, document_type, label, terms)


def _valid_evidence(entries: Iterable[Dict[str, Any]]) -> List[Dict[str, Any]]:
    output = []
    seen = set()
    for entry in entries:
        if not entry:
            continue
        key = (entry.get("document_type"), entry.get("page"), entry.get("label"), entry.get("quote"))
        if key not in seen:
            output.append(entry)
            seen.add(key)
    return output


# ---------------------------------------------------------------------------
# Review item construction
# ---------------------------------------------------------------------------


def _item(
    item_id: str,
    category: str,
    title: str,
    status: str,
    severity: str,
    summary: str,
    values: Optional[Dict[str, Any]] = None,
    evidence: Optional[List[Dict[str, Any]]] = None,
    sub_items: Optional[List[Dict[str, Any]]] = None,
) -> Dict[str, Any]:
    result: Dict[str, Any] = {
        "id": item_id,
        "category": category,
        "category_label": CATEGORY_LABELS[category],
        "title": title,
        "status": status,
        "severity": severity,
        "summary": summary,
        "values": values or {},
        "evidence": _valid_evidence(evidence or []),
    }
    if sub_items:
        result["sub_items"] = sub_items
    return result


def _status_from_missing(*values: Any) -> Optional[Tuple[str, str]]:
    if any(value in (None, "", [], {}) for value in values):
        return "UNDETERMINED", "warning"
    return None


def _product_map(products: Sequence[Dict[str, Any]]) -> Dict[str, Dict[str, Any]]:
    return {product.get("model", ""): product for product in products if product.get("model")}


def build_review_items(
    documents: DocumentSet,
    contract: Dict[str, Any],
    cqp: Dict[str, Any],
    ta: Dict[str, Any],
) -> List[Dict[str, Any]]:
    items: List[Dict[str, Any]] = []
    contract_pdf, cqp_pdf, ta_pdf = documents.contract_physical, documents.cqp, documents.ta

    # ---------------- Source completeness ----------------
    ta_has_content = bool(documents.ta and (ta.get("products") or ta.get("configurations") or _find_ta_start(documents.ta)))
    source_ok = bool(documents.contract_physical and documents.cqp and ta_has_content)
    items.append(_item(
        "source_completeness",
        CATEGORY_CUSTOMER,
        "文件来源完整性",
        "PASS" if source_ok else "MISMATCH",
        "info" if source_ok else "blocker",
        (
            "合同、CQP和TA均可读取；TA位于合同附件中。" if source_ok and documents.ta_embedded
            else "合同、CQP和独立TA均可读取。" if source_ok
            else "TA既未作为独立文件提供，也未在合同中识别到可验证的技术协议内容。"
        ),
        {
            "合同": "已找到" if documents.contract_physical else "缺失",
            "CQP": "已找到" if documents.cqp else "缺失",
            "TA": "合同内嵌" if documents.ta_embedded else ("独立文件" if ta_has_content else "缺失"),
        },
        [
            _evidence(contract_pdf, "contract", "合同", [contract.get("contract_number", ""), "销售合同"]),
            _evidence(cqp_pdf, "cqp", "CQP", [cqp.get("cqp_number", ""), "报价"]),
            _evidence(ta_pdf, "ta", "TA", ["Technical Agreement", "技术协议书", "技术协议"]),
        ],
    ))

    # ---------------- Customer information ----------------
    contract_no = contract.get("contract_number", "")
    cqp_ref, cqp_no, ta_no = contract.get("cqp_reference", ""), cqp.get("cqp_number", ""), ta.get("contract_number", "")
    explicit_link_mismatch = bool((cqp_ref and cqp_no and cqp_ref != cqp_no) or (ta_no and contract_no and ta_no != contract_no))
    link_missing = not contract_no or not cqp_no or not cqp_ref or (ta_has_content and not ta_no)
    link_status = "MISMATCH" if explicit_link_mismatch else ("WARNING" if link_missing else "PASS")
    link_severity = "blocker" if explicit_link_mismatch else ("warning" if link_missing else "info")
    items.append(_item(
        "document_linkage", CATEGORY_CUSTOMER, "文件编号与版本关联", link_status, link_severity,
        "合同、CQP和TA编号关联一致。" if link_status == "PASS" else (
            "文件编号存在明确冲突。" if explicit_link_mismatch else "部分编号或版本未提取，无法完成全部关联校验。"
        ),
        {"合同编号": contract_no or "未提取", "合同内CQP引用": cqp_ref or "未提取", "CQP编号": cqp_no or "未提取", "TA合同编号": ta_no or "未提取", "CQP版本": cqp.get("version") or "未提取"},
        [
            _evidence(contract_pdf, "contract", "合同编号", [contract_no]),
            _evidence(contract_pdf, "contract", "CQP引用", [cqp_ref]),
            _evidence(cqp_pdf, "cqp", "CQP编号", [cqp_no]),
            _evidence(ta_pdf, "ta", "TA合同编号", [ta_no]),
        ],
    ))

    seller = contract.get("seller_name", "")
    expected_seller = SELLER_PREFIX_MAP.get(contract_no[:1], "")
    seller_values = [value for value in (seller, cqp.get("seller_name", ""), ta.get("seller_name", "")) if value]
    seller_conflict = bool(seller and any(not _contains_equivalent(seller, value) for value in seller_values[1:]))
    prefix_conflict = bool(expected_seller and seller and not _contains_equivalent(expected_seller, seller))
    seller_missing = not seller
    seller_status = "MISMATCH" if seller_conflict or prefix_conflict else ("UNDETERMINED" if seller_missing else "PASS")
    seller_severity = "blocker" if seller_conflict or prefix_conflict else ("warning" if seller_missing else "info")
    items.append(_item(
        "seller_entity", CATEGORY_CUSTOMER, "卖方法定实体", seller_status, seller_severity,
        "卖方主体与合同号前缀规则一致。" if seller_status == "PASS" else (
            "卖方主体与其他文件或合同号前缀规则冲突。" if seller_status == "MISMATCH" else "卖方主体无法完整识别，需人工校验盖章页。"
        ),
        {"合同": seller or "未提取", "CQP": cqp.get("seller_name") or "未提取", "TA": ta.get("seller_name") or "未提取", "合同号预期实体": expected_seller or "无映射"},
        [
            _evidence(contract_pdf, "contract", "合同卖方", [seller, expected_seller]),
            _evidence(cqp_pdf, "cqp", "CQP卖方", [cqp.get("seller_name", "")]),
            _evidence(ta_pdf, "ta", "TA卖方", [ta.get("seller_name", "")]),
        ],
    ))

    buyer, customer, ta_buyer = contract.get("buyer_name", ""), cqp.get("customer_name", ""), ta.get("buyer_name", "")
    ta_buyer_conflict = bool(buyer and ta_buyer and not _contains_equivalent(buyer, ta_buyer))
    direct_buyer_match = bool(buyer and customer and _contains_equivalent(buyer, customer))
    if not buyer:
        buyer_status, buyer_severity, buyer_summary = "UNDETERMINED", "blocker", "合同买方未提取，无法确认签约主体。"
    elif ta_buyer_conflict:
        buyer_status, buyer_severity, buyer_summary = "MISMATCH", "blocker", "合同与TA买方名称明确不一致。"
    elif customer and not direct_buyer_match:
        buyer_status, buyer_severity, buyer_summary = "WARNING", "warning", "CQP可能使用客户简称；需通过客户主数据确认其与合同买方为同一主体。"
    elif not customer:
        buyer_status, buyer_severity, buyer_summary = "WARNING", "warning", "CQP客户名称未提取，合同买方作为法律依据。"
    else:
        buyer_status, buyer_severity, buyer_summary = "PASS", "info", "买方主体一致。"
    items.append(_item(
        "buyer_customer_identity", CATEGORY_CUSTOMER, "买方与CQP客户", buyer_status, buyer_severity, buyer_summary,
        {"合同买方": buyer or "未提取", "CQP客户": customer or "未提取", "TA买方": ta_buyer or "未提取"},
        [
            _evidence(contract_pdf, "contract", "合同买方", [buyer]),
            _evidence(cqp_pdf, "cqp", "CQP客户", [customer]),
            _evidence(ta_pdf, "ta", "TA买方", [ta_buyer]),
        ],
    ))

    contract_address, cqp_address = contract.get("buyer_address", ""), cqp.get("customer_address", "")
    address_ok = _address_match(contract_address, cqp_address)
    address_status = "PASS" if address_ok else "WARNING"
    items.append(_item(
        "customer_address", CATEGORY_CUSTOMER, "客户地址", address_status, "info" if address_ok else "warning",
        "合同与CQP地址核心信息一致。" if address_ok else "客户地址缺失或仅能部分匹配，需要人工确认。",
        {"合同": contract_address or "未提取", "CQP": cqp_address or "未提取"},
        [
            _evidence(contract_pdf, "contract", "合同客户地址", [contract_address]),
            _evidence(cqp_pdf, "cqp", "CQP客户地址", [cqp_address]),
        ],
    ))

    end_user, cqp_end_user = contract.get("end_customer_name", ""), cqp.get("end_user", "")
    end_user_conflict = bool(end_user and cqp_end_user and not _contains_equivalent(end_user, cqp_end_user))
    if end_user_conflict:
        project_status, project_severity, project_summary = "MISMATCH", "blocker", "合同与CQP最终用户名称不一致。"
    elif not end_user and not cqp_end_user:
        project_status, project_severity, project_summary = "UNDETERMINED", "warning", "最终用户信息未提取。"
    elif end_user and not cqp_end_user:
        project_status, project_severity, project_summary = "WARNING", "warning", "合同写明最终用户，但CQP未明确列示；以合同为主并备注。"
    else:
        project_status, project_severity, project_summary = "PASS", "info", "最终用户信息一致。"
    items.append(_item(
        "project_end_user", CATEGORY_CUSTOMER, "项目、最终用户与安装地点", project_status, project_severity, project_summary,
        {"合同项目": contract.get("project_name") or "未提取", "CQP项目": cqp.get("project_name") or "未提取", "合同最终用户": end_user or "未提取", "CQP最终用户": cqp_end_user or "未提取", "安装地点": contract.get("installation_location") or "未提取"},
        [
            _evidence(contract_pdf, "contract", "最终用户", [end_user]),
            _evidence(contract_pdf, "contract", "安装地点", [contract.get("installation_location", "")]),
            _evidence(cqp_pdf, "cqp", "CQP项目/最终用户", [cqp.get("project_name", ""), cqp_end_user]),
        ],
    ))

    # ---------------- Product information ----------------
    c_products, q_products, t_products = _product_map(contract.get("products", [])), _product_map(cqp.get("products", [])), _product_map(ta.get("products", []))
    expected_models = set(c_products) | set(q_products) | set(t_products)
    model_mismatches = [model for model in sorted(expected_models) if model not in c_products or model not in q_products or (ta_has_content and model not in t_products)]
    no_models = not expected_models
    model_status = "UNDETERMINED" if no_models else ("MISMATCH" if model_mismatches else "PASS")
    items.append(_item(
        "product_models", CATEGORY_PRODUCT, "机器人型号", model_status, "blocker" if model_status != "PASS" else "info",
        "合同、CQP和TA中的机器人型号一致。" if model_status == "PASS" else ("未提取到机器人型号。" if no_models else "以下型号未同时出现在全部文件：" + "、".join(model_mismatches)),
        {"合同": "、".join(c_products) or "未提取", "CQP": "、".join(q_products) or "未提取", "TA": "、".join(t_products) or "未提取"},
        _evidence_many(contract_pdf, "contract", "合同型号", list(c_products)) + _evidence_many(cqp_pdf, "cqp", "CQP型号", list(q_products)) + _evidence_many(ta_pdf, "ta", "TA型号", list(t_products)),
    ))

    quantity_subitems, quantity_evidence = [], []
    quantity_mismatch = no_models
    for model in sorted(expected_models):
        cq, qq = int(c_products.get(model, {}).get("qty", 0)), int(q_products.get(model, {}).get("qty", 0))
        tq = int(t_products.get(model, {}).get("qty", 0)) if ta_has_content else 0
        ok = cq > 0 and qq > 0 and tq > 0 and cq == qq == tq
        quantity_mismatch = quantity_mismatch or not ok
        evidence = _valid_evidence(
            _evidence_many(contract_pdf, "contract", model + "数量", [model, str(cq)]) +
            _evidence_many(cqp_pdf, "cqp", model + "数量", [model, str(qq)]) +
            _evidence_many(ta_pdf, "ta", model + "数量", [model, str(tq)])
        )
        quantity_subitems.append({
            "id": "quantity_" + re.sub(r"[^a-z0-9]+", "_", model.lower()).strip("_"), "title": model,
            "status": "PASS" if ok else "MISMATCH", "summary": f"Contract {cq or '未提取'} / CQP {qq or '未提取'} / TA {tq or '未提取'}",
            "values": {"Contract": cq or "未提取", "CQP": qq or "未提取", "TA": tq or "未提取"}, "evidence": evidence,
        })
        quantity_evidence.extend(evidence)
    items.append(_item(
        "product_quantities", CATEGORY_PRODUCT, "各型号数量", "MISMATCH" if quantity_mismatch else "PASS", "blocker" if quantity_mismatch else "info",
        "各型号数量一致。" if not quantity_mismatch else "至少一个型号的数量不一致或无法提取。",
        {"合同总数": contract.get("total_qty") or "未提取", "CQP总数": cqp.get("total_qty") or "未提取", "TA总数": ta.get("total_qty") or "未提取"}, quantity_evidence, quantity_subitems,
    ))

    totals = (int(contract.get("total_qty", 0)), int(cqp.get("total_qty", 0)), int(ta.get("total_qty", 0)))
    total_ok = all(totals) and totals[0] == totals[1] == totals[2]
    items.append(_item(
        "total_quantity", CATEGORY_PRODUCT, "机器人总数量", "PASS" if total_ok else "MISMATCH", "info" if total_ok else "blocker",
        f"三份文件的机器人总数量均为{totals[0]}。" if total_ok else "机器人总数量不一致或未提取；总数应由各产品行数量求和。",
        {"合同": totals[0] or "未提取", "CQP": totals[1] or "未提取", "TA": totals[2] or "未提取"}, quantity_evidence,
    ))

    product_code_subitems = []
    for model, product in q_products.items():
        code = product.get("item_code", "")
        product_code_subitems.append({
            "id": "product_code_" + re.sub(r"[^a-z0-9]+", "_", model.lower()).strip("_"), "title": model,
            "status": "PASS" if code else "UNDETERMINED", "summary": f"产品代码：{code or '未提取'}",
            "values": {"CQP产品代码": code or "未提取"}, "evidence": [_evidence(cqp_pdf, "cqp", model + "产品代码", [code, model], page_hint=product.get("page"))],
        })
    all_codes = bool(product_code_subitems) and all(item["status"] == "PASS" for item in product_code_subitems)
    items.append(_item(
        "product_codes", CATEGORY_PRODUCT, "产品代码", "PASS" if all_codes else "UNDETERMINED", "info" if all_codes else "warning",
        "CQP产品代码均已提取。" if all_codes else "部分CQP产品代码未提取。", {},
        [entry for sub in product_code_subitems for entry in sub["evidence"]], product_code_subitems,
    ))

    contract_warranty = contract.get("warranty", {})
    cqp_warranty, ta_warranty = cqp.get("warranty_details_by_model", {}), ta.get("warranty_details_by_model", {})
    warranty_subitems, warranty_problem = [], not bool(contract_warranty.get("raw"))
    for model in sorted(expected_models):
        contract_period_info = _warranty_for_model(contract_warranty, model)
        contract_period = contract_period_info.get("period", "")
        contract_class = _warranty_class_from_text(contract_period_info.get("context", "") or contract_warranty.get("raw", ""), [contract_period_info] if contract_period else [])
        c_detail, t_detail = cqp_warranty.get(model, {}), ta_warranty.get(model, {})
        c_code, t_code = c_detail.get("code", ""), t_detail.get("code", "")
        codes_match = bool(c_code and t_code and c_code == t_code)
        classes = [value for value in (contract_class, c_detail.get("classification"), t_detail.get("classification")) if value and value != "Unknown"]
        class_match = len(set(classes)) <= 1
        ok = bool(contract_period and codes_match and class_match)
        warranty_problem = warranty_problem or not ok
        warranty_subitems.append({
            "id": "warranty_" + re.sub(r"[^a-z0-9]+", "_", model.lower()).strip("_"), "title": model,
            "status": "PASS" if ok else "MISMATCH", "summary": f"合同 {contract_period or '未提取'} ({contract_class}) / CQP {c_code or '未提取'} / TA {t_code or '未提取'}",
            "values": {"合同质保": contract_period or "未提取", "合同分类": contract_class, "CQP代码": c_code or "未提取", "CQP描述": c_detail.get("description") or "未提取", "TA代码": t_code or "未提取", "TA描述": t_detail.get("description") or "未提取"},
            "evidence": _valid_evidence([
                _evidence(contract_pdf, "contract", model + "合同质保", [contract_period, "5.2", "质保"]),
                _evidence(cqp_pdf, "cqp", model + " CQP质保", [c_code, c_detail.get("description", "")]),
                _evidence(ta_pdf, "ta", model + " TA质保", [t_code, t_detail.get("description", "")]),
            ]),
        })
    items.append(_item(
        "warranty_by_model", CATEGORY_PRODUCT, "按型号校对质保", "MISMATCH" if warranty_problem else "PASS", "blocker" if warranty_problem else "info",
        "合同5.2质保分类与CQP/TA质保配置一致。" if not warranty_problem else "合同5.2、CQP或TA的质保期限/分类/配置代码存在缺失或冲突。",
        {"合同5.2原文": contract_warranty.get("raw") or "未提取", "合同质保分类": contract_warranty.get("classification") or "Unknown"},
        [entry for sub in warranty_subitems for entry in sub["evidence"]], warranty_subitems,
    ))

    cqp_configs, ta_configs = cqp.get("configurations", []), ta.get("configurations", [])
    config_subitems, real_mismatches, ignored_codes = [], [], []
    for config in cqp_configs:
        model, code = config.get("model", ""), config.get("code", "")
        if not model or not code:
            continue
        if code in CQP_ONLY_ALLOWED_CODES:
            ignored_codes.append(code)
            continue
        matched, method = _config_semantic_match(config, ta_configs, _model_section_text(ta_pdf, model))
        if not matched:
            real_mismatches.append((model, code))
        config_subitems.append({
            "id": "config_" + re.sub(r"[^a-z0-9]+", "_", (model + "_" + code).lower()).strip("_"), "title": f"{model} · {code}",
            "status": "PASS" if matched else "MISMATCH", "summary": "TA中存在相同代码或可验证的等价描述。" if matched else "CQP配置在TA中未找到相同代码或可信等价描述。",
            "values": {"CQP描述": config.get("description") or "", "TA匹配方式": method},
            "evidence": _valid_evidence([
                _evidence(cqp_pdf, "cqp", f"{model} {code}", [code], page_hint=config.get("page")),
                _evidence(ta_pdf, "ta", f"{model} 对应配置", [code, *CONFIG_FEATURES.get(code, ())]),
            ]),
        })
    cqp_keys = {(item.get("model"), item.get("code")) for item in cqp_configs}
    ta_only = sorted({(item.get("model", ""), item.get("code", "")) for item in ta_configs if (item.get("model"), item.get("code")) not in cqp_keys and not str(item.get("code", "")).startswith("438-")})
    config_unavailable = not cqp_configs or not ta_configs
    config_status = "UNDETERMINED" if config_unavailable else ("MISMATCH" if real_mismatches else ("WARNING" if ta_only else "PASS"))
    config_severity = "blocker" if config_status in {"UNDETERMINED", "MISMATCH"} else ("warning" if ta_only else "info")
    items.append(_item(
        "configuration_consistency", CATEGORY_PRODUCT, "CQP与TA技术配置", config_status, config_severity,
        "CQP与TA技术配置一致。" if config_status == "PASS" else (
            "无法从CQP或TA中提取足够配置项，不能完成技术闭环。" if config_unavailable else
            "存在CQP配置未在TA中找到。" if real_mismatches else "TA存在CQP未列示的附加配置，需确认是否属于技术说明或范围变更。"
        ),
        {"允许仅存在于CQP": "、".join(sorted(set(ignored_codes))) or "无", "CQP缺失于TA": "、".join(f"{m}:{c}" for m, c in real_mismatches) or "无", "TA附加项": "、".join(f"{m}:{c}" for m, c in ta_only) or "无"},
        [entry for sub in config_subitems for entry in sub["evidence"]], config_subitems,
    ))

    naming_warning = bool(ta.get("lps_name_in_supply") and ta.get("lps_name_in_parameters") and ta.get("lps_name_in_supply") != ta.get("lps_name_in_parameters"))
    items.append(_item(
        "lps_lite_naming", CATEGORY_PRODUCT, "LPS / Lite+型号命名", "WARNING" if naming_warning else "PASS", "warning" if naming_warning else "info",
        "TA不同章节使用LPS与Lite+；系统已按等价别名映射，但建议确认正式型号名称。" if naming_warning else "未发现LPS/Lite+命名冲突。",
        {"供货范围": ta.get("lps_name_in_supply") or "未提取", "技术参数": ta.get("lps_name_in_parameters") or "未提取"},
        [_evidence(ta_pdf, "ta", "LPS/Lite+命名", ["LPS", "Lite+"])],
    ))

    repeatability = ta.get("repeatability_gen2", "")
    unit_warning = bool(repeatability and re.search(r"0[.]0*\d+\s*m(?!m)", repeatability, re.I))
    items.append(_item(
        "technical_parameter_units", CATEGORY_PRODUCT, "技术参数单位合理性", "WARNING" if unit_warning else "PASS", "warning" if unit_warning else "info",
        "重复定位精度使用m而非mm，疑似单位错误。" if unit_warning else "未发现明显技术参数单位异常。",
        {"重复定位精度": repeatability or "未提取"}, [_evidence(ta_pdf, "ta", "重复定位精度", [repeatability])],
    ))

    # ---------------- Other information ----------------
    for item_id, title, left_key, right_key in (
        ("untaxed_amount", "未税金额", "untaxed_amount", "untaxed_total"),
        ("tax_included_amount", "含税金额", "tax_included_amount", "tax_included_total"),
    ):
        left, right = float(contract.get(left_key, 0) or 0), float(cqp.get(right_key, 0) or 0)
        status, severity, diff = _amount_status(left, right)
        if status == "PASS":
            summary = "合同与CQP金额一致。"
        elif status == "WARNING":
            summary = f"合同与CQP相差人民币{diff:.2f}元，小于1元，按舍入差异处理。"
        elif status == "UNDETERMINED":
            summary = "合同或CQP金额未提取，无法核对。"
        else:
            summary = f"合同与CQP相差人民币{diff:.2f}元，超过允许的舍入范围。"
        items.append(_item(
            item_id, CATEGORY_OTHER, title, status, severity, summary,
            {"合同": f"¥{left:,.2f}", "CQP": f"¥{right:,.2f}", "差异": f"¥{diff:,.2f}"},
            [_evidence(contract_pdf, "contract", "合同" + title, [f"{left:,.2f}", str(left)]), _evidence(cqp_pdf, "cqp", "CQP" + title, [f"{right:,.2f}", str(right)])],
        ))

    contract_vat, cqp_vat = float(contract.get("vat_rate", 0) or 0), float(cqp.get("vat_rate", 0) or 0)
    vat_ok = abs(contract_vat - EXPECTED_CN_ROBOT_VAT) < 0.0001 and abs(cqp_vat - EXPECTED_CN_ROBOT_VAT) < 0.0001
    items.append(_item(
        "vat_rate", CATEGORY_OTHER, "增值税率与税额", "PASS" if vat_ok else "MISMATCH", "info" if vat_ok else "blocker",
        "合同与CQP VAT均为13%。" if vat_ok else "VAT缺失、合同与CQP不一致，或不是中国机器人裸机合同的13%标准税率。",
        {"合同VAT": f"{contract_vat*100:.2f}%", "CQP VAT": f"{cqp_vat*100:.2f}%", "CQP税额": f"¥{float(cqp.get('tax_amount', 0) or 0):,.2f}"},
        [_evidence(contract_pdf, "contract", "合同VAT", [f"{contract_vat*100:g}%"]), _evidence(cqp_pdf, "cqp", "CQP VAT", [f"{cqp_vat*100:g}%"])],
    ))

    arithmetic_checks = []
    for source, untaxed, vat, gross in (
        ("合同", float(contract.get("untaxed_amount", 0) or 0), contract_vat, float(contract.get("tax_included_amount", 0) or 0)),
        ("CQP", float(cqp.get("untaxed_total", 0) or 0), cqp_vat, float(cqp.get("tax_included_total", 0) or 0)),
    ):
        expected = round(untaxed * (1 + vat), 2) if untaxed and vat else 0
        diff = round(abs(expected - gross), 2) if gross else 0
        arithmetic_checks.append({"source": source, "expected": expected, "gross": gross, "difference": diff, "ok": bool(expected and gross and diff < AMOUNT_ROUNDING_TOLERANCE)})
    arithmetic_ok = all(entry["ok"] for entry in arithmetic_checks)
    items.append(_item(
        "tax_arithmetic", CATEGORY_OTHER, "税前税后金额计算", "PASS" if arithmetic_ok else "MISMATCH", "info" if arithmetic_ok else "blocker",
        "合同和CQP的未税金额、VAT与含税金额计算闭环。" if arithmetic_ok else "未税金额、VAT与含税金额之间存在超过1元的计算差异或关键金额缺失。",
        {entry["source"]: f"应为¥{entry['expected']:,.2f} / 文件¥{entry['gross']:,.2f} / 差¥{entry['difference']:,.2f}" for entry in arithmetic_checks}, [],
    ))

    rounding_issues = [entry for entry in cqp.get("line_rounding", []) if 0 < abs(entry.get("difference", 0)) < AMOUNT_ROUNDING_TOLERANCE]
    large_line_issues = [entry for entry in cqp.get("line_rounding", []) if abs(entry.get("difference", 0)) >= AMOUNT_ROUNDING_TOLERANCE]
    line_status = "MISMATCH" if large_line_issues else ("WARNING" if rounding_issues else "PASS")
    items.append(_item(
        "cqp_line_rounding", CATEGORY_OTHER, "CQP单价与行总额舍入", line_status, "blocker" if large_line_issues else ("warning" if rounding_issues else "info"),
        "CQP行金额存在超过1元的计算差异。" if large_line_issues else ("CQP显示单价与行总额存在小于1元的隐藏精度/舍入差。" if rounding_issues else "CQP行金额计算一致。"),
        {entry["model"]: f"显示计算 {entry['shown_calculation']:,.2f} / 行总额 {entry['line_total']:,.2f}" for entry in rounding_issues + large_line_issues}, [],
    ))

    payment_status, payment_severity, payment_summary = _payment_terms_consistency(contract.get("payment_terms", {}), cqp.get("payment_terms", {}))
    items.append(_item(
        "payment_terms", CATEGORY_OTHER, "付款条件与开票", payment_status, payment_severity, payment_summary,
        {"合同附件二原文": contract.get("payment_terms", {}).get("raw") or "未提取", "CQP付款条件": cqp.get("payment_terms", {}).get("raw") or "未提取", "BT09填写规则": "原文照抄合同附件二，不得简写比例或省略备注", "收款账户": contract.get("payment_terms", {}).get("account_number") or "未提取"},
        [_evidence(contract_pdf, "contract", "合同付款条件", [contract.get("payment_terms", {}).get("raw", ""), "付款条件"]), _evidence(cqp_pdf, "cqp", "CQP付款条件", [cqp.get("payment_terms", {}).get("raw", ""), "付款条件"])],
    ))

    schedule, q_weeks = contract.get("delivery_schedule", []), int(cqp.get("delivery_weeks", 0) or 0)
    contract_weeks = sorted({int(entry.get("weeks", 0)) for entry in schedule if entry.get("weeks")})
    if not contract_weeks:
        delivery_status, delivery_severity, delivery_summary = "UNDETERMINED", "blocker", "合同交期无法提取，BT09关键字段无法确认。"
    elif not q_weeks:
        delivery_status, delivery_severity, delivery_summary = "WARNING", "warning", "CQP交期未提取；BT09按合同交期填写。"
    elif any(week != q_weeks for week in contract_weeks) or (contract.get("delivery_trigger") and cqp.get("delivery_trigger") and contract.get("delivery_trigger") != cqp.get("delivery_trigger")):
        delivery_status, delivery_severity, delivery_summary = "WARNING", "warning", "合同与CQP交期或起算条件不同；这是非阻塞差异，BT09以合同为最终依据。"
    else:
        delivery_status, delivery_severity, delivery_summary = "PASS", "info", "合同与CQP交期一致；BT09仍以合同为准。"
    items.append(_item(
        "delivery_period", CATEGORY_OTHER, "交付周期与起算条件", delivery_status, delivery_severity, delivery_summary,
        {"合同": "；".join(f"{entry.get('model')} {entry.get('weeks')}周" for entry in schedule) or "未提取", "合同起算": contract.get("delivery_trigger") or "未提取", "CQP": cqp.get("delivery_time") or "未提取", "CQP起算": cqp.get("delivery_trigger") or "未提取", "BT09来源": "合同"},
        _evidence_many(contract_pdf, "contract", "合同交付周期", [str(week) + "周" for week in contract_weeks] + [contract.get("delivery_trigger", "")]) + _evidence_many(cqp_pdf, "cqp", "CQP交付周期", [cqp.get("delivery_time", "")]),
    ))

    incoterm = _infer_incoterm(contract, cqp)
    items.append(_item(
        "incoterm_delivery_place", CATEGORY_OTHER, "贸易术语、交付地点与Ship-to", incoterm["status"], incoterm["severity"], incoterm["reason"],
        {"结论": incoterm["conclusion"], "合同勾选": incoterm["contract_selected"] or "未识别", "合同交付地点": incoterm["delivery_location"] or "未提取", "CQP": cqp.get("delivery_term") or "未提取", "BT09 Incoterm 2": incoterm["incoterm_2"] or "待确认", "BT09 Ship-to名称": incoterm["ship_to_name"] or "待确认", "BT09 Ship-to地址": incoterm["ship_to_address"] or "待确认", "规则": incoterm["ship_to_rule"]},
        [_evidence(contract_pdf, "contract", "合同贸易术语选项", [line.get("text", "") for line in contract.get("incoterm_detection", {}).get("lines", [])]), _evidence(contract_pdf, "contract", "合同交付地点", [contract.get("delivery_location", "")]), _evidence(cqp_pdf, "cqp", "CQP贸易术语", [cqp.get("delivery_term", "")])],
    ))

    signature_placeholders = sorted(set(contract.get("signature_placeholders", []) + ta.get("signature_placeholders", [])))
    signature_issue = bool(signature_placeholders or contract.get("blank_signature_dates") or ta.get("blank_signature_dates"))
    items.append(_item(
        "signature_completeness", CATEGORY_OTHER, "签字、盖章与日期完整性", "WARNING" if signature_issue else "PASS", "warning" if signature_issue else "info",
        "文件仍含签字/盖章占位符或空日期；草稿阶段可接受，最终签署版必须补全。" if signature_issue else "签署信息完整。",
        {"占位符": "、".join(signature_placeholders) or "无", "合同日期空白": bool(contract.get("blank_signature_dates")), "TA日期空白": bool(ta.get("blank_signature_dates"))}, [],
    ))

    attachments = contract.get("attachments", {})
    missing_critical = []
    if not ta_has_content: missing_critical.append("技术协议")
    if not attachments.get("payment") or not contract.get("payment_terms", {}).get("raw"): missing_critical.append("附件二付款条件")
    integrity_missing = not attachments.get("integrity")
    attachment_status = "MISMATCH" if missing_critical else ("WARNING" if integrity_missing else "PASS")
    attachment_severity = "blocker" if missing_critical else ("warning" if integrity_missing else "info")
    items.append(_item(
        "attachment_completeness", CATEGORY_OTHER, "附件完整性与文件优先级", attachment_status, attachment_severity,
        "关键附件内容完整。" if attachment_status == "PASS" else ("缺少关键附件内容：" + "、".join(missing_critical) if missing_critical else "诚信条款附件未识别，需人工确认。"),
        {"技术协议内容": ta_has_content, "付款条件内容": bool(contract.get("payment_terms", {}).get("raw")), "诚信条款": bool(attachments.get("integrity")), "优先级": contract.get("file_priority") or "未提取"}, [],
    ))

    responsibilities = ta.get("responsibilities", {})
    responsibility_ok = bool(responsibilities) and all(responsibilities.values())
    items.append(_item(
        "scope_responsibility", CATEGORY_OTHER, "供货范围与责任边界", "PASS" if responsibility_ok else "UNDETERMINED", "info" if responsibility_ok else "warning",
        "TA已提取买卖双方的系统集成、安装与供货责任边界。" if responsibility_ok else "责任边界未完整提取，需人工确认TA相关条款。",
        responsibilities, [],
    ))

    return items


# ---------------------------------------------------------------------------
# Legacy compatibility and orchestration
# ---------------------------------------------------------------------------


def _legacy_check(item: Dict[str, Any]) -> Dict[str, Any]:
    return {
        "check_name": item.get("title", ""),
        "status": item.get("status", ""),
        "detail": item.get("summary", ""),
        "is_blocker": item.get("severity") == "blocker" and item.get("status") != "PASS",
    }


def _serialize_extracted(data: Dict[str, Any]) -> Dict[str, Any]:
    return data


def _run_optional_llm(
    contract: Dict[str, Any],
    cqp: Dict[str, Any],
    ta: Dict[str, Any],
    items: List[Dict[str, Any]],
) -> Dict[str, Any]:
    legacy = [_legacy_check(item) for item in items]
    incoterm_item = next((item for item in items if item["id"] == "incoterm_delivery_place"), {})
    incoterm_result = {
        "conclusion": incoterm_item.get("values", {}).get("结论", "UNDETERMINED"),
        "contract_evidence": incoterm_item.get("values", {}).get("合同勾选", ""),
        "cqp_evidence": incoterm_item.get("values", {}).get("CQP", ""),
        "consistent": incoterm_item.get("status") == "PASS",
        "status": incoterm_item.get("status", ""),
        "reason": incoterm_item.get("summary", ""),
    }
    warranty_item = next((item for item in items if item["id"] == "warranty_by_model"), {})
    config_item = next((item for item in items if item["id"] == "configuration_consistency"), {})
    financial = {
        "vat_check": next((item for item in items if item["id"] == "vat_rate"), {}),
        "untaxed_check": next((item for item in items if item["id"] == "untaxed_amount"), {}),
        "tax_included_check": next((item for item in items if item["id"] == "tax_included_amount"), {}),
    }
    try:
        return run_llm_contract_review(
            contract_data=contract,
            cqp_data=cqp,
            ta_data=ta,
            incoterm_result=incoterm_result,
            consistency_results=legacy,
            warranty_result={
                "consistent": warranty_item.get("status") == "PASS",
                "detail": warranty_item.get("summary", ""),
                "contract_warranty": contract.get("warranty", {}),
                "cqp_warranty_codes": cqp.get("warranty_codes_by_model", {}),
                "ta_warranty_codes": ta.get("warranty_codes_by_model", {}),
            },
            config_result={
                "overall_consistent": config_item.get("status") == "PASS",
                "models_compared": config_item.get("sub_items", []),
            },
            financial_result=financial,
        )
    except Exception as exc:
        return {"error": str(exc), "overall_assessment": "Unknown", "summary": "AI审核不可用，规则检查结果不受影响。"}



def _header_key(value: Any) -> str:
    return _compact(str(value or ""))


def _load_customer_master_rows(path: str) -> List[Dict[str, Any]]:
    extension = os.path.splitext(path)[1].lower()
    rows: List[Dict[str, Any]] = []
    if extension in {".xlsx", ".xlsm"}:
        from openpyxl import load_workbook
        workbook = load_workbook(path, read_only=True, data_only=True)
        for sheet in workbook.worksheets:
            iterator = sheet.iter_rows(values_only=True)
            headers = next(iterator, None)
            if not headers:
                continue
            header_names = [str(value or "").strip() for value in headers]
            for values in iterator:
                row = {header_names[index]: values[index] for index in range(min(len(header_names), len(values))) if header_names[index]}
                if any(value not in (None, "") for value in row.values()):
                    row["__sheet__"] = sheet.title
                    rows.append(row)
        workbook.close()
        return rows
    if extension == ".csv":
        with open(path, "r", encoding="utf-8-sig", newline="") as handle:
            for row in csv.DictReader(handle):
                if any(value not in (None, "") for value in row.values()):
                    rows.append(dict(row))
        return rows
    raise ValueError("客户主数据仅支持 .xlsx、.xlsm 或 .csv。")


def _row_values_by_headers(row: Dict[str, Any], aliases: Sequence[str]) -> List[str]:
    alias_keys = [_header_key(alias) for alias in aliases]
    values: List[str] = []
    for header, value in row.items():
        if str(header).startswith("__") or value in (None, ""):
            continue
        key = _header_key(header)
        if any(alias == key or alias in key for alias in alias_keys):
            values.append(_clean_inline(str(value)))
    return values


def _match_customer_master(path: Optional[str], contract: Dict[str, Any], incoterm: Dict[str, Any]) -> Dict[str, Any]:
    if not path:
        return {"status": "not_configured", "available": False, "matched": False, "note": "未提供客户主数据表。"}
    if not os.path.exists(path):
        return {"status": "error", "available": False, "matched": False, "note": f"客户主数据表不存在：{path}"}
    try:
        rows = _load_customer_master_rows(path)
    except Exception as exc:
        return {"status": "error", "available": False, "matched": False, "note": f"客户主数据读取失败：{exc}"}

    target_names = [
        contract.get("buyer_name", ""),
        contract.get("end_customer_name", ""),
        incoterm.get("ship_to_name", ""),
    ]
    target_addresses = [
        contract.get("buyer_address", ""),
        contract.get("end_customer_address", ""),
        incoterm.get("ship_to_address", ""),
    ]
    name_aliases = ("客户名称", "买方名称", "最终用户名称", "收货方名称", "ship-to name", "ship to name", "customer name", "end customer")
    address_aliases = ("客户地址", "买方地址", "最终用户地址", "收货地址", "ship-to address", "ship to address", "customer address")
    ship_to_aliases = ("ship-to id", "ship to id", "shiptoid", "收货方id", "收货客户id")
    end_customer_aliases = ("end customer id", "endcustomerid", "最终用户id", "终端客户id")
    gis_aliases = ("gis号", "gis number", "gis no", "gis")

    best: Optional[Tuple[int, Dict[str, Any]]] = None
    for row in rows:
        row_names = _row_values_by_headers(row, name_aliases)
        row_addresses = _row_values_by_headers(row, address_aliases)
        score = 0
        for target in target_names:
            if target and any(_contains_equivalent(target, candidate) for candidate in row_names):
                score += 4
                break
        for target in target_addresses:
            if target and any(_address_match(target, candidate) for candidate in row_addresses):
                score += 3
                break
        if best is None or score > best[0]:
            best = (score, row)
    if not best or best[0] < 4:
        return {"status": "not_found", "available": True, "matched": False, "note": "客户主数据中未找到可信匹配，ID/GIS保持为空。", "rows_checked": len(rows)}

    row = best[1]
    def first_value(aliases: Sequence[str]) -> str:
        values = _row_values_by_headers(row, aliases)
        return values[0] if values else ""
    return {
        "status": "matched",
        "available": True,
        "matched": True,
        "score": best[0],
        "sheet": row.get("__sheet__", ""),
        "ship_to_id": first_value(ship_to_aliases),
        "end_customer_id": first_value(end_customer_aliases),
        "gis_number": first_value(gis_aliases),
        "note": "客户主数据已按名称/地址匹配；空字段仍需人工补充。",
    }


def _customer_master_review_item(result: Dict[str, Any]) -> Dict[str, Any]:
    status = result.get("status")
    if status == "matched":
        item_status, severity, summary = "PASS", "info", "客户主数据已匹配，已提取可用的Ship-to ID、End Customer ID和GIS。"
    elif status == "not_configured":
        item_status, severity, summary = "INFO", "info", "未提供客户主数据表；相关ID保持为空且不影响合同审核结论。"
    else:
        item_status, severity, summary = "WARNING", "warning", result.get("note") or "客户主数据未匹配，需人工补充。"
    return _item(
        "customer_master_ids", CATEGORY_CUSTOMER, "客户ID与GIS匹配", item_status, severity, summary,
        {
            "Ship-to ID": result.get("ship_to_id") or "未匹配",
            "End Customer ID": result.get("end_customer_id") or "未匹配",
            "GIS": result.get("gis_number") or "未匹配",
            "匹配说明": result.get("note") or "",
        }, [],
    )

def _build_bt09_draft(
    contract: Dict[str, Any],
    cqp: Dict[str, Any],
    ta: Dict[str, Any],
    items: Sequence[Dict[str, Any]],
    customer_master: Optional[Dict[str, Any]] = None,
) -> Dict[str, Any]:
    del ta  # Reserved for future template fields; technical closure is represented in review items.
    customer_master = customer_master or {}
    incoterm_item = next((item for item in items if item.get("id") == "incoterm_delivery_place"), {})
    delivery_item = next((item for item in items if item.get("id") == "delivery_period"), {})
    warranty_item = next((item for item in items if item.get("id") == "warranty_by_model"), {})
    blockers = [item.get("id") for item in items if item.get("severity") == "blocker" and item.get("status") != "PASS"]
    values = incoterm_item.get("values", {})
    incoterm = values.get("结论", "UNDETERMINED")
    template_type = f"BT09_{incoterm}" if incoterm in {"DDP", "EXW"} else ""
    return {
        "ready": not blockers,
        "blocked_by": blockers,
        "template_type": template_type,
        "contract_number": contract.get("contract_number", ""),
        "cqp_number": cqp.get("cqp_number", ""),
        "buyer_name": contract.get("buyer_name", ""),
        "buyer_address": contract.get("buyer_address", ""),
        "end_customer_name": contract.get("end_customer_name", ""),
        "end_customer_address": contract.get("end_customer_address", ""),
        "incoterm": incoterm,
        "incoterm_2": values.get("BT09 Incoterm 2", ""),
        "ship_to_name": values.get("BT09 Ship-to名称", ""),
        "ship_to_address": values.get("BT09 Ship-to地址", ""),
        "delivery_terms": delivery_item.get("values", {}).get("合同", ""),
        "delivery_trigger": contract.get("delivery_trigger", ""),
        "payment_terms_verbatim": contract.get("payment_terms", {}).get("raw", ""),
        "warranty_summary": warranty_item.get("summary", ""),
        "vat_rate": contract.get("vat_rate", 0),
        "untaxed_amount": contract.get("untaxed_amount", 0),
        "tax_included_amount": contract.get("tax_included_amount", 0),
        "sales_person": contract.get("sales_person") or cqp.get("sales_person", ""),
        "pm": contract.get("pm", ""),
        "ship_to_id": customer_master.get("ship_to_id", ""),
        "end_customer_id": customer_master.get("end_customer_id", ""),
        "gis_number": customer_master.get("gis_number", ""),
        "customer_master_note": customer_master.get("note", "未提供客户主数据表时保持为空，不得编造。"),
    }


def _format_bt09_draft(fields: Dict[str, Any]) -> str:
    """Render BT09 fields as plain text for the legacy browser panel."""
    lines = [
        f"模板：{fields.get('template_type') or '待确认'}",
        f"合同号：{fields.get('contract_number') or '待确认'}",
        f"CQP号：{fields.get('cqp_number') or '待确认'}",
        f"Incoterm：{fields.get('incoterm') or '待确认'}",
        f"Incoterm 2：{fields.get('incoterm_2') or '待确认'}",
        f"Ship-to：{fields.get('ship_to_name') or '待确认'}",
        f"Ship-to地址：{fields.get('ship_to_address') or '待确认'}",
        f"Ship-to ID：{fields.get('ship_to_id') or '待补充'}",
        f"End Customer ID：{fields.get('end_customer_id') or '待补充'}",
        f"GIS：{fields.get('gis_number') or '待补充'}",
        f"交期：{fields.get('delivery_terms') or '待确认'}",
        f"交期起算：{fields.get('delivery_trigger') or '待确认'}",
        "付款条件（合同附件二原文）：",
        fields.get("payment_terms_verbatim") or "待确认",
        f"质保：{fields.get('warranty_summary') or '待确认'}",
    ]
    blocked_by = fields.get("blocked_by") or []
    if blocked_by:
        lines.extend(("", "当前不可发起BT09，阻塞项：" + "、".join(str(item) for item in blocked_by)))
    elif fields.get("customer_master_note"):
        lines.extend(("", "客户主数据：" + str(fields.get("customer_master_note"))))
    return "\n".join(lines)


def _legacy_incoterm_payload(item: Dict[str, Any]) -> Dict[str, Any]:
    values = item.get("values", {})
    return {
        "conclusion": values.get("结论", "UNDETERMINED"),
        "contract_evidence": "；".join(
            value for value in (
                f"勾选={values.get('合同勾选')}" if values.get("合同勾选") else "",
                f"交付地点={values.get('合同交付地点')}" if values.get("合同交付地点") else "",
            ) if value
        ),
        "cqp_evidence": values.get("CQP", ""),
        "consistent": item.get("status") == "PASS",
        "status": item.get("status", ""),
        "detail": item.get("summary", ""),
    }


def _legacy_warranty_payload(item: Dict[str, Any], cqp: Dict[str, Any]) -> Dict[str, Any]:
    return {
        "consistent": item.get("status") == "PASS",
        "detail": item.get("summary", ""),
        "cqp_warranty_codes": sorted(set(cqp.get("warranty_codes_by_model", {}).values())),
        "status": item.get("status", ""),
    }


def _legacy_financial_payload(items: Sequence[Dict[str, Any]], contract: Dict[str, Any], cqp: Dict[str, Any]) -> Dict[str, Any]:
    by_id = {item.get("id"): item for item in items}
    untaxed = by_id.get("untaxed_amount", {})
    taxed = by_id.get("tax_included_amount", {})
    vat = by_id.get("vat_rate", {})

    def amount_payload(item: Dict[str, Any]) -> Dict[str, Any]:
        raw_diff = item.get("values", {}).get("差异", "")
        match = re.search(r"([\d,]+(?:[.]\d+)?)", str(raw_diff))
        diff = float(match.group(1).replace(",", "")) if match else 0.0
        return {
            "status": item.get("status", ""),
            "diff": diff,
            "is_rounding": item.get("status") == "WARNING" and diff < AMOUNT_ROUNDING_TOLERANCE,
            "detail": item.get("summary", ""),
        }

    return {
        "vat_check": {
            "status": vat.get("status", ""),
            "contract_vat": float(contract.get("vat_rate", 0) or 0),
            "cqp_vat": float(cqp.get("vat_rate", 0) or 0),
            "detail": vat.get("summary", ""),
        },
        "untaxed_check": amount_payload(untaxed),
        "tax_included_check": amount_payload(taxed),
    }


def run_review(
    pdf_paths: List[str],
    customer_db_path: str = None,
    template_path: str = None,
    file_roles: Optional[Dict[str, str]] = None,
) -> Dict[str, Any]:
    del template_path  # Template selection is exposed through bt09_draft.template_type.
    customer_db_path = customer_db_path or os.environ.get("PM_CUSTOMER_MASTER_PATH", "")
    parsed_by_path = {os.path.abspath(path): parse_pdf(path) for path in pdf_paths}
    documents = _resolve_documents(parsed_by_path, file_roles)
    if documents.contract_physical is None or documents.cqp is None:
        raise ValueError("Contract 和 CQP 都必须上传且能够读取。")

    contract = extract_contract(documents.contract_body)
    cqp = extract_cqp(documents.cqp)
    ta = extract_ta(documents.ta)
    review_items = build_review_items(documents, contract, cqp, ta)
    incoterm_data = _infer_incoterm(contract, cqp)
    customer_master = _match_customer_master(customer_db_path, contract, incoterm_data)
    review_items.append(_customer_master_review_item(customer_master))
    key_checks = [_legacy_check(item) for item in review_items]
    blockers = [
        {"type": item["id"], "detail": item["summary"]}
        for item in review_items
        if item.get("severity") == "blocker" and item.get("status") != "PASS"
    ]
    non_blockers = [
        {"type": item["id"], "detail": item["summary"]}
        for item in review_items
        if item.get("severity") == "warning" and item.get("status") != "PASS"
    ]
    conclusion = "Blocked" if blockers else ("Pass with notes" if non_blockers else "Pass")
    llm_review = _run_optional_llm(contract, cqp, ta, review_items)
    incoterm_item = next((item for item in review_items if item["id"] == "incoterm_delivery_place"), {})
    warranty_item = next((item for item in review_items if item["id"] == "warranty_by_model"), {})
    configuration_item = next((item for item in review_items if item["id"] == "configuration_consistency"), {})
    bt09_fields = _build_bt09_draft(contract, cqp, ta, review_items, customer_master)

    return {
        "conclusion": conclusion,
        "review_categories": [
            {"id": CATEGORY_CUSTOMER, "title": CATEGORY_LABELS[CATEGORY_CUSTOMER], "order": 1},
            {"id": CATEGORY_PRODUCT, "title": CATEGORY_LABELS[CATEGORY_PRODUCT], "order": 2},
            {"id": CATEGORY_OTHER, "title": CATEGORY_LABELS[CATEGORY_OTHER], "order": 3},
        ],
        "document_sources": {
            "contract": {"physical_role": "contract", "embedded": False},
            "cqp": {"physical_role": "cqp", "embedded": False},
            "ta": {"physical_role": "contract" if documents.ta_embedded else "ta", "embedded": documents.ta_embedded},
        },
        "source_recognition": {
            "contract": {"status": "found", "page_count": len(documents.contract_physical.pages)},
            "cqp": {"status": "found", "page_count": len(documents.cqp.pages)},
            "ta": {"status": "embedded" if documents.ta_embedded else ("found" if documents.ta else "not_found"), "page_count": len(documents.ta.pages) if documents.ta else 0},
        },
        "extracted_data": {
            "contract": _serialize_extracted(contract),
            "cqp": _serialize_extracted(cqp),
            "ta": _serialize_extracted(ta),
        },
        "customer_master": customer_master,
        "key_checks": key_checks,
        "blockers": blockers,
        "non_blockers": non_blockers,
        "review_items": review_items,
        "llm_review": llm_review,
        # Legacy response shapes retained for the existing browser UI.
        "incoterm": _legacy_incoterm_payload(incoterm_item),
        "warranty": _legacy_warranty_payload(warranty_item, cqp),
        "configuration": configuration_item,
        "financial": _legacy_financial_payload(review_items, contract, cqp),
        "bt09_fields": bt09_fields,
        "bt09_draft": _format_bt09_draft(bt09_fields),
    }
